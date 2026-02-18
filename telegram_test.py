from __future__ import annotations
import logging
from types import CoroutineType
from telegram import Update, Message, Chat, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ApplicationBuilder, CallbackContext, CommandHandler, MessageHandler, ContextTypes, CallbackQueryHandler
from telegram.ext import filters
from telegram_settings_local import TOKEN
from telegram_settings_local import FRIENDS_USER
from telegram_settings_local import SPECIAL_ENTITIES

from typing import Any, Callable, Awaitable, Tuple, Union, Iterable, Literal, TypedDict, NamedTuple, Optional

import json
from dataclasses import dataclass

import enum
import html
class FriendsUser(enum.StrEnum):
    FLOCON = 'flocon'
    KOROLEVA_LION = 'koroleva-lion'
    LOUKOUM = 'loukoum'
    JOKÈRE = 'jokère'
    SHOKO = 'shoko'
    QSNAKES = 'QSNAKES'.lower()
    KERRICYBERGOOSE = 'KERRICYBERGOOSE'.lower()
    DANCING_UNICORN = 'DANCING_UNICORN'.replace('_', '-').lower()
    KARL = "Karl"
    BIRD_FLOCK_MASTER = "bird-flock-master"

class SpecialUsers(enum.StrEnum):
    CRAZY_JAM = 'CRAZY_JAM'.lower().replace('_', '-')  # Crazy Jam Channel
    CRAZY_JAM_BACKEND = 'CRAZY_JAM_BACKEND'.lower().replace('_', '-') # Utka Banda
    CRAZY_JAM_BACKEND_THREAD_IN = 'CRAZY_JAM_BACKEND_THREAD_IN'.lower().replace('_', '-')

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

logging.getLogger('httpx').setLevel(logging.WARN)

async def start(update: Update, context: CallbackContext):
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text="I'm a bot, please talk to me! To get a tour of functionalities, send a message to my creator t.me/robertvend")
    print("Someone started me!")


async def ids(update:Update, context):
    send = make_send(update, context)

    await send(str(dict(
        user_id=update.effective_user.id,
        chat_id=update.effective_chat.id,
        thread_id=update.effective_message.message_thread_id)))

from telegram.ext.filters import MessageFilter

class CrazyJamFilter(MessageFilter):
    def filter(self, message: Message):
        return message.chat.id == SPECIAL_ENTITIES[SpecialUsers.CRAZY_JAM]

class CrazyJamFwdError(Exception):
    pass

class DayOfWeekFilter(MessageFilter):
    def filter(self, message: Message):
        try:
            read_chat_settings = make_read_chat_settings_from_chat_id(message.chat.id)
            return do_if_setting_on(read_chat_settings('event.commands.dayofweek'))
        except Exception as error: 
            logging.error("Error", exc_info=error)
            return False

async def on_crazy_jam_message(update: Update, context):
    if not update.message:
        return
    try:
        original_message_id = update.effective_message.id
        sent_message = await update.message.forward(SPECIAL_ENTITIES[SpecialUsers.CRAZY_JAM_BACKEND], message_thread_id=SPECIAL_ENTITIES[SpecialUsers.CRAZY_JAM_BACKEND_THREAD_IN])
        new_message_id = sent_message.id

        simple_sql(('''insert into FwdRelation(original_message_id, fwd_message_id, original_chat_id, original_chat_username) VALUES (?,?,?,?)''', (original_message_id, new_message_id, update.effective_chat.id, update.effective_chat.username)))

        # silent bot in the main channel
    except Exception as e:
        raise CrazyJamFwdError(str(e))

import re

import regex 

import funcoperators

@funcoperators.infix
def fullmatches(string, reg):
    import regex
    return regex.compile(reg).fullmatch(string) 

def fullmatches_with_flags(flags):
    @funcoperators.infix 
    def fullmatches_with_flags__inner(string, reg):
        return regex.compile(reg, flags).fullmatch(string)
    return fullmatches_with_flags__inner

fullmatchesI = fullmatches_with_flags(re.I)

from functools import partial
MONEY_CURRENCIES_ALIAS = {
    "eur": "eur",
    "euro": "eur",
    "euros": "eur",
    "€": "eur",
    "₺": "try",
    "try": "try",
    "tl": "try",
    "lira": "try",
    "brl": "brl",
    "real": "brl",
    "reais": "brl",
    "rub": "rub",
    "руб": "rub",
    "₽": "rub",
    '$': 'usd',
    'usd': 'usd',
    'cad': 'cad',
    'sol': 'pen',
}
MONEY_RE = re.compile('(\\d+[.]?\\d*) ?(' + '|'.join(map(re.escape, MONEY_CURRENCIES_ALIAS)) + ')', re.I)

def read_pure_json(filename):
    import json 
    with open(filename, encoding='utf-8') as f:
        return json.load(f)

WIKTIONARY_LANGUAGES = read_pure_json('wiktionary_languages.json')
LAROUSSE_LANGUAGES = read_pure_json('larousse_languages.json')
DEFAULT_CURRENCIES = ['eur', 'usd', 'rub', 'brl', 'cad']

EVENT_ICS_TEMPLATE = '''\
BEGIN:VCALENDAR
VERSION:2.0
PRODID:-//hacksw/handcal//NONSGML v1.0//EN
BEGIN:VEVENT
DTSTAMP:{dt_created_utc:%Y%m%dT%H%M%SZ}
DTSTART:{dt_start_utc:%Y%m%dT%H%M%SZ}
DTEND:{dt_end_utc:%Y%m%dT%H%M%SZ}
SUMMARY:{name_ical_formatted}
END:VEVENT
END:VCALENDAR\
''' # UID:uid1@example.com, GEO:48.85299;2.36885, ORGANIZER;CN=John Doe:MAILTO:john.doe@example.com

def get_reply(message):
    if not message.is_topic_message:
        return message.reply_to_message
    elif message.reply_to_message.id == message.message_thread_id:
        return None
    else:
        return message.reply_to_message

def update_get_reply(update: Update):
    return get_reply(update.effective_message)

def strip_botname(update: Update, context: CallbackContext):
    # TODO analyse message.entities with message.parse_entity and message.parse_entities
    bot_mention: str = '@' + context.bot.username
    if update.message.text.startswith(bot_mention):
        return update.message.text[len(bot_mention):].strip()
    return update.message.text.strip()

async def hello_responder(msg:str, send: AsyncSend, *, update, context):
    user = update.effective_user
    if user.id == FRIENDS_USER.get(FriendsUser.LOUKOUM):
        if msg.lower().startswith("hello"):
            await send("Hello my loukoum !")
        if all(word in msg.lower() for word in ('bebeğimin', 'botu')):
            await send("İyi günler Loukoum ! Çok tatlısın 🍬")
    elif user.id == FRIENDS_USER.get(FriendsUser.FLOCON):
        if msg.lower().startswith("hello"):
            await send("Bonjour flocon ! J'espère que ta journée sera artistique !")
    elif user.id == FRIENDS_USER.get(FriendsUser.JOKÈRE):
        if msg.lower().startswith("hello"):
            await send("Ʒokère ! Nous nous retrouvons ! Pas de spam en public !")
    elif user.id == FRIENDS_USER.get(FriendsUser.KOROLEVA_LION):
        if msg.lower().startswith("hello"):
            await send("Hellow you wild sladkij ^^ Hope your day will improve your life !")
    elif user.id == FRIENDS_USER.get(FriendsUser.SHOKO):
        if (english := msg.lower().startswith("hello")) or (russian := msg.lower().startswith("привет")):
            assert english or russian
            await send("Do you want some lezhunki ? Or some Tango ? It seems like you do !" if english else "Хочешь лежунки или Танго? Мне кстати Да !")
    elif user.id == FRIENDS_USER.get(FriendsUser.QSNAKES):
        if msg.lower().startswith("hello"):
            await send("Santa Claus and Heisenberg say... hello")
    elif user.id == FRIENDS_USER.get(FriendsUser.KERRICYBERGOOSE):
        if msg.lower().startswith("hello"):
            await send("Wanna protecc ?")
    elif user.id == FRIENDS_USER.get(FriendsUser.DANCING_UNICORN):
        if msg.lower().startswith("hello"):
            await send("¡Acepto!")
    elif user.id == FRIENDS_USER.get(FriendsUser.KARL):
        if msg.lower().startswith("hello"):
            await send("Hello, your Majesty!")
    elif user.id == FRIENDS_USER.get(FriendsUser.BIRD_FLOCK_MASTER):
        if msg.lower().startswith("hello"):
            birds = (
                'Moineau',
                'Geai de chênes',
                'Mésange à longue queue',
                'Albatross',
                'Calopsitte',
                'Perruche ondulée',
                'Corbeau',
                'Hirondelle',
                'Hibou',
                'Vautour',
                'Pygargue à tête blanche',
                'Chouette',
                'Inséparable',
                'Pigeon',
            )
            todays_bird = birds[(Datetime.now(UTC) - Datetime(2026, 2, 2, tzinfo=UTC)).days % len(birds)]

            await send(f"Bonjour Maître des Oiseaux! Connais-tu l'oiseau du jour (UTC)? J'ai nommé... {todays_bird} 🐦")
    else:
        if msg.lower().startswith('hello'):
            await send("Hello ! :3")
        elif msg == "Hi":
            await send("Yo")
        elif msg == "hi":
            await send("yë")

def detect_currencies(msg: str):
    return [(value, MONEY_CURRENCIES_ALIAS[currency_raw.lower()]) for value, currency_raw in MONEY_RE.findall(msg)]

def lazy_value(getter):
    sentinel = object()
    value = sentinel
    def f():
        nonlocal value
        if value is sentinel:
            value = getter()
        return value
    return f

AsyncSend = Callable[[Update, CallbackContext], Awaitable[None]]

async def money_responder(msg:str, send: AsyncSend, *, update, context):
    if detected_currencies := detect_currencies(msg):
        read_chat_settings = make_read_chat_settings(update, context)

        chat_currencies = remove_dup_keep_order(read_chat_settings('money.currencies') or DEFAULT_CURRENCIES)
        
        db_known_currencies = read_chat_settings('money.known_currencies')
        if db_known_currencies is not None:
            db_known_currencies = remove_dup_keep_order(db_known_currencies)

        rates = lazy_value(get_database_euro_rates)

        upper_detected_currencies = [(value, currency_lower.upper()) for value, currency_lower in detected_currencies]
        for value, currency in upper_detected_currencies:
            if db_known_currencies is None or currency not in db_known_currencies:
                if currency in chat_currencies:
                    currencies_to_convert = [x for x in chat_currencies if x != currency]
                    amount_base = Decimal(value)
                    amounts_converted = [convert_money(amount_base, currency_base=currency, currency_converted=currency_to_convert, rates=rates()) for currency_to_convert in currencies_to_convert]
                    await send(format_currency(currency_list=[currency] + currencies_to_convert, amount_list=[amount_base] + amounts_converted))

class DoNotAnswer(Exception):
    pass

async def locationdistance_responder(msg:str, send: AsyncSend, *, update, context):
    if match := msg /fullmatches_with_flags(re.I)/ 'now\s+[@]\s+(.*)':
        loc = match.group(1)
        dists = location_distance_apply(loc, chat_id=update.effective_chat.id).dists
        if len(dists) > 1:
            return await send('\n'.join(f"• {dist} | {name}" for name, dist in dists.items()))

def split_based_on_indices(L, indices):
    if len(indices) == 0:
        return [L[:]]
    O = []
    c = 0
    for b in indices:
        O.append(L[c:b])
        c = b+1
    O.append(L[c:])
    return O

def separate_based_on_indices(L, indices):
    if len(indices) == 0:
        return [L[:]]
    O = []
    c = 0
    for b in indices:
        O.append(L[c:b])
        c = b
    O.append(L[c:])
    return O

async def distfrom(update, context):
    send = make_send(update, context)

    if not context.args:
        return await send("Usage:\n/distfrom place\n/distfrom place to: destination")

    tos = [i for i, x in enumerate(context.args) if x.lower() in ('to:', ':to')]

    if not tos:
        targets = None
        loc = ' '.join(context.args)
    else:
        bits = split_based_on_indices(context.args, tos)
        loc = ' '.join(bits[0])
        targets = [' '.join(sub) for sub in bits[1:]]
        
    dists = (d := location_distance_apply(loc, chat_id=update.effective_chat.id, targets=targets)).dists

    if targets is None:
        display = dists.keys()
    else:
        display = list(map(str.lower, targets))

    return await send('\n'.join(f"• {dists[name]} | {name}" for name in display))


async def pathfrom(update, context):
    send = make_send(update, context)

    if not context.args:
        return await send("Usage:\n/pathfrom place\n/pathfrom place to: destination")

    tos = [i for i, x in enumerate(context.args) if x.lower() in ('to:', ':to')]

    if not tos:
        targets = None
        loc = ' '.join(context.args)
    else:
        bits = split_based_on_indices(context.args, tos)
        loc = ' '.join(bits[0])
        targets = [' '.join(sub) for sub in bits[1:]]

    assert_true(targets is None or len(targets) <= 1, UserError("Not implemented"))
        
    dijkstra = location_distance_apply(loc, chat_id=update.effective_chat.id, targets=targets)

    dists, prevs = dijkstra.dists, dijkstra.prevs

    if targets is None:
        return await send('\n'.join(f"• {dists[name]} | {name} (by {prevs.get(name)})" for name in dists))

    c = targets[0].lower()
    path = L = [c]
    while c in prevs:
        L.append(prevs[c])
        c = prevs[c]
    path.reverse()
        
    return await send('\n'.join(f"• {name} (at {dists[name]})" for name in path))

@dataclass
class DijkstraResult:
    dists: dict
    prevs: dict

def location_distance_apply(loc, *, chat_id, targets=None):
    targets = set(map(str.lower, targets)) if targets is not None else None
    loc = loc.lower()      

    edges = []
    with get_connection() as conn:
        my_simple_sql = partial(simple_sql_args, connection=conn)
        S = set()
        for graph_id, in my_simple_sql('select rowid from LocationDistanceGraph where chat_id = ?', (chat_id, )):
            if graph_id not in S:
                edges += my_simple_sql('select source, dest, distance from LocationDistanceEdge where graph_id = ?', (graph_id, ))
                S.add(graph_id)

        for graph_id, in my_simple_sql('select graph_id from LocationDistanceImportedGraph where chat_id = ?', (chat_id, )):
            if graph_id not in S:
                edges += my_simple_sql('select source, dest, distance from LocationDistanceEdge where graph_id = ?', (graph_id, ))
                S.add(graph_id)

    from collections import defaultdict
    Graph = defaultdict(list)
    for source, dest, distance in edges:
        Graph[source.lower()].append((dest.lower(), distance))
        Graph[dest.lower()].append((source.lower(), distance))

    open_list = {loc: 0}
    dists = {}
    prevs = {}
    while open_list:
        current_name, current_dist = min(open_list.items(), key=lambda t:t[1])
        del open_list[current_name]

        assert current_name not in dists, "Strange"
        dists[current_name] = current_dist

        if targets is not None:
            targets.discard(current_name)
            if not targets:
                break

        for neigh_name, neigh_dist in Graph[current_name]:
            if neigh_name not in dists:
                new_dist = current_dist + neigh_dist
                if neigh_name not in open_list or new_dist < open_list[neigh_name]:
                    open_list[neigh_name] = new_dist
                    prevs[neigh_name] = current_name
    
    return DijkstraResult(dists=dists, prevs=prevs)

class locationdistance:
    async def locationinfo(update, context):
        send = make_send(update, context)

        async def print_usage():
            return await send(
                "Usage:\n"
                "/graph addedge|listedges|listgraphs|switch|current|import|unimport|namespace"
            )
        
        Args = InfiniteEmptyList(context.args)
        
        if Args[0] /fullmatchesI/ r'addedges?':
            return await locationdistance.addedges(Args[1:], update, context)
        if Args[0] /fullmatchesI/ r'switch':
            return await locationdistance.switchgraph(Args[1:], update, context)
        if Args[0] /fullmatchesI/ r'current':
            return await send(f'Current graph: {locationdistance.get_current_graph(update.effective_chat.id)}')
        if Args[0] /fullmatchesI/ r'list|listedge|listedges':
            return await locationdistance.listlocationinfo(Args[1:], update, context)
        if Args[0] /fullmatchesI/ r'listgraph|listgraphs':
            return await locationdistance.listgraphs(Args[1:], update, context)
        if Args[0] /fullmatchesI/ r'import':
            return await locationdistance.importgraph(Args[1:], update, context)
        if Args[0] /fullmatchesI/ r'unimport':
            return await locationdistance.unimportgraph(Args[1:], update, context)
        if Args[0] /fullmatchesI/ r'namespace':
            return await locationdistance.graphnamespace(Args[1:], update, context)
        return await print_usage()
    
    async def addedges(args, update, context):
        send = make_send(update, context)

        def parse_edge(edge):
            if len(bits := edge.split()) == 3:
                source, dest, dist = bits
            elif len(bits := edge.split(' / ')) == 2:
                source, dest = bits
                B = InfiniteEmptyList(dest.split())
                if B[-1] /fullmatches/ '\d+' or B[-1] == 'delete':
                    dist = B[-1]
                    dest = ' '.join(B[:-1])
                else:
                    raise ValueError
            else:
                source, dest, dist = bits
            source, dest, dist = map(str.strip, (source, dest, dist))
            if dist.lower() == 'delete':
                dist = 'delete'
            else:
                dist = int(dist)
            return source, dest, dist
        
        def parse_semi_edge(edge):
            if not (bits := InfiniteEmptyList(edge.split())) and (bits[-1] /fullmatches/ '\d+'):
                raise ValueError
            *destL, dist = bits
            return ' '.join(destL), int(dist)

        edges = []
        try:
            for edge in ' '.join(args).split('//'):
                if m := edge /fullmatchesI/ '(path|star)\s*:\s*(.*)':
                    pattern, data = m.group(1), m.group(2)
                    data_list = InfiniteEmptyList(data.split('/'))
                    
                    prev = data_list[0]
                    if not prev:
                        raise ValueError

                    if pattern.lower() == 'star':
                        first = prev
                        for target in data_list[1:]:
                            dest, dist = parse_semi_edge(target)
                            edges.append((first, dest, dist))

                    elif pattern.lower() == 'path':
                        if len(data_list) == 1:
                            # /locationinfo path: Hello 5 World in house 10 Tada
                            S = prev.split()
                            breaks = []
                            for i, b in enumerate(S):
                                if b /fullmatches/ '\d+':
                                    breaks.append(i)
                            if not breaks:
                                raise ValueError
                            c = breaks[0]
                            prev = ' '.join(S[0:c])
                            for i in breaks[1:] + [len(S)]:
                                dist, *destL = S[c:i]
                                dest = ' '.join(destL)
                                assert_true(prev.strip() and dest.strip(), ValueError)
                                edges.append((prev.strip(), dest.strip(), dist))
                                prev = dest
                                c = i
                        else:
                            # /locationinfo path: Hello / World 5 / Tada 5
                            for target in data_list[1:]:
                                dest, dist = parse_semi_edge(target.strip())
                                edges.append((prev, dest, dist))
                                prev = dest
                else:
                    edges.append(parse_edge(edge.strip()))
            edges = [(a, b, int(c) if c != 'delete' else c) for a,b,c in edges]
        except ValueError as e:
            return await send(
                'Usage:\n'
                '/graph addedges from to distance\n'
                '/graph addedges from / to / distance\n'
                '/graph addedges from / to / distance // from / to / distance\n'
                '/graph addedges path: A dist B dist C\n'
                '/graph addedges star: A / B dist / C dist\n')

        chat_id = update.effective_chat.id
        with sqlite3.connect('db.sqlite') as conn:
            conn.execute('begin transaction')
            graph_id = locationdistance.get_current_graph_id(chat_id, conn)
            graph_chat_id, = only_one(conn.execute(''' select chat_id from LocationDistanceGraph where rowid = ? ''', (graph_id, )))
            graph_name = locationdistance.get_current_graph(chat_id, conn)
            
            if graph_chat_id != chat_id:
                return await send(f'You are not allowed to modify graph "{graph_name}" as it belongs to another chat')
            
            for source, dest, distance in edges:
                locationdistance.save_location_distance(chat_id, source, dest, distance, conn, graph_id)
            conn.execute('end transaction')

        return await send(f'[Graph "{graph_name}"] Edges modified: {len(edges)}')

    async def listlocationinfo(args, update, context):
        send = make_send(update, context)

        with sqlite3.connect('db.sqlite') as conn:
            my_simple_sql = partial(simple_sql_args, connection=conn)

            conn.execute('begin transaction')
            graph_id = locationdistance.get_current_graph_id(update.effective_chat.id, conn)
            graph_name = locationdistance.get_current_graph(update.effective_chat.id, conn)

            cursor = conn.cursor()
            cursor.execute('select source, dest, distance from LocationDistanceEdge where graph_id = ?', (graph_id, ))
            edges = cursor.fetchall()
            conn.execute('end transaction')

        return await send(f'[Graph "{graph_name}"]\n' + (' //\n'.join(f"{source} / {dest} / {distance}" for source, dest, distance in edges) or '/'))

    async def listgraphs(args, update, context):
        send = make_send(update, context)

        with get_connection() as conn:
            my_simple_sql = partial(simple_sql_args, connection=conn)

            conn.execute('begin transaction')
            
            chat_id = update.effective_chat.id

            out = []
            L = []
            L += my_simple_sql('select rowid from LocationDistanceGraph where chat_id=?', (chat_id, ))
            L += my_simple_sql('select graph_id from LocationDistanceImportedGraph where chat_id=?', (chat_id, ))

            for graph_id, in L:
                out.append("- {}".format(locationdistance.graph_full_name(chat_id=chat_id, graph_id=graph_id, connection=conn)))

            conn.execute('end transaction')
        
        def key(x):
            return ('.' in x, x)

        return await send('\n'.join(sorted(out, key=key)) or '/')

    def save_location_distance(chat_id, source, dest, distance, conn, graph_id):
        if distance == 'delete':
            conn.execute('delete from LocationDistanceEdge where chat_id = ? and graph_id = ? and (LOWER(source) = LOWER(?) and LOWER(dest) = LOWER(?) or LOWER(dest) = LOWER(?) and LOWER(source) = LOWER(?))', (chat_id, graph_id, source, dest, source, dest))
        elif conn.execute('select count(*) from LocationDistanceEdge where chat_id = ? and graph_id = ? and (LOWER(source) = LOWER(?) and LOWER(dest) = LOWER(?) or LOWER(dest) = LOWER(?) and LOWER(source) = LOWER(?))', (chat_id, graph_id, source, dest, source, dest)).fetchone()[0] == 0:
            conn.execute('insert into LocationDistanceEdge(chat_id, graph_id, source, dest, distance) values (?, ?, ?, ?, ?)', (chat_id, graph_id, source, dest, distance))
        else:
            conn.execute('update LocationDistanceEdge set distance = ? where chat_id = ? and graph_id = ? and (LOWER(source) = LOWER(?) and LOWER(dest) = LOWER(?) or LOWER(dest) = LOWER(?) and LOWER(source) = LOWER(?))', (distance, chat_id, graph_id, source, dest, source, dest))

    def graph_full_name(chat_id, graph_id, connection=None):
        my_simple_sql = partial(simple_sql_args, connection=connection)

        graph_name, graph_visibility, graph_chat_id = only_one(my_simple_sql(''' select name, visibility, chat_id from LocationDistanceGraph where rowid = ?''', (graph_id, )))
        
        return locationdistance.get_full_name_from_db_infos(chat_id=chat_id, graph_name=graph_name, graph_visibility=graph_visibility, graph_chat_id=graph_chat_id)        

    def get_current_graph(chat_id, connection=None):
        my_simple_sql = partial(simple_sql_args, connection=connection)
        alls = my_simple_sql(''' select g.name, g.visibility, g.chat_id from LocationDistanceCurrentGraphChat c inner join LocationDistanceGraph g on c.graph_id=g.rowid where c.chat_id=? ''', (chat_id, ))
        graph_name, graph_visibility, graph_chat_id = only_one(alls) if alls else ('chat', 'chat', chat_id)

        return locationdistance.get_full_name_from_db_infos(chat_id=chat_id, graph_name=graph_name, graph_visibility=graph_visibility, graph_chat_id=graph_chat_id)

    def get_full_name_from_db_infos(*, chat_id, graph_name, graph_visibility, graph_chat_id, connection=None):
        my_simple_sql = partial(simple_sql_args, connection=connection)

        if graph_chat_id == chat_id and graph_visibility == 'chat':
            namespace = None
        else:
            namespace, = only_one(my_simple_sql(''' select namespace from LocationDistanceGraphNamespace where chat_id = ?''', (graph_chat_id, )) or [(None, )])
            if namespace is None:
                namespace = '?'
        return (namespace + '.' if namespace else '') + graph_name + (' (readonly)' if graph_chat_id != chat_id else '')

    def get_current_graph_id(chat_id, connection=None):
        my_simple_sql = partial(simple_sql_args, connection=connection)
        my_simple_sql_create = partial(simple_sql_create_args, connection=connection)
        alls = my_simple_sql(''' select graph_id from LocationDistanceCurrentGraphChat where chat_id=? ''', (chat_id, ))
        graph_id, = only_one(alls) if alls else (None, )
        if graph_id is None:
            if r := my_simple_sql('''select rowid from LocationDistanceCurrentGraphChat where chat_id=? and visibility='chat' and name='chat' ''', (chat_id, )):
                graph_id, = only_one(r)
            else:
                r = my_simple_sql_create(''' insert into LocationDistanceCurrentGraphChat(chat_id, visibility, name) VALUES (?,?,?)''', (chat_id, 'chat', 'chat'))
                graph_id = r.lastrowid
            my_simple_sql('insert into LocationDistanceCurrentGraphChat(chat_id, graph_id) VALUES (?, ?)''', (chat_id, graph_id))
        return graph_id

    def get_graph_id_default_chat(chat_id, connection=None):
        my_simple_sql = partial(simple_sql_args, connection=connection)
        alls = my_simple_sql(''' select rowid from LocationDistanceGraph where c.chat_id=? AND visibility="chat" and name="chat" ''', (chat_id, ))
        g_infos = only_one(alls) if alls else (None, )
        return g_infos[0]

    async def switchgraph(args, update, context):
        send = make_send(update, context)

        async def print_usage(current_graph=False):
            return await send(
                "Usage:\n"
                "/graph switch public.NAME\n"
                "/graph switch NAME\n" + (
                    f"\nCurrent graph: {locationdistance.get_current_graph(update.effective_chat.id)}"
                    if current_graph else ''   
                ))
        
        if not args:
            return await print_usage(current_graph=True)
        
        module_name, = args

        chat_id = update.effective_chat.id
        with get_connection() as conn:
            my_simple_sql_create = partial(simple_sql_create_args, connection=conn)
            my_simple_sql_dict = partial(simple_sql_dict_args, connection=conn)
            my_simple_sql = partial(simple_sql_args, connection=conn)

            DOTTED = rf'({ListLangRegexes.NAME})[.]({ListLangRegexes.NAME})'
            SIMPLE = ListLangRegexes.NAME

            imported_chat_id = chat_id
            if module_name /fullmatchesI/ 'public[.]chat':
                raise UserError("public.chat is a reserved name")
            
            elif m := module_name /fullmatchesI/ DOTTED:
                actual_module = m.group(2).lower()
                visibility = 'public'
                namespace = m.group(1).lower()
                imported_chat_id, = only_one(my_simple_sql('select chat_id from LocationDistanceGraphNamespace where namespace=?', (namespace, )))
                
            elif m := module_name /fullmatchesI/ SIMPLE:
                actual_module = m.group(0).lower()
                visibility = 'chat'
            
            else:
                raise UserError("Wrong format for a graph name")
        
            all_graphs = my_simple_sql_dict(''' select rowid, visibility from LocationDistanceGraph where chat_id=? AND name=? AND visibility=?''', (imported_chat_id, actual_module, visibility))
            
            if len(all_graphs) == 0:
                r = my_simple_sql_create('''insert into LocationDistanceGraph(chat_id, name, visibility) VALUES (?,?,?)''', (chat_id, actual_module, visibility))
                graph = {'rowid': r['lastrowid'], 'visibility': visibility}
            elif len(all_graphs) == 1:
                graph = all_graphs[0]
            else:
                raise AssertionError
            
            if my_simple_sql_dict(''' select rowid from LocationDistanceCurrentGraphChat where chat_id=? ''', (chat_id, )):
                my_simple_sql_dict(''' update LocationDistanceCurrentGraphChat set graph_id=? where chat_id=? ''', (graph['rowid'], chat_id, ))
            else:
                my_simple_sql_create(''' insert into LocationDistanceCurrentGraphChat(graph_id, chat_id) VALUES(?,?)''', (graph['rowid'], chat_id, ))
            
        return await send(f"Current graph: {locationdistance.get_current_graph(chat_id)}")

    async def importgraph(args, update, context):
        send = make_send(update, context)
        chat_id = update.effective_chat.id

        graph_full_name, = args
        graph_full_name = graph_full_name.lower()

        if '.' not in graph_full_name:
            raise UserError("Must specify the namespace as namespace.graph")
        
        try:
            graph_namespace, graph_simple_name = graph_full_name.split('.')
        except ValueError:
            raise UserError("Only one dot in notation namespace.graph")
        
        with get_connection() as conn:
            my_simple_sql = partial(simple_sql_args, connection=conn)

            try:
                imported_chat_id, = only_one_specific(my_simple_sql('''select chat_id from LocationDistanceGraphNamespace where namespace=? ''', (graph_namespace, )))
            except NoRecords:
                raise UserError(f"No public namespace {graph_namespace}")

            try:
                graph_id, = only_one_specific(my_simple_sql(''' select rowid as graph_id from LocationDistanceGraph where chat_id=? and name=? and visibility="public" ''', (imported_chat_id, graph_simple_name, )))
            except NoRecords:
                raise UserError("There is no public graph with that name")
            
            if not my_simple_sql(''' select rowid from LocationDistanceImportedGraph where graph_id = ? and chat_id = ?''', (graph_id, chat_id)):
                my_simple_sql(''' insert into LocationDistanceImportedGraph(graph_id, chat_id) values (?,?)''', (graph_id, chat_id))

        return await send(f'Graph {graph_full_name} is now imported in the chat')
    
    async def listimport(args, update, context):
        ''' select g.name from imported_graph i join graph g on igraph_id=g.rowid '''

    async def unimportgraph(args, update, context):
        send = make_send(update, context)
        chat_id = update.effective_chat.id

        graph_full_name, = args
        graph_full_name = graph_full_name.lower()

        if '.' not in graph_full_name:
            raise UserError("Must specify the namespace as namespace.graph")

        try:
            graph_namespace, graph_simple_name = graph_full_name.split('.')
        except ValueError:
            raise UserError("Only one dot in notation namespace.graph")

        with get_connection() as conn:
            my_simple_sql = partial(simple_sql_args, connection=conn)

            try:
                imported_chat_id, = only_one_specific(my_simple_sql('''select chat_id from LocationDistanceGraphNamespace where namespace=? ''', (graph_namespace, )))
            except NoRecords:
                raise UserError(f"No public namespace {graph_namespace}")

            try:
                graph_id, = only_one_specific(my_simple_sql(''' select rowid as graph_id from LocationDistanceGraph where chat_id=? AND name=? and visibility="public" ''', (imported_chat_id, graph_simple_name, )))
            except NoRecords:
                raise UserError("There is no public graph with that name")
            
            if my_simple_sql(''' select rowid from LocationDistanceImportedGraph where graph_id = ? and chat_id = ?''', (graph_id, chat_id)):
                my_simple_sql(''' delete from LocationDistanceImportedGraph where graph_id = ? and chat_id = ? ''', (graph_id, chat_id))

        return await send(f"Graph {graph_full_name} isn't imported in the chat anymore")

    def get_current_graph_namespace(*, chat_id, connection) -> Optional[str]:
        my_simple_sql = partial(simple_sql_args, connection=connection)

        return only_one(only_one(my_simple_sql('''select namespace from LocationDistanceGraphNamespace where chat_id=?''', (chat_id, )) or [[None]]))

    async def graphnamespace(args, update, context):
        send = make_send(update, context)
        
        chat_id = update.effective_chat.id

        try:
            namespace, = args
        except:
            with get_connection() as conn:
                current_namespace = locationdistance.get_current_graph_namespace(chat_id=chat_id, connection=conn)
                return await send(f"Usage: /graph namespace NAMESPACE\n\nCurrent graph namespace: {current_namespace}")
        
        namespace = namespace.lower()

        if not namespace /fullmatches/ ListLangRegexes.NAME:
            raise UserError("Wrong format for a graph namespace")

        with get_connection() as conn:
            my_simple_sql = partial(simple_sql_args, connection=conn)

            current_namespace = locationdistance.get_current_graph_namespace(chat_id=chat_id, connection=conn)
            if current_namespace is None:
                my_simple_sql('''insert into LocationDistanceGraphNamespace(namespace, chat_id) values (?,?)''', (namespace, chat_id))
            else:
                my_simple_sql('''update LocationDistanceGraphNamespace set namespace=? where chat_id=? ''', (namespace, chat_id))
        
        return await send(f"Current graph namespace: {namespace}")

async def whereisanswer_responder(msg:str, send: AsyncSend, *, update, context):
    reply = update_get_reply(update)

    assert_true(reply and reply.text, DoNotAnswer)
    assert_true(reply.text.startswith('/whereis') or reply.text.startswith('/whereis@' + context.bot.username), DoNotAnswer)
    
    key = ' '.join(InfiniteEmptyList(reply.text.split())[1:])
    value = msg
    
    await save_thereis(key, value, update=update, context=context)

"""
List doc and tests:

# the equal statement with a (non empty) list
a = list
a = 
- b
- c
assert a == ["b", "c"]

a = 
- b
- c
assert "a" not in globals() 

a = tasklist
a = 
- b
- c
assert a == ["[ ] b", "[ ] c"]
"""

class ListLang:
    OPERATIONS_ONE_LINE = sorted((
        # crud
        "add", "append",
        "delete", "del",
        "insert",
        "replace", "rep",
        # list
        "clear",
        "shuffle",
        # reading
        "print", "list",
        "enumerate", "enum",
        # tasklist
        "check",
        "uncheck",
        # tree
        'addchild', 'insertchild', 
    ), key=len, reverse=True)

    DYNAMIC_TYPES = [
        'flashcard.current',
        # 'event.today',
        'flashcard.page',
    ]

    POSSIBLE_TYPES = [
        'list',
        'tasklist',
        'tree',
        'tasktree',
        'alias',
    ] + ['dynamic' + '.' + x for x in DYNAMIC_TYPES]
    
    OPS_1L = '|'.join(map(re.escape, OPERATIONS_ONE_LINE))

    IsTask = re.compile("^\\[\\s*(x|)\\s*\\]\\s*(.*)\\s*$")

class ListLangRegexes:
    NAME = r"[\p{L}-_][\p{L}-_\d]*"
    ONE_LINE, OP_MULTI_LINE, MULTI_EQUALS_PLUS_TYPE = (
        regex.compile(fr"({NAME})(\s*[.]\s*|\s+)({ListLang.OPS_1L}|[=])\s*(.*?)", regex.IGNORECASE),
        regex.compile(fr"({NAME})\s*([+][=]|[=])\s*()\s*\n(.*)", regex.DOTALL),
        regex.compile(fr"({NAME})\s*([=])\s*(.*?)\s*\n(.*)", regex.DOTALL))

async def list_responder(msg: str, send: AsyncSend, *, update, context):
    import regex

    read_chat_settings = make_read_chat_settings(update, context)

    RE = ListLangRegexes

    if (match := RE.ONE_LINE.fullmatch(msg)) or (match_multi := RE.OP_MULTI_LINE.fullmatch(msg)) or (match_multi_equals_plus_type := RE.MULTI_EQUALS_PLUS_TYPE.fullmatch(msg)):
        if match: # one line operation
            list_name, _, operation_raw, parameters = match.groups()

            operation = 'createempty' if operation_raw == '=' else operation_raw

            requested_type = parameters
            operation = operation.lower()

        elif match_multi or match_multi_equals_plus_type: # multi line operation
            list_name, operation_symbol, post_operation_symbol, parameters_text = (match_multi or match_multi_equals_plus_type).groups()
            if match_multi:
                operation = {'=':'editmulti', '+=': 'extendmulti'}[operation_symbol]
            else:
                operation = 'createassign'
                requested_type = post_operation_symbol

            parameters_lines = parameters_text.splitlines()
            parameters_lines = list(map(str.strip, parameters_lines))
            parameters_lines = list(filter(None, parameters_lines))
            if any(map(lambda x:x.startswith("-"), parameters_lines)) and not all(map(lambda x:x.startswith("-"), parameters_lines)):
                raise UserError("Either use dash notation or don't, not a mix")
            parameters_lines = [line[1:] if line.startswith("-") else line for line in parameters_lines]
            parameters_lines = list(map(str.strip, parameters_lines))
            parameters_lines = list(filter(None, parameters_lines))
            parameters = parameters_lines

        list_name: str
        operation: str
        parameters: str | list[str]  # list in case of multiline operation
        requested_type: None | str # in case of new list creation (a = list; or multi-line variant)

        if operation in ('createempty', 'createassign'):
            # the list should be (re) created
            requested_type = requested_type.lower()

            with sqlite3.connect("db.sqlite") as conn:
                conn.execute('begin transaction')
                chat_id, user_id = update.effective_chat.id, update.effective_user.id

                if requested_type[-1:] == '!':
                    # name = list!
                    # name = []!
                    # name = copy from other!
                    requested_type = requested_type[:-1]
                    force_creation = True
                else:
                    force_creation = False

                re_spaces = '\s+'
                re_spaces0 = '\s*'
                re_group = lambda x: '(' + x + ')'
                re_list_name = r'\p{L}(?:\p{L}|[-_\d])*' # examples: abc, a_b_c, a-b-c, abc34
                re_bits = lambda *args: ''.join(args)
                re_type_id = r'\p{L}(?:\p{L}|[.])*'  # examples: abc, abc.def

                if alias_list_match := requested_type /fullmatchesI/ re_bits('alias', re_spaces, re_group(re_list_name)):
                    target_alias = alias_list_match.group(1)
                    type_list = ('alias', target_alias)

                elif re.fullmatch(re.escape('[') + '\s*' + re.escape(']'), requested_type):
                    # name = list
                    # name = []
                    # name = [ ]
                    type_list = 'list'

                elif param_match := regex.compile(f'copy\s+((of|from)\s*)?({re_list_name})').fullmatch(requested_type):
                    # name = copy of other
                    # name = copy from other
                    _, _, copy_from_name = param_match.groups()
                    type_list = ('copy', copy_from_name)

                elif param_match := regex.compile('(tasktree)\s+((of|from)\s*)?({re_list_name})').fullmatch(requested_type):
                    copy_from_type, _, _, copy_from_name = param_match.groups()
                    type_list = (copy_from_type, copy_from_name)

                elif requested_type in set(ListLang.POSSIBLE_TYPES):
                    type_list = requested_type

                elif general_match := requested_type /fullmatchesI/ f'({re_type_id})(?:{re_spaces}(.*))?':
                    general_type, general_args = general_match.group(1) or '', general_match.group(2) or ''
                    general_type = general_type.lower()
                    general_args = general_args.strip()

                    if general_type in ListLang.DYNAMIC_TYPES:
                        if general_type == 'flashcard.page':
                            if ' ' in general_args:
                                raise UserError("Not a page format")
                    
                        type_list = ('dynamic', general_type) if not general_args else ('dynamic', general_type, general_args)
                    elif general_args:
                        raise DoNotAnswer
                    else:
                        raise UserError(f"List creation of type {requested_type!r} not implemented, use = list, for example")
                else:
                    raise DoNotAnswer

                type_list: str | tuple[str, ...]
                force_creation: bool

                try:
                    listsmodule.forcecreatelist.do_it(conn=conn, chat_id=chat_id, name=list_name, user_id=user_id, type_list=type_list, force_creation=force_creation)
                except listsmodule.ListAlreadyExist:
                    raise UserError(f"List already exist, use {requested_type+'!'!r} to delete old list and force creation of new list")
                
                if operation == 'createassign':
                    if requested_type == 'tasklist':
                        listsmodule.extendmultitasklist.do_it(conn=conn, chat_id=chat_id, name=list_name, values=parameters) 
                    elif requested_type == 'list':
                        listsmodule.extendmultilist.do_it(conn=conn, chat_id=chat_id, name=list_name, values=parameters) 
                    else:
                        raise UserError('Not possible at the moment with {requested_type!r}')

                conn.execute('end transaction')
                await send(f"List named {list_name!r} created")

        elif operation in ListLang.OPERATIONS_ONE_LINE + ['extendmulti', 'editmulti', ]:
            # simple operation on the list
            with sqlite3.connect("db.sqlite") as conn:
                conn.execute('begin transaction')

                chat_id = update.effective_chat.id

                async def operation_one_line(*, list_name, previous_aliases):
                  if list_name in previous_aliases:
                      raise UserError("Alias loop")
                  if listsmodule.list_exists(conn=conn, chat_id=chat_id, name=list_name):
                    
                    P = lambda: dict(conn=conn, name=list_name, chat_id=chat_id)
                    PP = lambda: P() | dict(parameters=parameters)

                    list_type = listsmodule.get_list_type(**P())
                    list_type_is_tree = list_type in ('tree', 'tasktree')
                    dynamic_match = list_type /fullmatchesI/ ('dynamic' + re.escape('.') + '(.*)')
                    dynamic_list = dynamic_match.group(1) if dynamic_match else None
                    alias_match = list_type /fullmatches/ ('alias' + re.escape('.') + '(.*)')
                    alias_list = alias_match.group(1) if alias_match else None

                    if alias_match:
                        if listsmodule.list_exists(conn=conn, chat_id=chat_id, name=alias_list):
                            return await operation_one_line(list_name=alias_list, previous_aliases=previous_aliases + (list_name, ))
                        else:
                            raise UserError(f"Alias target {alias_list!r} does not exist")

                    did_edit = True
                    if operation in ('add', 'append'):
                        if list_type in ('tasklist', 'tasktree'):
                            modified_value = listsmodule.make_task(parameters)
                            listsmodule.addtolist.do_it(**P(), value=modified_value)
                        elif list_type in ('list', 'tree'):
                            listsmodule.addtolist.do_it(**P(), value=parameters)
                        elif dynamic_list:
                            listsmodule.dynamic_add.do_it(**P(), value=parameters, dynamic_list=dynamic_list)
                        else:
                            raise DoNotAnswer
                                
                    elif operation in ('print', 'list', ):
                        space_between_lines = do_if_setting_on(read_chat_settings('list.space_between_lines'))
                        if list_type_is_tree:
                            indent = int_or_none(read_chat_settings('list.indent'))
                            await send(listsmodule.printtree.it(**P(), parameters=parameters, indent=indent, space_between_lines=space_between_lines))
                        elif dynamic_list:
                            await send(listsmodule.print_dynamic.it(**P(), parameters=parameters, dynamic_list=dynamic_list))
                        elif list_type in ('list', 'tasklist', ):
                            await send(listsmodule.printlist.it(**P(), parameters=parameters, space_between_lines=space_between_lines))
                        else:
                            raise DoNotAnswer
                        did_edit = False

                    elif operation in ('clear', ):
                        if list_type in ('list', 'tree', 'tasklist', 'tasktree'):
                            listsmodule.clearlist.do_it(conn=conn, name=list_name, chat_id=chat_id)
                        elif dynamic_list:
                            listsmodule.clear_dynamic.do_it(**P(), dynamic_list=dynamic_list)
                        else:
                            raise DoNotAnswer

                    elif operation in ('extendmulti', ):
                        if list_type in ('tasktree', ):
                            listsmodule.extendmultitasktree.do_it(conn=conn, name=list_name, chat_id=chat_id, values=parameters)
                        elif list_type in ('tree', ):
                            listsmodule.extendmultitree.do_it(conn=conn, name=list_name, chat_id=chat_id, values=parameters)
                        elif list_type in ('tasklist', ):
                            listsmodule.extendmultitasklist.do_it(conn=conn, name=list_name, chat_id=chat_id, values=parameters)
                        elif list_type in ('list', ):
                            listsmodule.extendmultilist.do_it(conn=conn, name=list_name, chat_id=chat_id, values=parameters)
                        elif dynamic_list:
                            listsmodule.extend_multi_dynamic.do_it(**P(), values=parameters, dynamic_list=dynamic_list)
                        else:
                            raise DoNotAnswer
                    
                    elif operation in ('editmulti', ):
                        if list_type in ('tasktree', ):
                            raise UserError("Impossible at the moment")
                        if list_type in ('tasklist', ):
                            listsmodule.editmultitasklist.do_it(conn=conn, name=list_name, chat_id=chat_id, values=parameters)
                        elif list_type in ('list', 'tree'):
                            listsmodule.editmultilist.do_it(conn=conn, name=list_name, chat_id=chat_id, values=parameters)
                        else:
                            raise DoNotAnswer

                    elif operation in ('shuffle', ):
                        if list_type_is_tree:
                            raise UserError("Operation not possible")
                        elif list_type in ('list', 'tasklist'):
                            listsmodule.shuffle.do_it(conn=conn, name=list_name, chat_id=chat_id)
                        else:
                            raise DoNotAnswer

                    elif operation in ('enum', 'enumerate', ):
                        space_between_lines = do_if_setting_on(read_chat_settings('list.space_between_lines'))
                        if list_type_is_tree:
                            indent = int_or_none(read_chat_settings('list.indent'))
                            await send(listsmodule.enumeratetree.it(**P(), parameters=parameters, indent=indent, space_between_lines=space_between_lines))
                        elif dynamic_list:
                            await send(listsmodule.enumerate_dynamic.it(**P(), parameters=parameters, dynamic_list=dynamic_list))
                        elif list_type in ('list', 'tasklist'):
                            await send(listsmodule.enumeratelist.it(**P(), parameters=parameters, space_between_lines=space_between_lines))
                        else:
                            raise DoNotAnswer
                        
                        did_edit = False
                    
                    elif operation in ('del', 'delete'):
                        if list_type_is_tree:
                            listsmodule.delintree(**P()).run(parameters=parameters)
                        elif list_type in ('list', 'tasklist'):
                            listsmodule.delinlist.do_it(**P(), value=parameters)
                        elif dynamic_list:
                            listsmodule.delindynamic.do_it(**P(), value=parameters, dynamic_list=dynamic_list)
                        else:
                            raise DoNotAnswer
                    
                    elif operation in ('insert', ):
                        if list_type in ('tasklist', ):
                            listsmodule.insertintasklist.do_it(**P(), parameters=parameters)
                        elif list_type in ('tasktree', ):
                            listsmodule.insertintasktree.do_it(**P(), parameters=parameters)
                        elif list_type in ('tree', ):
                            listsmodule.insertintree(**P()).run(parameters=parameters)
                        elif list_type in ('list', ):
                            listsmodule.insertinlist.do_it(**P(), parameters=parameters)
                        else:
                            raise DoNotAnswer
                    
                    elif operation in ('rep', 'replace'):
                        if list_type in ('tasklist', ):
                            listsmodule.replaceintasklist.do_it(**PP())
                        elif list_type_is_tree:
                            if list_type in ('tasktree', ):
                                listsmodule.replaceintasktree.do_it(**PP())
                            else:
                                listsmodule.replaceintree.do_it(**PP())
                        elif list_type in ('list', ):
                            listsmodule.replaceinlist.do_it(**PP())
                        else:
                            raise DoNotAnswer
                    
                    elif operation in ('insertchild', 'addchild', ):
                        if list_type_is_tree:
                            if list_type in ('tasktree', ):
                                listsmodule.tasktreeinsertchild.do_it(**PP())
                            else:
                                listsmodule.treeinsertchild(**P()).run(parameters=parameters)
                        else:
                            raise DoNotAnswer

                    elif operation in ('check', 'uncheck', ):
                        if list_type in ('tasklist', 'tasktree'):
                            if operation == 'check':
                                if list_type == 'tasktree':
                                    listsmodule.tasktreecheck(**P()).run(value=parameters, direction='x')
                                else:
                                    listsmodule.tasklistcheck.do_it(conn=conn, name=list_name, chat_id=chat_id, value=parameters, direction='x')
                            elif operation == 'uncheck':
                                if list_type == 'tasktree':
                                    listsmodule.tasktreecheck(**P()).run(value=parameters, direction=' ')
                                else:
                                    listsmodule.tasklistcheck.do_it(conn=conn, name=list_name, chat_id=chat_id, value=parameters, direction=' ')
                        else:
                            raise DoNotAnswer

                    else:
                        raise AssertionError(f"On operation {operation}")
                
                    if did_edit:
                        await send(f"List {list_name!r} edited")

                await operation_one_line(list_name=list_name, previous_aliases=())
                conn.execute('end transaction')
            
        else:
            return await send(f"I should do the operation {operation!r} on the list named {list_name!r} (not implemeted yet)")

def int_or_none(setting):
    if setting is None:
        return None
    try:
        return int(setting)
    except ValueError:
        return None

async def eventedit_responder(msg:str, send: AsyncSend, *, update, context):
    reply = update_get_reply(update)
    assert_true(reply and reply.text, DoNotAnswer)
    user_id, chat_id = update.effective_user.id, update.effective_chat.id

    try:
        event = addevent_analyse_from_bot(update, context, reply.text)
    except ValueError:
        raise DoNotAnswer
    
    if match_postpone := re.fullmatch('([+]|[-])\s*(\d+)\s*(h|hours|min|minute|minutes|day|days|week|weeks)', msg):
        event_db = retrieve_event_from_db(update=update, context=context, what=event['what'], when=event['when'])
        
        sign, amount, units = match_postpone.groups()
        amount = int(amount)
        sign = +1 if sign == '+' else -1
        delta = amount * sign * timedelta(**{{'h':'hours', 'hours':'hours', 'min':'minutes', 'minute':'minutes', 'minutes':'minutes', 'day':'days', 'days':'days', 'weeks':'weeks', 'week':'weeks'}[units]:1})

        before = DatetimeDbSerializer.strptime(event_db['date'])
        after = before + delta

        read_chat_settings = make_read_chat_settings(update, context)
        do_event_admin_check('edit', setting=read_chat_settings('event.admins'), user_id=update.effective_user.id)

        simple_sql(('''UPDATE Events SET date=? where rowid=?''', (DatetimeDbSerializer.strftime(after), event_db['rowid'], )))

    elif match_field_edit := msg /fullmatches_with_flags(re.I)/ '(\p{L}*)[ ]*([=]?)[ ]*(.*)':
        read_chat_settings = make_read_chat_settings(update, context)

        field_name = match_field_edit.group(1).lower()
        equal_present = match_field_edit.group(2)
        new_value = match_field_edit.group(3)
        possible_fields = ['name', 'date', 'datetime', 'time', 'when', 'what', 'where', 'location']

        if not field_name:
            # Guessing
            time, rest = ParseEvents.parse_time(new_value.split())
            if time and not rest:
                field_name = 'time'
            else:
                raise DoNotAnswer
        else:
            if not equal_present:
                raise DoNotAnswer
            if field_name not in possible_fields:
                raise DoNotAnswer

        event_db = retrieve_event_from_db(update=update, context=context, what=event['what'], when=event['when'])
        event_rich = split_event_with_where_etc({'what': event_db['name']})

        if field_name in ('name', 'what'):
            event_rich['what'] = new_value
            db_datetime = event_db['date']
        
        elif field_name in ('where', 'location'):
            event_rich['where'] = new_value
            db_datetime = event_db['date']

        elif field_name in ('datetime', 'when'):

            tz = induce_my_timezone(user_id=user_id, chat_id=chat_id)
            check_tz_in_chat(tz=tz, chat_timezones=read_chat_settings('event.timezones'))
            
            datetime_object = parse_datetime_point(update, context, when_infos=new_value, what_infos=event_db['name']).datetime.replace(tzinfo=tz)
            db_datetime = DatetimeDbSerializer.strftime(datetime_object.astimezone(UTC))
        
        elif field_name in ('time', ):
            time, rest = ParseEvents.parse_time(new_value.split())
            assert_true(not rest, UserError("Too much values for time"))
            assert_true(time, UserError("Cannot parse time"))

            tz = induce_my_timezone(user_id=user_id, chat_id=chat_id)
            check_tz_in_chat(tz=tz, chat_timezones=read_chat_settings('event.timezones'))

            datetime_object = DatetimeDbSerializer.strptime(event_db['date']).replace(tzinfo=UTC).astimezone(tz)
            datetime_object = Datetime.combine(datetime_object.date(), time).replace(tzinfo=tz)
            db_datetime = DatetimeDbSerializer.strftime(datetime_object.astimezone(UTC))

        elif field_name in ('date', ):
            mini_event, rest = parse_event_date(new_value.split())
            assert_true(not rest, UserError("Too much values for date"))
            assert_true(mini_event.date_str, UserError("Cannot parse date"))

            tz = induce_my_timezone(user_id=user_id, chat_id=chat_id)
            check_tz_in_chat(tz=tz, chat_timezones=read_chat_settings('event.timezones'))

            date, date_end = DatetimeText.to_date_range(mini_event.date_str, tz=tz)

            datetime_object = DatetimeDbSerializer.strptime(event_db['date']).replace(tzinfo=UTC).astimezone(tz)
            datetime_object = Datetime.combine(date, datetime_object.time()).replace(tzinfo=tz)
            db_datetime = DatetimeDbSerializer.strftime(datetime_object.astimezone(UTC))

        else:
            raise DoNotAnswer
        
        db_what = event_rich['what'] if not event_rich.get('where') else event_rich['what'] + ' @ ' + event_rich['where']
        simple_sql(('''UPDATE Events SET name=?, date=? where rowid=?''', (db_what, db_datetime, event_db['rowid'], )))

    else:
        raise DoNotAnswer

    return await send(format_event_emoji_style_from_event_id(event_db['rowid'], chat_id=chat_id, user_id=user_id))

def format_event_emoji_style_from_event_id(event_id, *, chat_id, user_id):
    new_event_db = only_one(simple_sql_dict(('select rowid, date, name from Events where rowid=?', (event_id,))))
    date_utc = new_event_db['date']
    name = new_event_db['name']
    tz = induce_my_timezone(chat_id=chat_id, user_id=user_id)
    datetime = DatetimeDbSerializer.strptime(date_utc).replace(tzinfo=ZoneInfo('UTC')).astimezone(tz)
    date, time = datetime.date(), datetime.time()
    read_chat_settings = make_read_chat_settings_from_chat_id(chat_id)
    chat_timezones = read_chat_settings("event.timezones")
    return format_event_emoji_style(name=name, datetime=datetime, date=date, time=time, tz=tz, chat_timezones=chat_timezones)


class GetOrEmpty(list):
    def __getitem__(self, i):
        try:
            return super().__getitem__(i)
        except IndexError:
            return ''
InfiniteEmptyList = GetOrEmpty

from collections import namedtuple
NamedChatDebt = namedtuple('NamedChatDebt', 'chat_id, debitor_id, creditor_id, amount, currency, reason')

async def sharemoney_responder(msg:str, send: AsyncSend, *, update, context):
    chat_id = update.effective_chat.id

    def Amount():
        from pyparsing import Word, nums, infix_notation, opAssoc, one_of

        class EvalConstant:
            def __init__(self, tokens):
                self.value = tokens[0]

            def eval(self):
                return int(self.value)
            
        class EvalOne:
            SIGNS = {"+": 1, "-": -1}
            def __init__(self, tokens) -> None:
                self.sign, self.value = tokens[0]

            def eval(self):
                return self.SIGNS[self.sign] * self.value.eval()
        
        def operator_operands(tokenlist):
            """ generator to extract operators and operands in pairs """
            it = iter(tokenlist)
            while True:
                try:
                    yield (next(it), next(it))
                except StopIteration:
                    break

        class EvalTwo:
            OPS = {
                '+': lambda x,y: x+y,
                '-': lambda x,y: x-y,
                '*': lambda x,y: x*y,
                '/': lambda x,y: x/y,
            }
            def __init__(self, tokens):
                self.value = tokens[0]
            
            def eval(self):
                acc = self.value[0].eval()
                for op, val in operator_operands(self.value[1:]):
                    acc = self.OPS[op](acc, val.eval())
                return acc


        arithmetics = infix_notation(
            Word(nums).set_parse_action(EvalConstant),
            [
                (one_of("+ -"), 1, opAssoc.RIGHT, EvalOne),
                (one_of("* /"), 2, opAssoc.LEFT, EvalTwo),
                (one_of("+ -"), 2, opAssoc.LEFT, EvalTwo),
            ]
        )
        return arithmetics

    import regex
    name = regex.compile(r"(\p{L}\w*)([.]([A-Za-z]+))?")
    amount = Amount()
    currency_re = regex.compile('[A-Z]{3}')
    Args = GetOrEmpty(msg.split())    

    # Name owes Name 50
    # Name owes Name 50 [EUR]
    # Name owes Name 50 [EUR] [for Something]

    if name.fullmatch(Args[0]) and Args[1] in ('owes', 'paid') and name.fullmatch(Args[2]) and amount.matches(Args[3]):
        i = 4
        if currency_re.fullmatch(Args[i]):
            currency_string: str = Args[i].upper()
            i += 1
        else:
            currency_string: str = None
        
        if Args[i] in ('for', '#') and Args[i+1]:
            reason = ' '.join(Args[i+1:])
        elif Args[i].startswith('#') and Args[i][1:]:
            reason = ' '.join([Args[i][1:]] + Args[i+1:])
        elif Args[i]:
            raise UserError("Too much infos")
        else:
            reason = None

        reason: str | None

        first_name, operation, second_name, amount_str = Args[:4]

        if operation == 'paid':
            first_name, second_name = second_name, first_name
            # now it's like owes

        first_currency, second_currency = map(lambda x: name.fullmatch(x).group(3), (first_name, second_name))

        if first_currency or second_currency:
            if not (first_currency and second_currency):
                raise UserError("Currencies must match")
            if not (first_currency.upper() == second_currency.upper()):
                raise UserError("Currencies must match")
        
        if currency_string:
            if first_currency or second_currency:
                if not(first_currency.upper() == second_currency.upper() == currency_string.upper()):
                    raise UserError("Currencies must match")
        
        the_currency = currency_string or first_currency or second_currency
        del currency_string, first_currency, second_currency
        
        if the_currency:
            first_name = name.fullmatch(first_name).group(1) + "." + the_currency.upper()
            second_name = name.fullmatch(second_name).group(1) + "." + the_currency.upper()
        
        debt = NamedChatDebt(
            debitor_id=first_name,
            creditor_id=second_name,
            chat_id=chat_id,
            amount=amount.parse_string(amount_str, parse_all=True)[0].eval(),
            reason=reason,
            currency=the_currency and the_currency.upper())
    
        read_chat_settings = make_read_chat_settings(update, context)

        if do_if_setting_on(read_chat_settings('sharemoney.required_for')):
            if not debt.reason:
                raise UserError('You must specify a reason (group policy)\nExample: John owes Maria 5 for bowling')
        
        simple_sql((
            'insert into NamedChatDebt(debitor_id, creditor_id, chat_id, amount, currency, reason) values (?,?,?,?,?,?)',
            (debt.debitor_id, debt.creditor_id, debt.chat_id, debt.amount, debt.currency, debt.reason)))
        
        return await send(' '.join(filter(None,
            ('Debt created:', f'"{debt.debitor_id}"', 'owes', f'"{debt.creditor_id}"', f'{debt.amount}', (f'# {debt.reason}' if debt.reason else ''))
        )))

async def flashcard_responder(msg, send, *, update, context):
    if ' / ' in msg:
        context.args = msg.split()
        return await add_flashcard(update, context, scope='general')

async def englishpractice_responder(msg: str, send: AsyncSend, *, update, context):
    pass

class EnglishPracticeFilter(MessageFilter):
    def filter(self, message: Message):

        def custom_make_read_chat_settings(chat_id):
            from functools import partial
            return partial(read_settings, id=chat_id, settings_type='chat')
        
        read_chat_settings = custom_make_read_chat_settings(chat_id=message.chat.id)

        return do_if_setting_on(read_chat_settings('englishpractice.active'))

RESPONDERS = (
    (hello_responder, 'hello', 'on'),
    (money_responder, 'money', 'on'),
    (sharemoney_responder, 'sharemoney', 'off'),
    (whereisanswer_responder, 'whereisanswer', 'on'),
    (eventedit_responder, 'eventedit', 'on'),
    (list_responder, 'list', 'off'),
    (englishpractice_responder, 'englishpractice', 'off'),
    (locationdistance_responder, 'locationdistance', 'off'),
    (flashcard_responder, 'flashcard', 'off'),
)

async def on_message(update: Update, context: CallbackContext):
    send = make_send(update, context)
    
    # if update.message:
    #     logging.info("@{username}: {text} (In {group})".format(
    #         username=update.message.from_user.username,
    #         text=update.message.text,
    #         group='private' if update.message.chat.type == ChatType.PRIVATE else
    #                "'{}'".format(update.message.chat.title) if update.message.chat.type in (ChatType.GROUP, ChatType.SUPERGROUP, ChatType.CHANNEL) else
    #                update.message.chat.type))
    # else:
    #     logging.info("{}".format(update))

    if update.message:
        msg = strip_botname(update, context)
        read_settings = make_read_chat_settings(update, context)

        for responder, setting, default in RESPONDERS:
            if (read_settings(setting + '.active') or default) == 'off':
                continue
            try:
                await responder(msg, send, update=update, context=context)
            except DoNotAnswer:
                pass
            except Exception as e:
                await log_error(e, send)

    if update.edited_message:
        pass

async def caps(update: Update, context: CallbackContext):
    text_caps = str(context.args).upper()
    send = make_send(update, context)
    await send(text_caps)

def unilinetext(x):
    import unicodedata
    return "U+{} {} {}".format(hex(ord(x))[2:].upper().zfill(4), x, unicodedata.name(x, '?'))

async def uniline(update, context):
    send = make_send(update, context)
    if not (reply := update_get_reply(update)) and not context.args:
        return await send("Usage: /uniline word1 word2\nCan also be used on a reply message")
    for arg in ([reply.text] if reply else []) + list(context.args):
        S = map(unilinetext, arg)
        await send('\n'.join(S) or '[]')

async def nuniline(update, context):
    send = make_send(update, context)
    nonascii = lambda x: ord(x) > 0x7F
    if not (reply := update_get_reply(update)) and not context.args:
        return await send("Usage: /nuniline word1 word2\nCan also be used on a reply message")
    for arg in ([reply.text] if reply else []) + list(context.args):
        S = map(unilinetext, filter(nonascii, arg))
        await send('\n'.join(S) or '[]')

async def befluent(update, context):
    send = make_send(update, context)

    await send("Hello")

    return ConversationHandler.END

async def ru(update: Update, context: CallbackContext):
    send = make_send(update, context)
    
    if reply := update_get_reply(update):
        pass
    elif not context.args:
        return await send("Usage: /ru word1 word2 word3...")
    else:
        pass
    d1 = ("azertyuiopqsdfghjklmwxcvbn",
          "азертыуиопясдфгхйклмвхцвбн")
    d2 = ("sh shch ch ye yu zh ya yo".split(),
          "ш  щ    ч  э  ю  ж  я  ё".split())
    d3 = ("' ''".split(), 
          'ь ъ'.split())
    D = (dict(zip(*d1))
       | dict(zip(*d2))
       | dict(zip(*map(str.upper, d1)))
       | dict(zip(map(str.upper, d2[0]), map(str.upper, d2[1])))
       | dict(zip(*d3))
       | dict(zip(map(str.capitalize, d2[0]), map(str.upper, d2[1]))))
    S = sorted(D, key=len, reverse=True)
    import re
    R = re.compile('|'.join(map(re.escape, S)))
    def to_cyrilic(word):
        return R.sub(lambda m: (lambda x: D.get(x,x))(m.group(0)), word)
    await send(' '.join(to_cyrilic(word) for word in context.args) if not reply or context.args else to_cyrilic(reply.text))

with open("json/ipa/fr_FR.json") as IPA_DATA_FR:
    IPA_DATA_FR = json.load(IPA_DATA_FR)

async def ipa_display(update:Update, context: ContextTypes.DEFAULT_TYPE, *, mode:Literal['ipa', 'ru']):
    send = make_send(update, context)

    if not context.args:
        raise UsageError("/{command} word+")

    *words, = context.args

    def tr(words, mode=mode):
        """
        >>> tr('je mange une pizza'.split())
        '/ʒə mɑ̃ʒ yn pidza/'
        """
        ipa_dict = IPA_DATA_FR['fr_FR'][0]
        to_ipa = lambda x: ipa_dict.get(x.lower(), '"{}"'.format(x.lower()))
        deal_multiple = lambda x: '({})'.format(x.replace(', ', '|').replace(',', '|')) if ',' in x else x
        strip_bars = lambda x: x.replace('/', '')
        mapping = lambda x: strip_bars(deal_multiple(to_ipa(x)))
        return {'ipa': '/', 'ru': '['}[mode] + ' '.join(map(mapping, words)) + {'ipa': '/', 'ru': ']'}[mode]
    
    def trru(words, mode=mode):
        """
        >>> trru('je mange une pizza'.split())
        '/жё ма̃ж юн пидза/'
        """
        mapping_ipa_ru = (
            'a b d e f g i j k l m n o p r s t u v w x y z ø ŋ œ ɑ ɔ ə ɛ ɡ ɥ ɪ ɲ ʁ ʃ ʊ ʒ œ\u0303'.split(),
            'а б д е ф г и й к л м н о п р с т у в у х ю з ё н ё а о ё э г ю и н р ш у ж у\u0303'.split())
        
        # todo: do not touch the quoted symbols

        mapping_ipa_ru_dict = dict(zip(*mapping_ipa_ru))

        import re
        Re = re.compile('|'.join(map(re.escape, sorted(mapping_ipa_ru_dict, key=len, reverse=True))))

        t = tr(words, mode=mode)
        mapping = lambda x: mapping_ipa_ru_dict.get(x,x)
        return Re.sub(lambda m: mapping(m.group(0)), t)

    if mode == 'pron':
        def func(words):
            return "{}\n{}".format(tr(words, mode='ipa'), trru(words, mode='ru'))
    else:
        func = {'ipa': tr, 'ru': trru}[mode]
    return await send(func(words))

async def ipa(update:Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        return await ipa_display(update, context, mode='ipa')
    except UsageError as e:
        send = make_send(update, context)
        return await send(str(e).format(command="ipa"))
    
async def iparu(update:Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        return await ipa_display(update, context, mode='ru')
    except UsageError as e:
        send = make_send(update, context)
        return await send(str(e).format(command="iparu"))

async def pron(update:Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        return await ipa_display(update, context, mode='pron')
    except UsageError as e:
        send = make_send(update, context)
        return await send(str(e).format(command="pron"))

def get_or_empty(L: list, i:int) -> str | object:
    try:
        return L[i]
    except IndexError:
        return ''

def make_read_my_settings(update: Update, context: CallbackContext=None, connection=None):
    from functools import partial
    return partial(read_settings, id=update.effective_user.id, settings_type='user')

def make_read_chat_settings(update: Update, context: CallbackContext=None, connection=None):
    from functools import partial
    return partial(read_settings, id=update.effective_chat.id, settings_type='chat')

def make_read_chat_settings_from_chat_id(chat_id, connection=None):
    from functools import partial
    return partial(read_settings, id=chat_id, settings_type='chat')

DICT_ENGINES = ('wikt', 'larousse', 'glosbe')

async def dict_command(update: Update, context: CallbackContext, *, engine:Literal['wikt'] | Literal['larousse'] | Literal['glosbe'], command_name:str):
    send = make_send(update, context)
    read_my_settings = make_read_my_settings(update, context)

    reply = update_get_reply(update)
    if not context.args:
        if not reply:
            return await send(f"Usage: /{command_name} word1 word2 word3...\nCan also be used on a reply message")

    if reply:
        reply_message_words = reply.text.split()
        
        def any_number(items):
            import re
            number = re.compile('-?\\d+')
            return any(map(number.fullmatch, items))

        def substitute_numbers(items):
            import re
            number = re.compile('-?\\d+')
            for item in items:
                if number.fullmatch(item):
                    i = int(item) 
                    if i == 0:
                        yield item
                    else:
                        idx = i if i < 0 else i-1
                        try:
                            yield reply_message_words[idx]
                        except IndexError:
                            yield item
                else:
                    yield item

    Args = InfiniteEmptyList(context.args)
    if Args[-1].startswith('/'):
        language = Args[-1][1:]
        parameter_words = Args[:-1]
        if ':' in language:
            base_lang, target_lang, *_ = language.split(':')
        else:
            base_lang, target_lang = '', language
    else:
        parameter_words = Args[:]
        base_lang = None
        target_lang = None
    
    if reply:
        if any_number(parameter_words):
            words = list(substitute_numbers(parameter_words))
        else:
            words = reply_message_words + parameter_words
    else:
        words = parameter_words
    
    new_words = []
    for w in words:
        if "'" in w:
            bits = w.strip("'").split("'")
            if bits:
                for bit in bits:
                    new_words.append(bit)
                    new_words.append("'")
                new_words.pop()
        else:
            new_words.append(w)
    words = new_words
    
    base_lang, target_lang

    base_lang = base_lang or read_my_settings(f'{command_name}.description')
    target_lang = target_lang or read_my_settings(f'{command_name}.text')

    # lang transformation
    if engine == 'wikt':
        target_lang = WIKTIONARY_LANGUAGES.get(base_lang or 'en', {}).get(target_lang, target_lang)
    elif engine == 'larousse':
        target_lang = LAROUSSE_LANGUAGES.get('fr', {}).get(target_lang or 'fr', target_lang)
        base_lang = LAROUSSE_LANGUAGES.get('fr', {}).get(base_lang or 'fr', base_lang)
    elif engine == 'glosbe':
        pass
    else:
        raise UsageError("Engine is misconfigured, please run /mysettings dict.engine {}".format('|'.join(DICT_ENGINES)))

    display_html = do_unless_setting_off(read_my_settings(f"{command_name}.html"))
    
    # url maker
    if engine == 'wikt':
      def url(x):
        x = x.lower()
        return (
            'https://wiktionary.com/wiki/'
            + ('{}:'.format(base_lang) if base_lang else '')
            + x
            + ('#{}'.format(target_lang) if target_lang else '')
        )
    elif engine == 'larousse':
      def url(x):
        x = x.lower()
        return (
            f'https://larousse.fr/dictionnaires/{target_lang}/{x}' if target_lang == base_lang else 
            f'https://larousse.fr/dictionnaires/{target_lang}-{base_lang}/{x}'
        )
    elif engine == 'glosbe':
      def url(x):
        x = x.lower()
        return f'https://glosbe.com/{target_lang}/{base_lang}/{x}'

    import html
    if display_html:
        send_html = partial(send, parse_mode='HTML', disable_web_page_preview=len(words) > 2)
        return await send_html(' | '.join(f'<a href="{html.escape(url(x))}">{html.escape(x)}</a>' for x in words))
    else:
        return await send('\n\n'.join(url(x) for x in words))
    

wikt = partial(dict_command, command_name='wikt', engine='wikt')
larousse = partial(dict_command, command_name='larousse', engine='larousse')

async def dict_(update: Update, context: CallbackContext):
    read_my_settings = make_read_my_settings(update, context)
    engine = read_my_settings('dict.engine')
    if not engine:
        raise UserError('Engine not set for /dict command, use "/mysettings dict.engine wikt" for example to set wiktionary engine')
    return await dict_command(update, context, command_name='dict', engine=engine)

class UsageError(Exception):
    pass

async def add_flashcard(update, context, *, scope=Literal['personal', 'general']):
    send = make_send(update, context)
    try:
      if reply := update_get_reply(update):
        sentence = reply.text 
        translation = ' '.join(context.args)
      else:
        def find_sentence_translation(args):
            if any(x in args for x in ("=", "/")):
                separator_position = args.index("=" if "=" in args else "/")
                sentence, translation = args[:separator_position], args[separator_position+1:]
                sentence, translation = map(' '.join, (sentence, translation))
            elif len(args) == 2:
                sentence, translation = args
            elif len(args) == 1:
                sentence, translation = args[0], ''
            else:
                raise UsageError
            return sentence, translation
        
        sentence, translation = find_sentence_translation(context.args)
    except UsageError:
        return await send("Usage:\n/flashcard word translation\n/flashcard words+ = translation+\n/flashcard words+ / translation+\nCan also be used on a reply message to replace the words")
    
    the_id = (update.effective_chat.id if scope == 'general' else
              update.effective_user.id if scope == 'personal' else raise_error(AssertionError))

    page_id = get_current_flashcard_page_id(chat_id=the_id)
    save_flashcard(sentence, translation, page_id=page_id)

    await send(f"New flashcard:\n{sentence}\n→ {translation}")

def get_current_flashcard_page_id(*, chat_id, connection=None) -> dict:
    def op(conn):
        results = conn.execute("select rowid from flashcardpage where chat_id=? and current=1", (chat_id,)).fetchall()
        if not results:
            return create_default_flashcard_page(chat_id=chat_id, connection=conn)
        else:
            page_id, = only_one(results)
            return page_id 

    return db_connect_or_use(connection, op)

def get_named_flashcard_page_id(*, page_name, chat_id, connection=None):
    def op(conn):
        results = conn.execute('select rowid from flashcardpage where chat_id=? and name=?', (chat_id, page_name)).fetchall()
        if not results:
            return create_named_flashcard_page(page_name=page_name, chat_id=chat_id, connection=conn)
        else:
            page_id, = only_one(results)
            return page_id
        
    return db_connect_or_use(connection, op)

def create_named_flashcard_page(*, page_name, chat_id, connection=None):
    def op(conn):
        cursor = conn.cursor()
        cursor.execute("insert into flashcardpage(chat_id, name, current) values (?,?,?)", (chat_id, page_name, 0))
        return cursor.lastrowid
    
    return db_connect_or_use(connection, op)

def create_default_flashcard_page(*, chat_id, connection=None):
    def op(conn):
        cursor = conn.cursor()
        cursor.execute("insert into flashcardpage(chat_id, name, current) values (?,?,?)", (chat_id, '1', 1))
        return cursor.lastrowid

    return db_connect_or_use(connection, op)

def db_connect_or_use(connection, op):
    if connection is None:
        with sqlite3.connect("db.sqlite") as conn:
            return op(conn)
    else:
        return op(connection)

def save_flashcard(sentence, translation, *, page_id, connection=None):
    query = ('insert into Flashcard(sentence, translation, page_id) values (?,?,?)', (sentence, translation, page_id))
    simple_sql(query, connection=connection)

def get_connection(connection=None):
    if connection is None:
        return sqlite3.connect('db.sqlite')
    else:
        return connection

def simple_sql(query, *, connection=None):
    conn = connection
    assert isinstance(query, (tuple, list))
    assert isinstance(query[1], (tuple, list))
    if conn:
        return conn.execute(*query).fetchall()
    with sqlite3.connect("db.sqlite") as conn:
        return conn.execute(*query).fetchall()

def simple_sql_args(text, args, *, connection):
    return simple_sql((text, args), connection=connection)

SimpleSqlModifyReturn = TypedDict('SimpleSqlModifyReturn', {'rowcount': int})
SimpleSqlCreateReturn = TypedDict('SimpleSqlModifyReturn', {'lastrowid': int})

def simple_sql_modify(query, *, connection=None) -> SimpleSqlModifyReturn:
    conn = connection
    assert isinstance(query, (tuple, list))
    assert isinstance(query[1], (tuple, list))
    if conn:
        cursor = conn.cursor()
        cursor.execute(*query).fetchall()
        return SimpleSqlModifyReturn(rowcount=cursor.rowcount)
    with sqlite3.connect("db.sqlite") as conn:
        cursor = conn.cursor()
        cursor.execute(*query).fetchall()
        return SimpleSqlModifyReturn(rowcount=cursor.rowcount)

def simple_sql_create(query, *, connection=None) -> SimpleSqlCreateReturn:
    conn = connection
    assert isinstance(query, (tuple, list))
    assert isinstance(query[1], (tuple, list))
    if conn:
        cursor = conn.cursor()
        cursor.execute(*query)
        return SimpleSqlCreateReturn(lastrowid=cursor.lastrowid)
    with sqlite3.connect("db.sqlite") as conn:
        cursor = conn.cursor()
        cursor.execute(*query)
        return SimpleSqlCreateReturn(lastrowid=cursor.lastrowid)

def simple_sql_modify_args(text, args, *, connection):
    return simple_sql_modify((text, args), connection=connection)

def simple_sql_create_args(text, args, *, connection) -> SimpleSqlCreateReturn:
    return simple_sql_create((text, args), connection=connection)

def simple_sql_dict(query, *, connection=None):
    conn = connection
    assert isinstance(query, (tuple, list))
    assert isinstance(query[1], (tuple, list))
    if conn:
        saved = conn.row_factory
        conn.row_factory = sqlite3.Row
        ret = conn.execute(*query).fetchall()
        conn.row_factory = saved
        return ret
    with sqlite3.connect("db.sqlite") as conn:
        conn.row_factory = sqlite3.Row
        return conn.execute(*query).fetchall()

def simple_sql_dict_args(text, args, *, connection):
    return simple_sql_dict((text, args), connection=connection)

async def practiceflashcards(update, context):
    send = make_send(update, context)

    async def print_usage():
        return await send("Usage:\n/practiceflashcards")

    args, kwargs = clean_kwarg_args(context.args)
    Args = InfiniteEmptyList(args)

    try:
        if n := kwargs.get('n'):
            n = int(only_one(n))
        elif Args[0]:
            n = int(Args[0])
        else:
            n = None
        
        if {'reversed', 'reverse'} & kwargs.keys():
            direction = 'reversed'
        else:
            direction = 'normal'

        if page_name := kwargs.get('page'):
            page_name = only_one(page_name)
        else:
            page_name = flashcard.Current

    except UsageError:
        await print_usage()
        return ConversationHandler.END
    
    chat_id = update.effective_chat.id

    with get_connection() as conn:
        DbChatInfo = dict(chat_id=chat_id, connection=conn)
        my_simple_sql = partial(simple_sql_args, connection=conn)
        my_simple_sql_dict = partial(simple_sql_dict_args, connection=conn)
        
        page_id = (get_current_flashcard_page_id(**DbChatInfo) if page_name is flashcard.Current else
                   get_named_flashcard_page_id(page_name=page_name, **DbChatInfo))
        
        query = ('''select sentence, translation from flashcard inner join flashcardpage on flashcard.page_id = flashcardpage.rowid
                    where flashcardpage.rowid=?''', (page_id, ))
        lines = my_simple_sql(*query)

    if not lines:
        page_info = only_one(my_simple_sql_dict('select rowid, name, current from flashcardpage where rowid=?', (page_id, )))
        await send(f"No flashcards for page {page_info['name']}")
        return ConversationHandler.END
    
    import random
    sample = random.sample(lines, n if n is not None else len(lines))
    sentences = [x[0] if direction == 'normal' else x[1] for x in sample[0:1]]
    
    await send(f"Welcome to the practice of {len(sample)} flashcards\n"
               "Type the answer to each question\n"
               "To stop the practice earlier, type /stop")
    
    await send('\n'.join(map("{} ?".format, sentences)))

    context.user_data['sample'] = sample
    context.user_data['direction'] = direction
    context.user_data['index'] = 0

    return 0

from telegram.ext import ConversationHandler
async def guessing_word(update, context):
    sample = context.user_data['sample']
    direction = context.user_data['direction']
    index = context.user_data['index']
    
    send = make_send(update, context)
    
    given_answer = update.effective_message.text
    if given_answer.lower() in ('/cancel', '/stop', '/finish'):
        context.user_data.clear()
        await send('< Practice Finished >')
        return ConversationHandler.END
    
    answers = [x[1] if direction == 'normal' else x[0] for x in sample[index:index+1]]
    await send('\n'.join(map("→ {}".format, answers)))

    if index + 1 < len(sample):
        context.user_data['index'] += 1
        index = context.user_data['index']
        sentences = [x[0] if direction == 'normal' else x[1] for x in sample[index:index+1]]
        await send('\n'.join(map("{} ?".format, sentences)))
        return
    
    context.user_data.clear()
    await send('< Practice Finished >')
    return ConversationHandler.END

SendSaveInfo = namedtuple('SendSaveInfo', 'chat_id thread_id')
def make_send(update: Update, context: CallbackContext, *, save_info: SendSaveInfo = None, **ckwargs) -> AsyncSend:
    if not save_info:
        save_info = make_send_save_info(update, context)

    async def send(m, **kwargs):
        return await context.bot.send_message(
            text=m,
            chat_id=save_info.chat_id,
            message_thread_id=save_info.thread_id,
            **ckwargs,
            **kwargs)
    return send

def make_send_save_info(update: Update, context: CallbackContext) -> SendSaveInfo:
    return SendSaveInfo(
        chat_id=update.effective_chat.id,
        thread_id=update.effective_message.message_thread_id if update.effective_message and update.effective_message.is_topic_message else None,
    )

async def switchpageflashcard(update, context):
    send = make_send(update, context)

    chat_id = update.effective_chat.id

    try:
        page_name, = context.args
    except:
        with get_connection() as conn:
            my_simple_sql_args = partial(simple_sql_args, connection=conn)
            page_id = get_current_flashcard_page_id(chat_id=chat_id, connection=conn)
            page_name, = only_one(my_simple_sql_args('select name from flashcardpage where rowid=?', (page_id, )))

            return await send(f"Usage: /switchpageflashcard page_name\nCurrent page: {page_name}")

    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        # 1. Remove current page, if any
        conn.execute("update flashcardpage set current=0 where chat_id=?", (chat_id,))
        
        # 2. Create or Update target page as current
        db_page_name, = conn.execute("select name from flashcardpage where name=? and chat_id=?", (page_name, chat_id)).fetchone() or (None,)
        if db_page_name is None:
            conn.execute("insert into flashcardpage(chat_id, name, current) values (?,?,1)", (chat_id, page_name))
        else:
            conn.execute("update flashcardpage set current=1 where chat_id=? and name=?", (chat_id, page_name))
        conn.execute("end transaction")

    if is_group := (update.effective_chat.id != update.effective_user.id):
        await send(f"The current flashcard page is now {page_name!r} (for this group)")
    else:
        await send(f"Your current flashcard page is now {page_name!r}")

class flashcard:
    Current = object()

    def parse_and_add(value, *, chat_id, page_name:str | Current, connection):
        conn = connection  

        if len(parsed := value.split()) == 2:
            sentence, translation = map(str.strip, parsed)
        elif len(parsed := value.split('=', maxsplit=1)) == 2:
            sentence, translation = map(str.strip, parsed)
        elif len(parsed := value.split('/', maxsplit=1)) == 2:
            sentence, translation = map(str.strip, parsed)
        else:
            raise UserError('Invalid flashcard')
                
        page_id = (get_current_flashcard_page_id(chat_id=chat_id) if page_name is flashcard.Current else
                   get_named_flashcard_page_id(page_name=page_name, chat_id=chat_id, connection=conn))
        
        save_flashcard(sentence, translation, page_id=page_id, connection=conn)

    def print_current_flashcards(*, chat_id, connection, select=None):
        page_id = get_current_flashcard_page_id(chat_id=chat_id)
        return flashcard.print_page_flashcards_from_id(chat_id=chat_id, page_id=page_id, connection=connection, select=select)
    
    def print_page_flashcards(chat_id, page_name, connection, select=None):
        page_id = get_named_flashcard_page_id(page_name=page_name, chat_id=chat_id, connection=connection)
        return flashcard.print_page_flashcards_from_id(chat_id=chat_id, page_id=page_id, connection=connection, select=select)

    def print_page_flashcards_from_id(chat_id, page_id, connection, select=None):
        my_simple_sql = partial(simple_sql, connection=connection)

        results = my_simple_sql((
            ''' select flashcard.sentence, flashcard.translation from flashcard inner join flashcardpage on flashcard.page_id=flashcardpage.rowid
                where flashcardpage.chat_id=? and flashcardpage.rowid=?
            ''', (chat_id, page_id)))

        return ('\n\n' if not select else '\n').join(
            f"{sentence}\n→ {translation}" if not select else f"- {sentence}" if select == 'first' else f'- {translation}' if select == 'second' else raise_error(AssertionError)
            for sentence, translation in results) or '/'
    
    def enumerate_flashcards(*, page_name:str | Current, chat_id, connection, select=None):
        my_simple_sql = partial(simple_sql, connection=connection)

        page_id = flashcard.get_page_id(chat_id=chat_id, page_name=page_name, connection=connection)
        
        results = my_simple_sql((
            ''' select flashcard.sentence, flashcard.translation from flashcard inner join flashcardpage on flashcard.page_id=flashcardpage.rowid
                where flashcardpage.chat_id=? and flashcardpage.rowid=?
            ''', (chat_id, page_id)))
        
        return ('\n\n' if not select else '\n').join(
            f"{n}. {sentence}\n→ {translation}" if not select else f"{n}. {sentence}" if select == 'first' else f"{n}. {translation}" if select == 'second' else raise_error(AssertionError)
            for n, (sentence, translation) in enumerate(results, start=1)) or '/'
    
    def get_page_id(*, chat_id, page_name, connection):
        return (get_current_flashcard_page_id(chat_id=chat_id, connection=connection) if page_name is flashcard.Current else
                get_named_flashcard_page_id(chat_id=chat_id, page_name=page_name, connection=connection))

    def clear_flashcards(chat_id, page_name, connection):
        my_simple_sql = partial(simple_sql_args, connection=connection)

        page_id = flashcard.get_page_id(page_name=page_name, chat_id=chat_id, connection=connection)
        my_simple_sql('''DELETE FROM flashcard where page_id=?''', (page_id, ))

    def delete_in_page(do_all_delete, chat_id, page_name, connection):
        my_simple_sql = partial(simple_sql_args, connection=connection)

        page_id = flashcard.get_page_id(chat_id=chat_id, page_name=page_name, connection=connection)
        rowids = my_simple_sql(''' select rowid from flashcard where page_id=? ''', (page_id, ))

        def delete(i):
            my_simple_sql(''' delete from flashcard where rowid=? ''', (rowids[i][0], ))

        do_all_delete(delete, len(rowids))
        
async def listflashcards(update, context):
    send = make_send(update, context)
    Args = InfiniteEmptyList(context.args)
    chat_id = update.effective_chat.id

    if Args[0].lower() == 'first':
        select = 'first'
    elif Args[0].lower() == 'second':
        select = 'second'
    else:
        select = None

    with get_connection() as conn:
        return await send(flashcard.print_current_flashcards(chat_id=chat_id, select=select, connection=conn))

async def listpageflashcards(update, context):
    send = make_send(update, context)
    Args = InfiniteEmptyList(context.args)
    chat_id = update.effective_chat.id

    results = simple_sql(('select name, current from FlashcardPage where chat_id = ?', (chat_id, )))
    await send('\n'.join("{marker} {name}".format(marker='→' if current else '•', name=name) for name, current in results) or '/')

async def exportflashcards(update, context):
    chat_id = update.effective_chat.id
    query = ('''
             select sentence, translation, flashcardpage.name from flashcard
             inner join flashcardpage on flashcard.page_id=flashcardpage.rowid where flashcardpage.chat_id=?
             ''',
             (chat_id, ))
    
    all_pages = simple_sql(query)

    from collections import defaultdict
    page_dict = defaultdict(list)
    for sentence, translation, page_name in all_pages:
        page_dict[page_name].append((sentence, translation))

    def export_tsv_utf8():
        import io
        file_content_io = io.StringIO()
        import csv
        csv.writer(file_content_io, dialect='excel-tab').writerows(lines)
        return file_content_io.getvalue().encode('utf-8')

    def export_xlsx():
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = '1'
        for sheet_name, lines in page_dict.items():
            if sheet_name not in wb.sheetnames:
                wb.active = wb.create_sheet(title=sheet_name)
            for line in lines:
                wb.active.append(line)
        import io
        bytes_io = io.BytesIO()
        wb.save(bytes_io)
        return bytes_io.getvalue()
    
    #file_content = export_tsv_utf8()
    #extension = 'tsv'
    file_content: bytes = export_xlsx()
    extension: str = 'xlsx'

    await context.bot.send_document(
        update.effective_chat.id,
        file_content,
        filename="flashcards." + extension,
        message_thread_id=update.effective_message.message_thread_id if update.effective_message.is_topic_message else None)

async def export_event(update, context, *, name, datetime_utc):
    from datetime import date, time, datetime, timedelta
    
    file_content_str = EVENT_ICS_TEMPLATE.format(
        dt_created_utc=datetime.now(UTC).replace(tzinfo=None),
        dt_start_utc=datetime_utc,
        dt_end_utc=datetime_utc + timedelta(hours=1),
        name_ical_formatted=name)
    
    file_content: bytes = file_content_str.encode('utf-8')
    
    await context.bot.send_document(
        update.effective_chat.id,
        file_content,
        filename="event.ics",
        message_thread_id=update.effective_message.message_thread_id if update.effective_message.is_topic_message else None)

import zoneinfo
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
UTC = ZoneInfo('UTC')

import functools

class DatetimeText:
    days_english = "monday tuesday wednesday thursday friday saturday sunday".split() 
    days_english_short = "mon tue wed thu fri sat sun".split() 
    days_french = "lundi mardi mercredi jeudi vendredi samedi dimanche".split()
    days_french_short = "lun mar mer jeu ven sam dim".split()
    _days_russian = "понедельник вторник среда четверг пятница суббота воскресенье".split()
    _days_russian_short = "пн вт ср чт пт сб вс".split()
    _days_russian_short_dotted = "пн. вт. ср. чт. пт. сб. вс.".split()
    days_russian_short = _days_russian_short
    days_russian = _days_russian

    class days:
        @staticmethod
        def in_lang(x):
            x = x.upper()
            if x == 'EN':
                return DatetimeText.days_english
            elif x == 'FR':
                return DatetimeText.days_french
            elif x == 'RU':
                return DatetimeText.days_russian
            return DatetimeText.days_english

    class days_short:
        @staticmethod
        def in_lang(x):
            x = x.upper()
            if x == 'EN':
                return DatetimeText.days_english_short
            elif x == 'FR':
                return DatetimeText.days_french_short
            elif x == 'RU':
                return DatetimeText.days_russian_short
            return DatetimeText.days_english_short

    months_english = [
        "january",
        "february",
        "march",
        "april",
        "may",
        "june",
        "july",
        "august",
        "september",
        "october",
        "november",
        "december"
    ]

    months_french = [
        "janvier",
        "février",
        "mars",
        "avril",
        "mai",
        "juin",
        "juillet",
        "août",
        "septembre",
        "octobre",
        "novembre",
        "décembre"
    ]

    month_in_russian = [
        *'январь февраль март апрель май июнь июль август сентябрь октябрь ноябрь декабрь'.split(),
    ]

    class months:
        @staticmethod
        def in_lang(x):
            x = x.upper()
            if x == 'EN':
                return DatetimeText.months_english
            elif x == 'FR':
                return DatetimeText.months_french
            elif x == 'RU':
                return DatetimeText.month_in_russian
            return DatetimeText.months_english

    _other_months_list = month_in_russian

    @classmethod
    def padezh_month(cls, month: int, day: int):
        """ padezh_month(8, 11) -> 11th of August -> августа """
        return cls.month_in_russian[month-1]

    months_value = functools.reduce(dict.__or__, ({
        x: i for i, x in enumerate(months_list, start=1)
    } for months_list in (months_english, months_french, *_other_months_list)))


    @classmethod
    def is_relative_day_keyword(cls, x:str):
        return x.lower() in ("today", "auj", "aujourdhui", "aujourd'hui", "aujourd’hui", "tomorrow", "demain")
    
    @classmethod
    def is_valid_weekday(cls, x:str):
        return cls.parse_valid_weekday(x) is not None

    @classmethod
    def parse_valid_weekday(cls, x:str) -> int:
        """0 == monday"""
        x = x.lower()
        for lang in (cls.days_english, cls.days_french, cls._days_russian, cls._days_russian_short, cls._days_russian_short_dotted):
            if x in lang:
                return lang.index(x)
        return None
    
    @classmethod
    def is_valid_month(cls, x:str):
        return cls.parse_valid_month(x) is not None
    
    @classmethod
    def parse_valid_month(cls, x:str):
        x = x.lower()
        return cls.months_value.get(x, None)
    
    @classmethod
    def to_datetime_range(self, name, *, time=None, reference=None, tz=None):
        from datetime import datetime as Datetime, time as Time
        date, date_end = r = self.to_date_range(name, reference=reference, tz=tz)
        datetime = Datetime.combine(date, time or Time(0,0)).replace(tzinfo=tz)
        return datetime, date_end

    @classmethod
    def to_date_range(self, name, *, reference=None, tz=None) -> tuple[Date, Date]:
        from datetime import datetime, timedelta, date, date as Date
        reference = reference or datetime.now().astimezone(tz).replace(tzinfo=None)
        today = reference.date()
        name = name.lower()

        day = None
        if parsed_tuple := ParseEvents.parse_valid_date(name):
            if len(parsed_tuple) == 3:
                day = date(*map(int, parsed_tuple))
            elif len(parsed_tuple) == 2:
                day = date(today.year, *map(int, parsed_tuple))
            else:
                raise AssertionError
        
        elif match := re.fullmatch("(\d{4})-(\d{2})-(\d{2})", name):
            day = date(*map(int, match.groups()))

        elif (match_eu := re.fullmatch(r"(\d{1,2}) (%s)( (\d{4}))?" % '|'.join(map(re.escape, self.months_value)), name)) or \
           (match_us := re.fullmatch(r"(%s) (\d{1,2})( (\d{4}))?" % '|'.join(map(re.escape, self.months_value)), name)):
            if match := match_eu:
                dstr,mstr,_,ystr = match.groups()
            elif match := match_us:
                mstr,dstr,_,ystr = match.groups()
            if not ystr:
                y = today.year
            else:
                y = int(ystr)
            d = int(dstr)
            m = self.months_value[mstr]
            day = date(y, m, d)

        if day is not None:
            return day, day + timedelta(days=1)
        
        if name in ("today", "auj", "aujourdhui", "aujourd'hui", "aujourd’hui"):
            return today, today + timedelta(days=1)
        
        if name in ("week", "semaine"):
            beg = today
            end = today + timedelta(days=7)
            return beg, end
        
        if name in ("tomorrow", "demain"):
            return today + timedelta(days=1), today + timedelta(days=2)
        
        if name in ('future', 'futur'):
            return today, date.max - timedelta(days=7)
        
        if name in ('past', 'passé'):
            return date.min + timedelta(days=7), today
        
        if name in ('yesterday', 'hier'):
            return today - timedelta(days=1), today
        
        if name in ('ereyesterday', 'avant-hier', 'avanthier'):
            return today - timedelta(days=2), today - timedelta(days=1)
        
        if name in ('overmorrow', 'après-demain', 'apres-demain', 'apresdemain', 'aprèsdemain'):
            return today + timedelta(days=2), today + timedelta(days=3)
        
        if not DatetimeText.is_valid_weekday(name):
            raise UnknownDateError(f"Unknown date {name}")

        i = DatetimeText.parse_valid_weekday(name)

        assert i is not None
        assert 0 <= i < 7 

        the_day = today + timedelta(days=1)
        while the_day.weekday() != i:
            the_day += timedelta(days=1)
        
        beg = the_day
        end = beg + timedelta(days=1)
        return beg, end
    
    @classmethod
    def format_td_T_minus(cls, td:timedelta, *, format='multiple'):
        assert format in ('unit', 'short', 'long', 'multiple')
        from datetime import timedelta
        sign = "-" if td >= timedelta(seconds=0) else "+"
        td = abs(td)
        d,rem = divmod(td, timedelta(days=1))
        m,s = divmod(rem.seconds, 60)
        h,m = divmod(m, 60)
        days, hours, minutes = (timedelta(**{x:1}) for x in 'days hours minutes'.split())
        if format == 'short':
            return (f"D{sign}{d}" if td > 5 * days else 
                    f"H{sign}{h + d*24}" if td > 3 * hours else 
                    f"M{sign}{m + h*60 + d*24}" if td > 10 * minutes else 
                    f"S{sign}{s + m*60 + h*60 + d*24}")
        if format == 'long':
            return (f"D{sign}{d}" if td > 10 * days else 
                    f"H{sign}{h + d*24}" if td > 10 * hours else 
                    f"M{sign}{m + h*60 + d*24}" if td > 60 * minutes else 
                    f"S{sign}{s + m*60 + h*60 + d*24}")
        if format == 'multiple':
            return (f"D{sign}{d}, H{sign}{h}" if d else
                    f"H{sign}{h}, M{sign}{m}" if h else
                    f"M{sign}{m}, S{sign}{s}" if m else
                    f"S{sign}{s}")
        
        return (                    
            "D{}{}".format(sign, d) if d else
            "H{}{}".format(sign, h) if h else 
            "M{}{}".format(sign, m) if m else 
            "S{}{}".format(sign, s)
        )


from collections import namedtuple
from datetime import date as Date, time as Time, datetime as Datetime, timedelta as Timedelta

class ParsedEventMiddleNoName(NamedTuple):
    date: str
    time: Optional[Time]
    day_of_week: str
    relative_day_keyword: str
    timezone: Optional[str]

class ParsedEventMiddle(NamedTuple):
    date: str
    time: Optional[Time]
    name: str 
    day_of_week: str
    relative_day_keyword: str
    timezone: Optional[str]

    @staticmethod
    def from_no_name(event: ParsedEventMiddleNoName, name:str):
        return ParsedEventMiddle(**event._asdict(), name=name)

class ParsedScheduleMiddle(NamedTuple):
    events: list[ParsedEventMiddle]
    each_activateds: list[bool]

class ParsedEventFinal(NamedTuple):
    date_str: str
    time: Optional[Time]
    name: str 
    date: Date
    date_end: Date 
    datetime: Datetime 
    datetime_utc: Datetime 
    tz: ZoneInfo
    tz_explicit: bool

@dataclass
class ParsedEventDate:
    day_of_week: str
    date_str: str
    relative_day_keyword: str

def parse_event_date(args) -> tuple[ParsedEventDate, list]:
    """
    ['Something', 'A', 'B', 'C'] -> 'Something', ['A', 'B', 'C']  # n = 1
    ['25', 'November', 'A', 'B', 'C'] -> '25 November', ['A', 'B', 'C']  # n = 2
    ['25', 'November', '2023', 'A', 'B', 'C'] -> '25 November 2023', ['A', 'B', 'C']  # n = 3
    ['25.11', 'A', 'B', 'C'] -> '25.11', ['A', 'B', 'C']  # n = 3
    ['25.11.2025', 'A', 'B', 'C'] -> '25.11.2025', ['A', 'B', 'C']  # n = 3
    ['today', 'friday', 'A', 'B', 'C'] -> ['A', 'B', 'C'], day_of_week
    """
    Args = GetOrEmpty(args)

    if DatetimeText.is_relative_day_keyword(Args[0]): # Examples: today, tomorrow
        relative_day_keyword = Args[0]
        args = args[1:]
        Args = GetOrEmpty(args)
    else:
        relative_day_keyword = '' # we don't use any relative day indicator

    if DatetimeText.is_valid_weekday(Args[0]): # Examples: Monday
        day_of_week = Args[0]
        args = args[1:]
        Args = GetOrEmpty(args)
    else:
        day_of_week = '' # we don't know the day of week

    date_str = None
    if ParseEvents.is_valid_date(Args[0]): # Example: 2020-12-31
        n = 1 # the first token is the date
    elif Args[0].isdecimal() and Args[1].lower() in DatetimeText.months_value \
    or Args[1].isdecimal() and Args[0].lower() in DatetimeText.months_value: # Example: 25 November
        if Args[2].isdecimal() and len(Args[2]) == 4: # Example: 2012
            n = 3 # has year
        else:
            n = 2 # no year
    else:
        if relative_day_keyword or day_of_week:
            # if only day_of_week: ir will be enough to know the date
            # if only relative_day_keyword: it will be enough to know the date
            # if both: we are based on the relative_day_keyword (today) and will later check that it corresponds to a "Friday" for example 
            date_str = relative_day_keyword or day_of_week
            n = 0
        else:
            n = 1 # the first token will be the date
    
    if date_str is None:
        date_str = ' '.join(args[:n])

    return ParsedEventDate(
        day_of_week=day_of_week,
        relative_day_keyword=relative_day_keyword,
        date_str=date_str,
    ), args[n:]

class ParseEvents:

    class TimeWithTz(Exception):
        time: Time 
        rest: list[str]
        timezone: str

        def __init__(self, time, rest, timezone):
            self.time = time
            self.rest = rest
            self.timezone = timezone

    @classmethod
    def is_valid_date(cls, value:str):
        return cls.parse_valid_date(value) is not None

    @classmethod
    def parse_valid_date(cls, value:str) -> tuple:
        import re
        ReDateRu = re.compile('(\d{2})[.](\d{2})')
        ReDateRuYear = re.compile('(\d{2})[.](\d{2})[.](\d{4})')
        ReBeYear = re.compile('(\d{2})[/](\d{2})[/](\d{4})')
        if regexes := [(R, M) for R in (ReDateRu, ReDateRuYear, ReBeYear) if (M := R.fullmatch(value))]:
            matching_re, match = only_one(regexes, ValueError(f"ProgrammingError: Multiple regex match the string {value}"))
            if ReDateRu is matching_re:
                a,b = match.groups()
                return b,a
            elif ReDateRuYear is matching_re:
                a,b,c = match.groups()
                return c,b,a
            elif ReBeYear is matching_re:
                a,b,c = match.groups()
                return c,b,a
            else:
                raise AssertionError("Problem")
        else:
            return None

    @classmethod
    def parse_time(cls, args: list) -> tuple[Optional[Time], list]:
        Args = InfiniteEmptyList(args)
        if match := re.compile('(\\d{1,2})[:hH](\\d{2})?').fullmatch(Args[0]):
            hours, minutes = match.group(1), match.group(2)
            time = Time(int(hours), int(minutes or '0'))
            args = args[1:]
        elif match := re.compile('(\\d{1,2})[:hH](\\d{2})?[:]([A-Za-z_]+|[A-Za-z_]+[/][A-Za-z_]+)').fullmatch(Args[0]):
            hours, minutes = match.group(1), match.group(2)
            time = Time(int(hours), int(minutes or '0'))
            timezone: str = match.group(3)
            args = args[1:]
            raise ParseEvents.TimeWithTz(time=time, rest=args, timezone=timezone)
        else:
            time = None
        return time, args
    
    @classmethod
    def parse_event_timed(cls, args: list) -> tuple[ParsedEventMiddleNoName, list]:
        date: str
        rest: list
        time: Optional[Time]
        
        parsed_event_date: ParsedEventDate
        parsed_event_date, rest = parse_event_date(args)
        
        day_of_week = parsed_event_date.day_of_week 
        date = parsed_event_date.date_str
        relative_day_keyword = parsed_event_date.relative_day_keyword

        try:
            time, rest = cls.parse_time(rest)
            timezone = None
        except ParseEvents.TimeWithTz as ret:
            time, rest = ret.time, ret.rest
            timezone = ret.timezone

        return ParsedEventMiddleNoName(date=date, time=time, day_of_week=day_of_week, relative_day_keyword=relative_day_keyword, timezone=timezone), rest
    
    @classmethod
    def parse_event(cls, args) -> ParsedEventMiddle:
        event_no_name: ParsedEventMiddleNoName
        event_no_name, rest = cls.parse_event_timed(args)
        return ParsedEventMiddle.from_no_name(event_no_name, name=" ".join(rest))

    @classmethod
    def parse_schedule(cls, args, *, tz) -> list[ParsedEventMiddleNoName]:
        default_tz = tz
        out: list[ParsedEventMiddleNoName] = []
        it = args
        while it:
            before = list(it)

            each_activated = False
            It = InfiniteEmptyList(it)
            if It[0].lower() in ('each', 'every', 'chaque', 'le'):
                each_activated = True
                each_activated_by = It[0]
                it = it[1:]
            elif tuple(map(str.lower, It[0:2])) in [('tous', 'les')]:
                each_activated = True
                each_activated_by = ' '.join(It[0:2])
                it = it[2:]
                
            event, it = cls.parse_event_timed(it)
            tz = event.timezone or default_tz

            try:
                DatetimeText.to_date_range(event.date, tz=tz)
            except UnknownDateError:
                it = before
                break

            if each_activated:
                if not event.day_of_week:
                    raise UserError(f"The keyword {each_activated_by!r} has to be applied on a day of the week")
                date_obj, date_obj_end = DatetimeText.to_date_range(event.date, tz=tz)

                It = InfiniteEmptyList(it)
                n = 4
                if (It[0].lower() in ("for", "pour")
                    and It[1].isdecimal()
                    and It[2].lower() in ("times", "fois")):
                    n = int(It[1])
                    it = it[3:]
                elif (It[0].lower() == 'n'
                    and It[1] == '='
                    and It[2].isdecimal()):
                    n = int(It[2])
                    it = it[3:]

                for i in range(n):
                    out.append(event._replace(date=str(date_obj + timedelta(weeks=i))))
            else:
                out.append(event)

        if sum(event.time is not None for event in out) == 1:
            the_time = next(event.time for event in out if event.time is not None)
            out = [event._replace(time=the_time) for event in out]

        name_fmt = ' '.join(it)

        bracket_extension = split_bracket_comma_format(name_fmt)

        to_return = []
        for event, n in zip(out, irange(1, len(out))):
            current_format = get_modulo(bracket_extension, n-1)
            name = safe_format(current_format, n=n)
            to_return.append( ParsedEventMiddle.from_no_name(event, name=name) )

        return to_return

def get_modulo(L, i):
    return L[i % len(L)]

def safe_format(fmt, **kwargs):
    """
    safe_format("Hello {n}", n=5) -> "Hello 5"
    safe_format("Hello {a}", n=5) -> "Hello {a}"
    """
    Re = re.compile(re.escape('{') + '[a-zA-Z_][a-zA-Z_[0-9]*' + re.escape('}'))
    return Re.sub(lambda m: str(kwargs.get(m.group(0)[1:-1], m.group(0))), fmt)

def split_bracket_comma_format(fmt):
    """
    split_bracket_comma_format("Hello {World, Life}") -> ["Hello World", "Hello Life"]
    split_bracket_comma_format("Hello World") -> ["Hello World"]
    split_bracket_comma_format("Hello {A,B} {C,D}") -> ["Hello A C", "Hello B D"]
    split_bracket_comma_format("Hello {A,B,C} {D,E}") -> ["Hello A D", "Hello B E", "Hello C D"]  # looping
    split_bracket_comma_format("Hello {A}") -> ["Hello {A}"]
    """
    Re = re.compile(re.escape('{') + '(.*?)' + re.escape('}'))
    if m := Re.findall(fmt):
        all_bits = []
        for p in m:
            bits = [x.strip() for x in p.split(',')]
            all_bits.append(bits)
        
        def make(i):
            counter = itertools.count(0)
            def sub(p):
                bits = all_bits[next(counter)]
                if len(bits) > 1:
                    return get_modulo(bits, i)
                else:
                    return '{' + (bits[0] if bits else '') + '}'
            return sub

        return [Re.sub(make(i), fmt) for i in range(max(len(b) for b in all_bits))]
    return [fmt]

def raise_error(error):
    raise error

def induce_my_timezone_from_update(update):
    return induce_my_timezone(user_id=update.effective_user.id, chat_id=update.effective_chat.id)

def induce_my_timezone(*, user_id, chat_id):
    if tz := get_my_timezone(user_id):
        return tz
    elif tzs := read_settings("event.timezones", id=chat_id, settings_type='chat'):
        if len(tzs) == 1:
            return tzs[0]
    raise UserError(
        "I don't know your timezone and the chat doesn't have one and only one timezone.\n"
        "\n"
        "Set your timezone by typing...\n"
        "- This: /mytimezone TIMEZONE\n"
        "- Example: /mytimezone Europe/Brussels\n"
        "- Example: /mytimezone America/Los_Angeles\n"
        "\n"
        "Or set the chat timezone by typing...\n"
        "- This: /chatsettings event.timezones TIMEZONE\n"
        "- Example: /chatsettings event.timezones Europe/Brussels\n")

def parse_datetime_point(update, context, when_infos=None, what_infos=None, has_inline_kargs=False, required_time=False) -> ParsedEventFinal:
    from datetime import datetime as Datetime, time as Time, date as Date, timedelta
    read_chat_settings = make_read_chat_settings(update, context)
    
    name = ''
    date_str = None
    if context.args and not has_inline_kargs:
        date_str, time, name, day_of_week, relative_day_keyword, timezone_event_str = ParseEvents.parse_event(context.args)
    if what_infos:
        name = name or what_infos
    if when_infos:
        if date_str:
            raise UserError("Multiple When specified")
        date_str, time, name_from_when_part, day_of_week, relative_day_keyword, timezone_event_str = ParseEvents.parse_event(when_infos.split())
        if name_from_when_part:
            raise UserError("Too much infos in the When part")
    if date_str is None:
        raise UserError("Must specify an event with date")

    if timezone_event_str:
        tz = ZoneInfoOrAlias(timezone_event_str, chat_id=update.effective_chat.id)
        # timezone explicit: no need to be in chat.event.timezones
        tz_explicit = True
    else:
        tz = induce_my_timezone_from_update(update)
        # timezone implicit: need to be in chat.event.timezones to avoid confusion
        tz_explicit = False

    if required_time:
        if time is None:
            raise UserError("Time must be specified (policy of the group)")

    date, date_end = DatetimeText.to_date_range(date_str, tz=tz)
    datetime = Datetime.combine(date, time or Time(0,0)).replace(tzinfo=tz)
    datetime_utc = datetime.astimezone(UTC)
    
    date_str: str
    time: Optional[Time]
    name: str 
    date: Date
    date_end: Date 
    datetime: Datetime 
    datetime_utc: Datetime 
    tz: ZoneInfo
    tz_explicit: bool
    Loc = locals()

    if not tz_explicit:
        chat_timezones = read_chat_settings("event.timezones")
        check_tz_in_chat(chat_timezones=chat_timezones, tz=tz)

    if relative_day_keyword:
        rdate, rdate_end = DatetimeText.to_date_range(relative_day_keyword, tz=tz)
        if rdate != date:
            raise UserError(f"{date_str!r} is not {relative_day_keyword!r}")

    if day_of_week:
        if not is_correct_day_of_week(date, day_of_week):
            raise UserError(f"{date_str!r} is not a {day_of_week!r}")
    
    return ParsedEventFinal(**{x: Loc[x] for x in ParsedEventFinal._fields})

def parse_datetime_schedule(*, tz, args) -> list[ParsedEventFinal]:
    # monday 15h tuesday 16h Party -> [[monday 15h Party], [tuesday 16h Party]]
    # monday tuesday 15h Party -> [[monday 15h Party], [tuesday 15h Party]]
    # monday 15h tuesday 16h Party {n} -> [[monday 15h Party 1], [tuesday 16h Party 2]]
    # monday 15h tuesday 16h Name is {Party, Sleep} -> [[monday 15h Name is Party], [tuesday 16h Name is Sleep]]
    
    commands_i = [i for i in range(len(args)) if args[i] == '//']

    if commands_i:
        intervals = [[0, None]]
        for i in commands_i:
            intervals[-1][1] = i
            intervals.append([i+1, None])
        intervals[-1][1] = len(args)

        all_events = []
        for a,b in intervals:
            all_events.extend(parse_datetime_schedule(tz=tz, args=args[a:b]))
        return all_events

    out = []
    event: ParsedEventMiddle
    date: datetime
    for event in ParseEvents.parse_schedule(args, tz=tz):
        time, name = event.time, event.name
        date, date_end = DatetimeText.to_date_range(event.date, tz=tz)
        datetime = Datetime.combine(date, time or Time(0, 0)).replace(tzinfo=tz)
        datetime_utc = datetime.astimezone(UTC)

        if day_of_week := event.day_of_week:
            if not is_correct_day_of_week(date, day_of_week):
                raise UserError(f"{event.date!r} is not a {day_of_week!r}")
        
        tz_explicit = event.timezone is not None
        final = ParsedEventFinal(date_str=event.date, time=time, name=name, date=date, date_end=date_end, datetime=datetime, datetime_utc=datetime_utc, tz=tz, tz_explicit=tz_explicit)
        out.append(final)

    return out 

def is_correct_day_of_week(date, day_of_week):
    return date.weekday() == DatetimeText.parse_valid_weekday(day_of_week)
    # return date.weekday() == (DatetimeText.days_english + DatetimeText.days_french).index(day_of_week.lower()) % 7

async def macro_event_follow(update, context):
    send = make_send(update, context)
    number_re = re.compile('[-]?\\d+')

    Args = InfiniteEmptyList(context.args)
    if not Args:
        return await event_action_follow(update, context)
    
    elif number_re.fullmatch(Args[0]):
        context.args = tuple(Args)
        return await event_action_follow(update, context)
    
    elif Args[0].lower() in ('delete', 'del'):
        if Args[1].lower() in ('follower', 'followers'):
            context.args = tuple(Args[2:])
            return await deleventacceptfollow(update, context)
        
        elif number_re.fullmatch(Args[1]):
            context.args = tuple(Args[1:])
            return await deleventfollow(update, context)
        
        elif Args[1].lower() in ('sub', 'subscription', 'subscriptions'):
            context.args = tuple(Args[2:])
            return await deleventfollow(update, context)
        
        return await send('/eventfollow delete (follower|subscription)')
        
    elif Args[0].lower() in ('accept', ):
        context.args = tuple(Args[1:])
        return await eventacceptfollow(update, context)
    
    elif Args[0].lower() in ('rename', ):
        if Args[1].lower() in ('follower', ):
            context.args = tuple(Args[2:])
            return await renameeventacceptfollow(update, context)
        
        elif Args[1].lower() in ('sub', 'subscription'):
            context.args = tuple(Args[2:])
            return await renameeventfollow(update, context)

        return await send('/eventfollow delete (follower|subscription')
    
    elif Args[0].lower() in ('list', ):
        if Args[1].lower() in ('followers', 'follower'):
            pass 
        elif Args[1].lower() in ('sub', 'subscription', 'subscriptions'):
            pass
        raise UserError("Not implemented yet")
      
    return await send("Wrong parameters")

async def event_action_follow(update, context):
    send = make_send(update, context)

    chat_id = update.effective_chat.id
    thread_id = make_send_save_info(update, context).thread_id

    if not context.args:
        return await send(
            f'Your chat id: {chat_id}\n\n'
            f'Use it so that other people can follow you!\n\n'
            f'To follow this chat:\n  /eventfollow {chat_id}\n\n'
            f'Usage: /eventfollow chat_id [other_chat_name]')
    
    target_chat_id = str(int(context.args[0]))
    my_relation_name = ' '.join(context.args[1:])

    # a follows b
    # a_chat_id = a.chat_id (immutable)
    # b_chat_id = b.chat_id (immutable)
    # a_name = the name a gave to b
    # b_name = the name b gave to a
    # a_thread_id = the thread id where a wants to receive events
    simple_sql(('insert into EventFollowPending(a_chat_id, b_chat_id, a_name, b_name, a_thread_id) VALUES (?,?,?,?,?)', (
        str(chat_id),
        str(target_chat_id),
        my_relation_name or str(target_chat_id),
        str(chat_id),
        str(thread_id) if thread_id is not None else '')))

    if True:  # do_unless_setting_off(the_target_chat . event.follow.notify_my_followers):
        await context.bot.send_message(
            text=f'Event follow request received!\n\nTo accept, type:\n/eventfollow accept {chat_id}\n\nOr:\n/eventfollow accept {chat_id} Custom Name',
            chat_id=target_chat_id)
            # message_thread_id=target_thread_id # read target_chat's "bot channel/admin channel" setting (ie. where they receive the follow requests)

    await send(f'Pending follow request sent to ' + (f'{target_chat_id}' if not my_relation_name else f'{target_chat_id} ({my_relation_name})'))

    # if receiving chat has the setting "automatically accept event following request"
    #   do it

async def eventacceptfollow(update, context):
    send = make_send(update, context)

    chat_id = update.effective_chat.id

    if not context.args:
        followers_pending = simple_sql(('select a_chat_id from EventFollowPending where b_chat_id = ?', (str(chat_id), )))
        return await send('No chats want to be your follower, keep rolling!' if not followers_pending else
            'These chats want to be your follower:\n{}'.format('\n'.join(map("-> {}".format, (
                str(x) for x, in followers_pending
            )))))

    source_chat_id = str(int(context.args[0]))
    my_relation_name = ' '.join(context.args[1:])

    with sqlite3.connect("db.sqlite") as conn:
        my_simple_sql = partial(simple_sql, connection=conn)

        if not (data := my_simple_sql(('select rowid, a_name, b_name, a_thread_id from EventFollowPending where a_chat_id = ? and b_chat_id = ?', (str(source_chat_id), str(chat_id))))):
            return await send(f"Cannot be followed by this chat ({source_chat_id}) because it didn't send a request")
        
        _, a_name, b_name, a_thread_id = data[0]
        
        my_simple_sql(('delete from EventFollowPending where a_chat_id = ? and b_chat_id = ?', (str(source_chat_id), str(chat_id))))

        my_simple_sql(('insert into EventFollow(a_chat_id, b_chat_id, a_name, b_name, a_thread_id) VALUES (?, ?, ?, ?, ?)', (
            str(source_chat_id),
            str(chat_id),
            a_name,
            my_relation_name or b_name,
            a_thread_id)))
        
    # todo: send them some notif
    await send(
        'You are now followed by this chat{}!'.format(" (that you named {})".format(my_relation_name) if my_relation_name else '') + " " +
        'Every event you add will be forwarded to them.' +
        "\n\n" +
        'To see and manage all your followers, see:\n/eventfollow list followers')

async def send_you_are_following_these_chats(update, context):
    send = make_send(update, context)

    chat_id = update.effective_chat.id

    followings = simple_sql(('select b_chat_id, a_name from EventFollow where a_chat_id = ?', (str(chat_id), )))
    await send('You are not following any chats' if not followings else
        'You are following these chats:\n{}'.format('\n'.join(map("-> {}".format, (
            f"{x} ({y})" if x != y else str(x) for x, y in followings
        )))))

async def send_these_chats_are_following_you(update, context):
    send = make_send(update, context)

    chat_id = update.effective_chat.id

    followers = simple_sql(('select a_chat_id, b_name from EventFollow where b_chat_id = ?', (str(chat_id), )))
    await send('No chats is following you' if not followers else
        'These chats are following you:\n{}'.format('\n'.join(map("-> {}".format, (
            f"{x} ({y})" if x != y else str(x) for x, y in followers
        )))))

async def deleventfollow(update, context):
    send = make_send(update, context)

    chat_id = update.effective_chat.id

    if not context.args:
        await send_you_are_following_these_chats(update, context)
        return await send('Usage: /eventfollow delete [chat_id]')

    target_chat_id = str(int(context.args[0]))

    with sqlite3.connect("db.sqlite") as conn:
        my_simple_sql = partial(simple_sql, connection=conn)
        my_simple_sql(('delete from EventFollowPending where a_chat_id = ? and b_chat_id = ?', (str(chat_id), str(target_chat_id))))
        my_simple_sql(('delete from EventFollow where a_chat_id = ? and b_chat_id = ?', (str(chat_id), str(target_chat_id))))

    return await send("Done! You don't follow this chat anymore")

async def deleventacceptfollow(update, context):
    send = make_send(update, context)

    chat_id = update.effective_chat.id

    if not context.args:
        await send_these_chats_are_following_you(update, context)
        return await send('Usage: /eventfollow delete follower [chat_id]')

    target_chat_id = str(int(context.args[0]))

    with sqlite3.connect("db.sqlite") as conn:
        my_simple_sql = partial(simple_sql, connection=conn)
        my_simple_sql(('delete from EventFollowPending where a_chat_id = ? and b_chat_id = ?', (str(target_chat_id), str(chat_id))))
        my_simple_sql(('delete from EventFollow where a_chat_id = ? and b_chat_id = ?', (str(target_chat_id), str(chat_id))))

    return await send("Done! This chat doesn't follow you anymore")

async def eventanyfollowrename(update, context, *, direction: Literal['follow', 'accept']):
    send = make_send(update, context)

    chat_id = update.effective_chat.id

    try:
        target_chat_id = str(int(context.args[0]))
        my_relation_name = ' '.join(context.args[1:])
    except IndexError:
        listing = {'follow': send_you_are_following_these_chats, 'accept': send_these_chats_are_following_you}[direction]
        await listing(update, context)
        return await send("Usage: /{command} chat_id new name".format(command={'follow': 'eventfollow follower', 'accept': 'eventfollow rename subscription'}[direction]))

    if direction == 'follow':
        base_query = 'update %s set a_name = ? where a_chat_id = ? and b_chat_id = ?'
    elif direction == 'accept':
        base_query = 'update %s set b_name = ? where b_chat_id = ? and a_chat_id = ?'
    else:
        raise AssertionError

    with sqlite3.connect("db.sqlite") as conn:
        my_simple_sql = partial(simple_sql, connection=conn)
        conn.execute('begin transaction')
        for table_name in ('EventFollow', 'EventFollowPending'):
            my_simple_sql((base_query % table_name, (my_relation_name, str(chat_id), str(target_chat_id))))
        conn.execute('end transaction')

    return await send({'follow': 'You now follow the chat {} as {!r}', 'accept': 'The chat following you {}, you call it {!r}'}[direction].format(target_chat_id, my_relation_name))

renameeventfollow = partial(eventanyfollowrename, direction='follow')
renameeventacceptfollow = partial(eventanyfollowrename, direction='accept')

def irange(a, b=None):
    if b is None:
        return irange(1, a)
    return range(a, b+1)

async def addschedule(update: Update, context: ContextTypes.DEFAULT_TYPE):
    send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)

    reply = update_get_reply(update)

    if not context.args:
        if reply:
            infos_event = {}
        else:
            return await send("Usage: /addschedule datetime+ name")
    else:
        infos_event = {}

    if not infos_event.get('link'):
        infos_event |= add_event_enrich_reply_with_link(update, context, reply=reply)

    chat_timezones = read_chat_settings("event.timezones")

    # Later commit: do_event_admin_check('add', setting=read_chat_settings('event.admins'), user_id=update.effective_user.id)

    source_user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    tz = induce_my_timezone_from_update(update)

    events: list[ParsedEventFinal] = parse_datetime_schedule(tz=tz, args=context.args)

    location_autocomplete = do_if_setting_on(read_chat_settings('event.location.autocomplete'))

    for event in events:
        if not do_if_setting_on(read_chat_settings('event.location.autocomplete')):
            name = update_name_using_locations(event.name, chat_id=chat_id)
        else:
            name = event.name

        new_event_id = add_event_to_db(name=name, chat_id=chat_id, source_user_id=source_user_id, datetime_utc=event.datetime_utc)

        if infos_event and infos_event.get('link'):
            simple_sql((''' insert into EventLinkAttr(event_id, link) VALUES (?,?)''', (new_event_id, infos_event['link'])))

        if not location_autocomplete:
            implicit_thereis(what=name, chat_id=chat_id)

    return await send(f"{len(events)} event(s) added")

def check_tz_in_chat(*, tz, chat_timezones):
    if chat_timezones and tz and tz not in chat_timezones:
        raise UserError('\n'.join([
            'Your timezone is not in chat timezones, this can be confusing, change your timezone or add your timezone to the chat timezones.',
            '- Your timezone: {tz}'.format(tz=tz),
            '- Chat timezones: {chat_timezone_str}'.format(chat_timezone_str=", ".join(map(str, chat_timezones))),
        ]))

def add_event_to_db(*, datetime_utc, name, chat_id, source_user_id) -> 'id':
    with sqlite3.connect('db.sqlite') as conn:
        cursor = conn.cursor()

        strftime = DatetimeDbSerializer.strftime

        cursor.execute("INSERT INTO Events(date, name, chat_id, source_user_id) VALUES (?,?,?,?)", (strftime(datetime_utc), name, chat_id, source_user_id))

        return cursor.lastrowid

class InteractiveAddEvent:
    @staticmethod
    async def ask_when(update, context):
        send = make_send(update, context)
        await send("When is the event ?\n\nExamples:\n- Today\n- Tomorrow\n- Sunday\n- 25.11\n- 31.12.2000")
        return 'ask-what-or-time'
    
    @staticmethod
    async def ask_what_or_time(update, context):
        send = make_send(update, context)

        when = update.message.text

        event: ParsedEventFinal = parse_datetime_point(update, context, when_infos=when, what_infos='')
        # no error: ok

        context.user_data['when'] = when

        if event.time:
            return await InteractiveAddEvent.ask_what(update, context)
        else:
            return await InteractiveAddEvent.ask_time(update, context)
        
    @staticmethod
    async def ask_time(update, context):
        send = make_send(update, context)

        await send('What is the time of the event ?\n\nExamples:\n- 8h\n- 16h\n- 20:15\n- /midnight\n- /empty')

        context.user_data['did_ask_time'] = 'on'

        return 'ask-what'
    @staticmethod
    async def ask_what_empty(update, context):
        if 'did_ask_time' in context.user_data:
            context.user_data['time'] = ''
        else:
            pass

        return await InteractiveAddEvent.ask_what(update, context)
    
    @staticmethod
    async def ask_what(update, context):
        send = make_send(update, context)
        if 'did_ask_time' in context.user_data:
            if 'time' in context.user_data:
                time = context.user_data['time']
            else:
                time = update.message.text
            if time.strip():
                context.user_data['when'] += ' ' + time
            else:
                pass # user_data.when is perfect
        else:
            pass # user_data.when is perfect

        parse_datetime_point(update, context, when_infos=context.user_data['when'], what_infos='')
        # no error: ok

        await send("What is the event about ?\nThe name of the event.\n\nExamples:\n- Party\n- /empty")
        return 'ask-where'

    @staticmethod
    async def ask_where(update, context):
        send = make_send(update, context)
        what = update.message.text
        context.user_data['what'] = what

        return await InteractiveAddEvent.really_ask_where(update, context)
    
    @staticmethod
    async def ask_where_empty(update, context):
        send = make_send(update, context)
        what = ''
        context.user_data['what'] = what

        return await InteractiveAddEvent.really_ask_where(update, context)

    @staticmethod
    async def really_ask_where(update, context):
        send = make_send(update, context)

        await send("Where is the event ?\n\nExamples:\n- My house\n- Miami Beach (123 Ocean Drive)\n- /skip\n- /empty")

        return 'ask-confirm'
    
    @staticmethod
    async def ask_confirm(update, context):
        where = update.message.text
        return await InteractiveAddEvent.continue_ask_confirm(update, context, where=where)
    
    @staticmethod
    async def ask_confirm_empty(update, context):
        where = ''
        return await InteractiveAddEvent.continue_ask_confirm(update, context, where=where)
    
    @staticmethod
    async def continue_ask_confirm(update, context, where):
        send = make_send(update, context)

        context.user_data['where'] = where
        when = context.user_data['when']
        what = context.user_data['what']
        await send(f"Do you want to add this event ?\nWhen: {when}\nWhat: {what}\nWhere: {where}\n")
        return 'do-add-event'
    
    @staticmethod
    async def do_add_event(update, context):
        send = make_send(update, context)

        when = context.user_data['when']
        what = context.user_data['what']
        where = context.user_data['where']
        context.user_data.clear()

        if update.message.text.lower() in ("no", "n"):
            await send("Event not added.\n\n/addevent can be however applied on the last message.")
            return ConversationHandler.END

        await InteractiveAddEvent.do_all_add_event(update, context, what=what, when=when, where=where)
        return ConversationHandler.END
    
    @staticmethod
    async def do_all_add_event(update, context, *, what, when, where):
        read_chat_settings = make_read_chat_settings(update, context)

        source_user_id = update.message.from_user.id
        chat_id = update.effective_chat.id

        real_what = (what if not where else what + ' @ ' + where).strip()
        
        required_time = do_if_setting_on(read_chat_settings('event.addevent.required_time'))
        date_str, time, name, date, date_end, datetime, datetime_utc, tz, tz_explicit = parse_datetime_point(update, context, when_infos=when, what_infos=real_what, required_time=required_time)
        
        add_event_to_db(datetime_utc=datetime_utc, name=name, chat_id=chat_id, source_user_id=source_user_id)

        chat_timezones = read_chat_settings("event.timezones")
        
        await post_event(update, context, link='', name=name, datetime=datetime, time=time, date_str=date_str, chat_timezones=chat_timezones, tz=tz, tz_explicit=tz_explicit, chat_id=chat_id, datetime_utc=datetime_utc)

def do_event_admin_check(type: Literal['add', 'del', 'edit', 'list'], *, setting, user_id):
    assert type in ('add', 'del', 'edit', 'list')
    if setting:
        # do an admin check
        if user_id in (admin_ids := set(map(lambda x:int(x.user_id), event_admins := setting))):
            if type in (event_admin := only_one(filter(lambda x:x.user_id == user_id, event_admins))).permissions:
                pass
            else:
                raise EventAdminError
        elif 0 in admin_ids:
            if type in (event_admin := only_one(filter(lambda x:x.user_id == 0, event_admins))).permissions:
                pass
            else:
                raise EventAdminError
        else:
            raise EventAdminError
    else:
        # no setting set = everyone is admin
        pass

def tg_url_id_from_chat_id(chat_id):
    """ chat_id negative and big into normal """
    return (abs(chat_id) - 10**12 if chat_id < 0 and abs(chat_id) - 10**12 > 0 else 
            abs(chat_id) if chat_id < 0 else 
            chat_id)

import sqlite3
async def add_event(update: Update, context: CallbackContext):
    send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)
    read_my_settings = make_read_my_settings(update, context)

    infos_event = {}
    if (reply := update_get_reply(update)) and reply.text:
        try:
            infos_event |= addevent_analyse(update, context)
        except EventAnalyseError as e:
            if context.args:
                pass
            else:
                raise e
    elif not context.args:
        return await send("Usage: /addevent date name\nUsage: /addevent date hour name\nInteractive version: /iaddevent")
    else:
        pass

    Args = InfiniteEmptyList(context.args)
    if has_inline_kwargs := Args[0].endswith(':'):
        infos_event |= addevent_analyse_args_as_object(Args)

    if not infos_event.get('link'):
        infos_event |= add_event_enrich_reply_with_link(update, context, reply=reply)

    infos_event = {k.lower():v for k,v in infos_event.items()}
    CanonInfo = add_event_canon_infos(infos_event=infos_event)

    source_user_id = update.effective_user.id
    chat_id = update.effective_chat.id

    required_time = do_if_setting_on(read_chat_settings('event.addevent.required_time'))
    date_str, time, name, date, date_end, datetime, datetime_utc, tz, tz_explicit = \
        parse_datetime_point(update, context, has_inline_kargs=has_inline_kwargs,
                             when_infos=CanonInfo.when_infos, what_infos=CanonInfo.what_infos, required_time=required_time)
    
    initial_name = name
    if do_unless_setting_off(read_chat_settings('event.location.autocomplete')):
        name = update_name_using_locations(name, chat_id=chat_id)

    chat_timezones = read_chat_settings("event.timezones")

    do_event_admin_check('add', setting=read_chat_settings('event.admins'), user_id=update.effective_user.id)

    new_event_id = add_event_to_db(datetime_utc=datetime_utc, name=name, chat_id=chat_id, source_user_id=source_user_id)

    if infos_event.get('link'):
        simple_sql((''' insert into EventLinkAttr(event_id, link) VALUES (?,?)''', (new_event_id, infos_event['link'])))

    if do_unless_setting_off(read_chat_settings('event.location.autocomplete')):
        implicit_thereis(what=initial_name, chat_id=chat_id)

    await post_event(update, context, name=name, datetime=datetime, time=time, date_str=date_str, chat_timezones=chat_timezones, tz=tz, tz_explicit=tz_explicit, chat_id=chat_id, datetime_utc=datetime_utc, link=infos_event and infos_event.get('link'))

def addevent_analyse_args_as_object(Args):
    if len(Args) == 0:
        return {}
    key = Args[0][:-1]
    i = 1
    while not (Args[i].endswith(':') or i >= len(Args)):
        i += 1
    return {key: ' '.join(Args[1:i])} | addevent_analyse_args_as_object(InfiniteEmptyList(Args[i:]))

def add_event_canon_infos(*, infos_event):
    other_infos = {k: infos_event[k] for k in infos_event.keys() - {'when', 'what', 'where', 'link'}}
    when_infos = infos_event.get('when') or ''
    what_infos = ' '.join(natural_filter([
        infos_event.get('what') or '',
    ] + [
        '{%s: %s}' % (item[0].capitalize(), item[1])
        for item in other_infos.items()
    ] + [
        "@ " + infos_event['where'] if infos_event.get('where') else '',
    ]))

    return DictJsLike(other_infos=other_infos, when_infos=when_infos, what_infos=what_infos)

def add_event_enrich_reply_with_link(update: Update, context: CallbackContext, *, reply):

    if not reply:
        return {}
    
    if (update.effective_message.chat.id, update.effective_message.message_thread_id) == (SPECIAL_ENTITIES[SpecialUsers.CRAZY_JAM_BACKEND], SPECIAL_ENTITIES[SpecialUsers.CRAZY_JAM_BACKEND_THREAD_IN]):
        orig_id_tuple = simple_sql(( ''' select original_message_id, original_chat_username, original_chat_id from FwdRelation where fwd_message_id = ?''', (reply.id, )))
        if orig_id_tuple:
            message_id, username, chat_id = only_one(orig_id_tuple)
            link = (f't.me/{username}/{message_id}' if username else 
                    f't.me/c/{tg_url_id_from_chat_id(chat_id)}/{message_id}')
        else:
            link = None
    else:
        link = None

    if not link:
        channel_pos_id = tg_url_id_from_chat_id(update.effective_chat.id)

        link = (f't.me/c/{channel_pos_id}/{reply.message_thread_id}/{reply.id}' if reply.message_thread_id else 
                f't.me/c/{channel_pos_id}/{reply.id}')
            
    return {'link': link}

class ImplicitLocations:
    Parens = re.compile("(.*)\\((.*)\\).*")

def update_name_using_locations(what, *, chat_id):
    event = split_event_with_where_etc({'what': what})
    if where := event.get('where'):
        results = simple_sql(('select value from EventLocation where chat_id=? and LOWER(key)=LOWER(?)', (chat_id, where,)))
        
        if results:
            long_location = only_one(results)[0]
        else:
            long_location = None

        if long_location:
            long_location = simplify_multi_line_location(long_location)
            return event.get('what') + ' @ ' + where + ' ' + '(' + long_location + ')'
        
    return what

def implicit_thereis(*, what:str, chat_id):
    what = what or ''
    _, *where = what.split("@", maxsplit=1)

    if not where:
        return
    
    where, = where
    Re = ImplicitLocations.Parens
    
    if not(m := Re.fullmatch(where)):
        return

    location, address = map(str.strip, m.groups())
    
    if not(location and address):
        return

    do_update_thereis_db(location, address, chat_id=chat_id)

async def post_event(update, context, *, name, datetime, time, link, date_str, chat_timezones, tz, tz_explicit=False, chat_id, datetime_utc):
    send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)

    emojis = EventFormatting.emojis

    infos = split_event_with_where_etc({'what': name, 'link': link})

    # 1. Send info in text

    displays = []
    if not time:
        displays = []
    elif not chat_timezones:
        displays.append('simple' if not tz_explicit else ('tz_conversion', tz))
    elif len(chat_timezones) == 1:
        displays.append('simple' if tz in chat_timezones else ('tz_conversion', chat_timezones[0]))
    else:
        displays.extend(('tz_conversion', timezone) for timezone in chat_timezones)

    await send(event_text := '\n'.join(filter(None, [
        f"Event added:",
        f"{emojis.Name} {infos['what']}",
    ] + ([
        f"{emojis.Location} {infos['where']}",
    ] if infos.get('where') else []) + [
        f"{emojis.Date} {datetime:%A} {datetime.date():%d/%m/%Y} ({date_str})"
    ] + [
        f"{emojis.Time} {datetime_tz:%H:%M}" if display == 'simple' else 
        f"{emojis.Time} {datetime_tz:%H:%M} ({timezone})" if datetime_tz.date() == datetime.date() else
        f"{emojis.Time} {datetime_tz:%H:%M} on {datetime_tz.date():%d/%m/%Y} ({timezone})"
        for display in displays
        for timezone in [tz if display == 'simple' else display[1] if display[0] == 'tz_conversion' else raise_error(AssertionError)]
        for datetime_tz in [datetime.astimezone(timezone)]
    ] + [
        f"{emojis.Link} {link}"
    ] * (bool(link) and do_if_setting_on(read_chat_settings('event.addevent.display_link'))))))
    
    if do_if_setting_on(read_chat_settings('event.addevent.display_file')):
        # 2. Send info as clickable ics file to add to calendar
        if do_unless_setting_off(read_chat_settings('event.addevent.help_file')):
            await send('Click the file below to add the event to your calendar:')
        await export_event(update, context, name=name, datetime_utc=datetime_utc)
    
    # 3. Forward it to other chats
    forward_ids = simple_sql(('select a_chat_id, a_name, a_thread_id from EventFollow where b_chat_id = ?', (str(chat_id), )))
    event_text_without_first_line = '\n'.join(list_del(event_text.splitlines(), 0))
    for forward_id, forward_my_chat_name, forward_thread_id in forward_ids:
        await context.bot.send_message(
            text=f'Event from {forward_my_chat_name}:' + '\n' + event_text_without_first_line,
            chat_id=forward_id,
            message_thread_id=forward_thread_id or None)

    if forward_ids:
        if do_unless_setting_off(read_chat_settings('event.addevent.display_forwarded_infos')):
            await send(f'Forwarded to {len(forward_ids)} chats')


from abc import ABC, abstractmethod

class GeneralAction(ABC):
    async def __call__(self, update: Update, context: CallbackContext):
        self.update = update
        self.context = context

        self.Args = InfiniteEmptyList(self.context.args)

        try:
            return await self.run()
        except UsageError as e:
            self.exception = e
            return await self.print_usage()
    
    def send(self, *a, **b):
        return make_send(self.update, self.context)(*a, **b)
    
    def chat_settings(self, *a, **b):
        return make_read_chat_settings(self.update, self.context)(*a, **b)
    
    @abstractmethod
    async def run(self):
        raise NotImplementedError
    
    async def print_usage(self):
        await self.send("Arguments not correct, please read the manual")

    def get_chat_id(self):
        return self.update.effective_chat.id
    
    def get_user_id(self):
        return self.update.effective_user.id

class OnListAction:
    def __init__(self, *, name, conn, chat_id):
        self.conn = conn
        self.name = name 
        self.chat_id = chat_id

    def list_id(self):
        if not hasattr(self, '_list_id'):
            self._list_id = only_one(only_one(self.my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (self.chat_id, self.name, )))))
        return self._list_id
    
    def my_simple_sql(self, *args, **kwargs):
        return partial(simple_sql, connection=self.conn)(*args, **kwargs)

class OnTreeAction(OnListAction):
    def itree(self, itree_str):
        return tuple(map(int, itree_str.split('.')))
    
    def assert_is_correct_itree(self, itree):
        assert len(itree)
        assert all(x >= 1 for x in itree)

    def tree_getnode(self, itree):
        self.list_id()

        prev = None
        for i in itree:
            if prev is None:
                x = "IS NULL"
                p = (self.list_id(), i-1)
            else:
                x = "= ?"
                p = (self.list_id(), prev, i-1)

            rowid, = only_one(
                self.my_simple_sql((f''' select rowid from ListElement where listid=? AND tree_parent {x} LIMIT 1 OFFSET ?''', p)),
                none=UserError(f"{'.'.join(map(str, itree))} does not exist in the tree"))
            
            prev = rowid
        return prev
    
class iameventadmin(GeneralAction):
    async def run(self):
        if self.Args:
            raise UsageError
        await self.send("Use: /chatsettings event.admins {0}\nOr: /chatsettings event.admins += {0}".format(self.get_user_id()))
iameventadmin = iameventadmin()

class events(GeneralAction):
    class DuplicatesUsageError(UsageError):
        pass

    async def run(self):
        match self.Args[0].lower():
            case 'deldups' | 'deldup' | 'deleteduplicates' | 'removeduplicates':
                return await self.delete_duplicates(args=self.Args[1:])
            case 'mergetimeduplicates':
                return await self.merge_time_duplicates(args=self.Args[1:])
            case 'add':
                return self.send("Not implements yet: use /addevent")
            case _:
                raise UsageError
    
    async def delete_duplicates(self, args):
        datetime_range = parse_datetime_range(self.update, args=args)
        beg, end, tz, when = (datetime_range[x] for x in ('beg_utc', 'end_utc', 'tz', 'when'))

        events = simple_sql_dict(('''
            SELECT rowid, date, name
            FROM Events
            WHERE ? <= date AND date < ?
            AND chat_id=?
            ORDER BY date''',
            (DatetimeDbSerializer.strftime(beg), DatetimeDbSerializer.strftime(end), self.update.effective_chat.id,)))
        
        to_delete = []
        for i in range(len(events)-1):
            if (events[i]['date'], events[i]['name']) == (events[i+1]['date'], events[i+1]['name']):
                to_delete.append(str(events[i]['rowid']))
        
        do_event_admin_check('del', setting=self.chat_settings('event.admins'), user_id=self.get_user_id())

        result = simple_sql_modify(('''DELETE FROM Events where rowid IN ({qmarks})'''.format(qmarks=','.join('?' * len(to_delete))), (*to_delete, ),))

        assert_true(result['rowcount'] == len(to_delete), ValueError("Wrong deletion"))

        return await self.send(f"{len(to_delete)} event(s) deleted")
    
    async def merge_time_duplicates(self, args):
        datetime_range = parse_datetime_range(self.update, args=args)
        beg, end, tz, when = (datetime_range[x] for x in ('beg_utc', 'end_utc', 'tz', 'when'))

        events = simple_sql_dict(('''
            SELECT rowid, date, name
            FROM Events
            WHERE ? <= date AND date < ?
            AND chat_id=?
            ORDER BY date''',
            (DatetimeDbSerializer.strftime(beg), DatetimeDbSerializer.strftime(end), self.update.effective_chat.id,)))
        
        def merge_names(names:list[str]):
            if not names:
                return ''
            if len(S := set(names)) == 1:
                return next(iter(S))
            return ' | '.join(names)  # will be better later, like analyzing locations (do not repeat locations)

        to_delete = []
        to_update = []
        i = 0
        while i < len(events) - 1:
            span = [i, None]
            while i < len(events) - 1 and events[i]['date'] == events[i+1]['date']:
                i += 1
            span[1] = i+1

            if span[1] - span[0] > 1:
                to_delete.extend(events[i]['rowid'] for i in range(1+span[0], span[1]))
                to_update.append((
                    events[span[0]]['rowid'],
                    merge_names([events[i]['name'] for i in range(*span)])
                ))
            else:
                i += 1
        
        do_event_admin_check('del', setting=self.chat_settings('events.admins'), user_id=self.get_user_id())
        do_event_admin_check('edit', setting=self.chat_settings('events.admins'), user_id=self.get_user_id())

        with sqlite3.connect('db.sqlite') as conn:
            my_simple_sql_modify = partial(simple_sql_modify, connection=conn)
            conn.execute('begin transaction')

            result = my_simple_sql_modify(('''DELETE FROM Events where rowid IN ({qmarks})'''.format(qmarks=','.join('?' * len(to_delete))), (*to_delete, ),))
            assert_true(result['rowcount'] == len(to_delete), ValueError("Wrong deletion"))

            for rowid, name in to_update:
                my_simple_sql_modify(('''UPDATE Events SET name=? WHERE rowid=?''', (name, rowid)))

            conn.execute('end transaction')

        return await self.send(f"{len(to_delete)} event(s) deleted, {len(to_update)} event(s) group found")

    async def print_usage(self):
        match self.exception:
            case self.DuplicatesUsageError:
                return await self.send("Usage:\n/events removeduplicates [when]")
            case _:
                return await self.send("Usage:\n/events add when [time] [what]\n/events removeduplicates [when]\n")

class listsmodule:

    class ListAlreadyExist(ValueError):
        pass

    @staticmethod
    def list_exists(*, chat_id, name, conn):
        my_simple_sql = partial(simple_sql, connection=conn)

        return bool(my_simple_sql((''' select 1 from List where chat_id=? and lower(name)=lower(?) ''', (chat_id, name,) )))
    
    @staticmethod
    def get_list_type(*, chat_id, name, conn):
        my_simple_sql = partial(simple_sql, connection=conn)

        return only_one(only_one(my_simple_sql(('''select type from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,)))))

    @staticmethod
    def get_list_id(*, chat_id, name, conn):
        my_simple_sql = partial(simple_sql, connection=conn)

        x, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))
        return x
    
    @staticmethod
    def load(*, chat_id, name, conn):
        my_simple_sql = partial(simple_sql, connection=conn)

        listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))
        return [x[0] for x in my_simple_sql(('''select value from ListElement where listid=?''', (listid, )))]
    
    @staticmethod
    def dump(*, chat_id, name, conn, values):
        listsmodule.editmultilist.do_it(conn=conn, chat_id=chat_id, name=name, values=values)

    class TreeNodeDump(NamedTuple):
        value: str
        rowid: int
        parent: int

    @staticmethod
    def treeload(*, chat_id, name, conn):
        my_simple_sql = partial(simple_sql, connection=conn)

        listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))
        return [listsmodule.TreeNodeDump(*x) for x in my_simple_sql(('''select value, rowid, tree_parent from ListElement where listid=?''', (listid, )))]
    
    @staticmethod
    def treedump(*, chat_id, name, conn, values):
        my_simple_sql = partial(simple_sql, connection=conn)

        listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))
        
        mapped = {}

        cursor = conn.cursor()
        for node in values:
            cursor.execute('''insert into ListElement(listid, value) VALUES (?, ?)''', (listid, node.value, ))
            mapped[node.rowid] = cursor.lastrowid
        
        for node in values:
            if node.parent is not None:
                cursor.execute('''update ListElement set tree_parent=? where rowid=?''', (mapped[node.parent], mapped[node.rowid]))
    class forcecreatelist:
        @staticmethod
        def do_it(*, conn, chat_id, name, user_id, type_list: object, force_creation: bool):
            my_simple_sql = partial(simple_sql, connection=conn)

            # pre creation operation : clear list if exists
            if listsmodule.list_exists(conn=conn, chat_id=chat_id, name=name):
                if force_creation:
                    listsmodule.clearlist.do_it(conn=conn, chat_id=chat_id, name=name)
                    listid = listsmodule.get_list_id(conn=conn, chat_id=chat_id, name=name)
                    my_simple_sql(('delete from ListElement where listid=?', (listid, )))
                    my_simple_sql(('delete from List where rowid=?', (listid, )))
                else:
                    raise listsmodule.ListAlreadyExist

            # pre creation operation : determine the type of the new list
            # - in case of copy, verify the other list exists, and save it's type
            match type_list:
                case 'copy', copy_from_name:
                    if listsmodule.list_exists(conn=conn, chat_id=chat_id, name=copy_from_name):
                        actual_type = target_type = only_one(only_one(my_simple_sql((''' select type from List where chat_id=? AND lower(name)=lower(?) ''', (chat_id, copy_from_name, )))))
                        if actual_type not in ('list', 'tasklist', 'tree', ):
                            raise UserError("Impossible at the moment")
                    else:
                        raise UserError(f'List {copy_from_name!r} does not exist')  # transaction will rollback
                case 'tasktree', copy_from_name:
                    actual_type = 'tasktree'
                    if listsmodule.list_exists(conn=conn, chat_id=chat_id, name=copy_from_name):
                        target_type = only_one(only_one(my_simple_sql((''' select type from List where chat_id=? AND lower(name)=lower(?) ''', (chat_id, copy_from_name, )))))
                        if target_type not in ('tree', 'tasktree'):
                            raise UserError("Impossible to create {} from {}".format('tasktree', target_type))
                    else:
                        raise UserError(f'List {copy_from_name!r} does not exist')  # transaction will rollback
                case 'dynamic', dynamic_list:
                    actual_type = 'dynamic' + '.' + dynamic_list
                case 'dynamic', dynamic_list, dynamic_list_params:
                    actual_type = 'dynamic' + '.' + dynamic_list + ' ' + dynamic_list_params
                case 'alias', alias_list:
                    actual_type = 'alias' + '.' + alias_list
                    if not listsmodule.list_exists(conn=conn, chat_id=chat_id, name=alias_list):
                        raise UserError(f"List {alias_list!r} does not exist")
                case _:
                    assert isinstance(type_list, str)
                    actual_type = type_list.lower()

            # creation operation
            my_simple_sql(('insert into List(name, chat_id, source_user_id, type) VALUES (?,?,?,?)', (name, chat_id, user_id, actual_type)))

            # post creation operation
            match type_list:
                case 'list' | 'tasklist' | 'tree' | 'tasktree':
                    pass # nothing to do more
                    
                case 'dynamic', _:
                    pass

                case 'dynamic', _, _:
                    pass

                case 'alias', _:
                    pass

                case 'copy', copy_from_name:
                    if actual_type in ('list', 'tasklist'):
                        values = listsmodule.load(conn=conn, chat_id=chat_id, name=copy_from_name)
                        listsmodule.dump(conn=conn, chat_id=chat_id, name=name, values=values)
                    elif actual_type in ('tree', 'tasktree'):
                        values = listsmodule.treeload(conn=conn, chat_id=chat_id, name=copy_from_name)
                        listsmodule.treedump(conn=conn, chat_id=chat_id, name=name, values=values)
                    else:
                        raise UserError(f'Unknown copy type {copy_from_name}')
                case 'tasktree', copy_from_name:
                    if target_type == 'tree':
                        values = listsmodule.treeload(conn=conn, chat_id=chat_id, name=copy_from_name)
                        values = [listsmodule.TreeNodeDump(value=listsmodule.make_task(x.value), rowid=x.rowid, parent=x.parent) for x in values]
                        listsmodule.treedump(conn=conn, chat_id=chat_id, name=name, values=values)
                    elif target_type == 'tasktree':
                        values = listsmodule.treeload(conn=conn, chat_id=chat_id, name=copy_from_name)
                        listsmodule.treedump(conn=conn, chat_id=chat_id, name=name, values=values)
                    else:
                        raise AssertionError
                case _:
                    raise AssertionError(f'Internal error on type_list variable: {type_list}')

    class createlist(GeneralAction):
        async def run(self):
            match len(self.Args):
                case 0:
                    name = "list"
                case 1:
                    name = self.Args[0]
                case _:
                    raise UsageError
                
            import regex 
            NAME = regex.compile(r"\p{L}+")
            assert_true(NAME.fullmatch(name), UserError("Name should be made of letters"))

            with sqlite3.connect("db.sqlite") as conn:
                my_simple_sql = partial(simple_sql, connection=conn)
                conn.execute('begin transaction')
                if my_simple_sql(('''select 1 from List where chat_id=? and lower(name)=lower(?)''', (self.get_chat_id(), name,))):
                    return await self.send(f'List {name!r} already exist')
                my_simple_sql(('insert into List(name, chat_id, source_user_id) VALUES (?,?,?)', (name, self.get_chat_id(), self.get_user_id())))
                conn.execute('end transaction')
            
            return await self.send(f'List named {name!r} created')
        
        async def print_usage(self):
            return await self.send("/createlist\n/createlist name")
    
    class tasktreeinsertchild:
        @staticmethod
        def do_it(*, parameters, **P):
            itree_str, value = parameters.split(maxsplit=1)
            value = listsmodule.make_task(value)
            listsmodule.treeinsertchild(**P).run(parameters=itree_str + ' ' + value)
    
    def make_task(x):
        if ListLang.IsTask.fullmatch(x):
            return x.strip()
        else:
            return '[ ]' + ' ' + x.strip()

    class treeinsertchild(OnTreeAction):
        def run(self, *, parameters):
            itree_str, value = parameters.split(maxsplit=1)
            itree = self.itree(itree_str)

            self.assert_is_correct_itree(itree)

            nodeid = self.tree_getnode(itree)
            self.my_simple_sql((''' insert into ListElement(listid, value, tree_parent) values (?,?,?)''', (self.list_id(), value, nodeid) ))
    
    class insertintasklist:
        @staticmethod
        def do_it(*, conn, chat_id, name, parameters):
            i, to_add = parameters.split(maxsplit=1)
            modified_value = listsmodule.make_task(to_add)
            listsmodule.insertinlist.do_it(conn=conn, chat_id=chat_id, name=name, parameters=i + ' ' + modified_value)
    
    class insertintasktree:
        @staticmethod
        def do_it(*, parameters, **P):
            i, to_add = parameters.split(maxsplit=1)
            modified_value = listsmodule.make_task(to_add)
            listsmodule.insertintree(**P).run(parameters=i + ' ' + modified_value)

    class replaceintasklist:
        @staticmethod
        def do_it(*, parameters, **P):
            i, to_rep = parameters.split(maxsplit=1)
            mv = listsmodule.make_task(to_rep)
            listsmodule.replaceinlist.do_it(parameters=' '.join([i, mv]), **P)
    class replaceintasktree:
        @staticmethod
        def do_it(*, parameters, **P):
            itree_str, to_rep = parameters.split(maxsplit=1)
            itree = tuple(map(int, itree_str.split('.')))
            
            mv = listsmodule.make_task(to_rep)

            listsmodule.replaceintree(**P).run(parameters=' '.join([itree_str, mv]))

    class replaceintree(OnTreeAction):
        def run(self, *, parameters):
            itree_str, to_rep = parameters.split(maxsplit=1)
            itree = self.itree(itree_str)

            node_rowid = self.tree_getnode(itree)

            self.my_simple_sql((''' update ListElement set value=? where listid=? and rowid=?''', (to_rep, self.list_id(), node_rowid, )))

    def get_listid(*, chat_id, name, my_simple_sql):
        return only_one(only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,)))))


    class insertintree(OnTreeAction):
        def run(self, *, parameters):
            i, to_add = parameters.split(maxsplit=1)
            
            rowids = self.my_simple_sql((''' select rowid, value from ListElement where listid=? AND tree_parent IS NULL''', (listid := self.list_id(), )))
            childrens = [self.my_simple_sql((''' select rowid from ListElement where tree_parent = ?''', (rowid, ))) for rowid, _ in rowids]
            childrens = [[only_one(x) for x in X] for X in childrens]

            i = int(i)
            i = i - 1
            assert i >= 0
            assert i < len(rowids)
            
            for j in range(i, len(rowids) - 1):
                rowid, value = rowids[j]
                nextrowid, nextvalue = rowids[j+1]
                children = childrens[j]
                nextchildren = childrens[j+1]
                
                self.my_simple_sql((''' update ListElement set value=? where rowid=? ''' , (value, nextrowid, )))
                for c in children:
                    self.my_simple_sql((''' update ListElement set tree_parent=? where rowid=? ''' , (nextrowid, c, )))

            if i < len(rowids):
                self.my_simple_sql((''' update ListElement set value=? where rowid=? ''', (to_add, rowids[i][0], )))
                
                cursor = self.conn.cursor()
                cursor.execute(''' insert into ListElement(listid, value, tree_parent) values(?,?,NULL)''', (listid, rowids[-1][1], ))
                new_rowid = cursor.lastrowid
                for c in childrens[-1]:
                    self.my_simple_sql((''' update ListElement set tree_parent=? where rowid=? ''' , (new_rowid, c, )))
            else:
                self.my_simple_sql((''' insert into ListElement(listid, value, tree_parent) values(?,?,NULL)''', (listid, to_add, )))
        

    class insertinlist(GeneralAction):
        @staticmethod
        def do_it(*, conn, chat_id, name, parameters):
            i, to_add = parameters.split(maxsplit=1)
            value = i
            
            my_simple_sql = partial(simple_sql, connection=conn)

            listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))

            rowids = my_simple_sql((''' select rowid, value from ListElement where listid=? ''', (listid, )))

            value = int(value)

            assert value in irange(-len(rowids), len(rowids)+1)
            assert value != 0

            if value < 0:
                value = len(rowids) + value
            else:
                value = value - 1

            for (rowid, v), (nextrowid, nextv) in zip(rowids[value:], rowids[value+1:]):
                my_simple_sql((''' update ListElement set value=? where rowid=? ''' , (v, nextrowid, )))

            if value < len(rowids):
                my_simple_sql((''' update ListElement set value=? where rowid=? ''', (to_add, rowids[value][0], )))
                my_simple_sql((''' insert into ListElement(listid, value) values(?,?)''', (listid, rowids[-1][1], )))
            else:
                my_simple_sql((''' insert into ListElement(listid, value) values(?,?)''', (listid, to_add, )))

    class replaceinlist(GeneralAction):
        @staticmethod
        def do_it(*, conn, chat_id, name, parameters):
            i, to_rep = parameters.split(maxsplit=1)
            value = i
            
            my_simple_sql = partial(simple_sql, connection=conn)

            listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))

            N = only_one(only_one(my_simple_sql(('''select count(*) from ListElement where listid=?''', (listid,)))))

            value = int(value)

            assert value in irange(-N, +N)
            assert value != 0

            if value < 0:
                value = N + value
            else:
                value = value - 1

            my_simple_sql(('''update ListElement set value=? where listid=? LIMIT 1 OFFSET ? ''', (to_rep, listid, value)))

    @staticmethod
    def parse_interval(value:str, N:int=None) -> range:
        """
        '1' → [1]: range
        '-5' → [-5]: range
        '2-4' → [2, 3, 4]: range
        '2:4' → [2, 3, 4]: range
        '-2:-1' → [-2, -1]: range
        """
        if (dot := ':' in value) or (dash := '-' in value):
            a,b = value.split(':' if dot else '-')
            if a and b:
                # interval: a:b or a-b or -a:b or -a:-b or a:-b
                int(a), int(b)
                assert int(a) <= int(b)
                return irange(int(a), int(b))
            
            elif a and not b:
                # 5: or 5-
                if N is None:
                    raise ValueError("Invalid format for number or range")
                else:
                    return irange(int(a), N)
            elif b and not a:
                if dot:
                    # :5
                    assert int(b) >= 0
                    return irange(1, int(b))
                    raise ValueError("Not implemented yet")
                elif dash:
                    # negative number
                    return irange(- int(b), - int(b))
            else:
                raise AssertionError
                
        else:
            return irange(int(value), int(value))
    
    @staticmethod
    def one_based_to_zero_based(i: int) -> int:
        return i if i < 0 else i - 1        

    @staticmethod
    def is_negative_range(r: range):
        return r.start < 0 or r.stop <= 0
    
    @staticmethod
    def to_positive_range(r: range, N: int, *, based:Literal[0, 1]):
        """
        range( 0, 5) (N = 10) (based = 0) → range(0, 5)
        range(-1, 0) (N = 10) (based = 0) → range(9, 10)
        range(-1, 0) (N = 10) (based = 1) → range(10, 11)
        """
        if listsmodule.is_negative_range(r):
            start = r.start + N + based if r.start < 0 else r.start
            stop = r.stop + N + based if r.stop <= 0 else r.stop

            assert start in range(based, N+based)
            assert stop in irange(based, N+based)
            assert start < stop

            return range(start, stop)
        else:
            return r
    
    @staticmethod
    def parse_interval_to_positive_range(value:str, *, N, based):
        return listsmodule.to_positive_range(listsmodule.parse_interval(value, N), N, based=based)
    
    @staticmethod
    def limit_offset_based1(r: range):
        limit = r.stop - r.start
        offset = listsmodule.one_based_to_zero_based(r.start)
        return limit, offset   
     
    class tasklistcheck:
        @staticmethod
        def do_it(*, conn, chat_id, name, value, direction:Literal['x', '', 'toggle']):
            direction = direction.strip()
            assert direction in ('x', '', 'toggle')
            direction = ' ' if direction == '' else direction

            my_simple_sql = partial(simple_sql, connection=conn)

            listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))

            rowids = my_simple_sql((''' select rowid, value from ListElement where listid=? ''', (listid, )))

            def action(i: int):
                assert i in irange(-len(rowids), len(rowids))
                assert i != 0

                if i < 0:
                    i = len(rowids) + i
                else:
                    i = i - 1
                
                rowid, old_value = rowids[i]

                if m := ListLang.IsTask.fullmatch(old_value):
                    new_check = direction if direction != 'toggle' else ('' if m.group(1) == 'x' else 'x')
                    new_value = '[' + new_check + ']' + ' ' + m.group(2).strip()
                else:
                    raise AssertionError('A non task is stored in a tasklist')
                
                my_simple_sql((''' update ListElement set value=? where rowid=?''', (new_value, rowid)))
            
            for value in value.split():
                for i in listsmodule.parse_interval(value, len(rowids)):
                    action(i)

    class tasktreecheck(OnTreeAction):
        def run(self, *, value, direction:Literal['x', ' ', 'toggle']):
            nodes = []
            for value in value.split():
                itree = tuple(map(int, value.split('.')))

                node = self.tree_getnode(itree)

                nodes.append(node)

            for node in nodes:

                old_value, = only_one(self.my_simple_sql(('''select value from ListElement where listid=? and rowid=?''', (self.list_id(), node))))

                if m := ListLang.IsTask.fullmatch(old_value):
                    if direction == 'toggle':
                        new_check = ' ' if m.group(1) == 'x' else 'x'
                    else:
                        new_check = direction
                    new_value = '[' + new_check + ']' + ' ' + m.group(2).strip()
                else:
                    raise AssertionError('A non task is stored in a tasklist')
                
                self.my_simple_sql(('''Update ListElement set value=? where listid=? and rowid=?''', (new_value, self.list_id(), node)))


    class delinlist(GeneralAction):
        @staticmethod
        def do_it(*, conn, chat_id, name, value):
            my_simple_sql = partial(simple_sql, connection=conn)

            listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))

            rowids = my_simple_sql((''' select rowid from ListElement where listid=? ''', (listid, )))

            for v in value.split():
                for i in map(listsmodule.one_based_to_zero_based, listsmodule.parse_interval(v, len(rowids))):
                    delrowid = rowids[i][0]
                    my_simple_sql((''' delete from ListElement where listid=? and rowid=?''', (listid, delrowid, )))

    class delindynamic:
        @staticmethod
        def do_it(*, conn, chat_id, name, value, dynamic_list):
            my_simple_sql = partial(simple_sql, connection=conn)

            def do_all_delete(delete, N):
                for v in value.split():
                    for i in map(listsmodule.one_based_to_zero_based, listsmodule.parse_interval(v, N)):
                        delete(i)

            match listsmodule.dynamic_list_analyze(dynamic_list):
                case 'flashcard.current':
                    flashcard.delete_in_page(do_all_delete=do_all_delete, connection=conn, chat_id=chat_id, page_name=flashcard.Current)
                case 'flashcard.page', page_name:
                    flashcard.delete_in_page(do_all_delete=do_all_delete, connection=conn, chat_id=chat_id, page_name=page_name)
                case 'event.today':
                    raise UserError("Not implemented yet")
                case _:
                    raise UserError("Unknown dynamic type")

    class delintree(OnTreeAction):
        def run(self, *, parameters):
            itree_str_multiple = parameters.split()
            
            itrees_rowid = []
            for itree_str in itree_str_multiple:
                itree = tuple(map(int, itree_str.split('.')))
            
                self.assert_is_correct_itree(itree)

                node_rowid = self.tree_getnode(itree)

                itrees_rowid.append(node_rowid)

            listid = self.list_id()
            
            def delete(X):
                children = self.my_simple_sql((''' select rowid from ListElement where listid=? and tree_parent=? ''', (listid, X)))
                for c, in children:
                    delete(c)
                self.my_simple_sql((''' delete from ListElement where listid=? and rowid=?''', (listid, X, )))
            
            for node_rowid in itrees_rowid:
                delete(node_rowid)

    def dynamic_list_analyze(dynamic_list):
        dyn_type, *dyn_params = dynamic_list.split(maxsplit=1)
        return dyn_type if not dyn_params else (dyn_type, ' '.join(dyn_params))
    class dynamic_add(GeneralAction):
        @staticmethod
        def do_it(*, conn, chat_id, name, value, dynamic_list):
            match listsmodule.dynamic_list_analyze(dynamic_list):
                case 'flashcard.current':
                    flashcard.parse_and_add(value, page_name=None, chat_id=chat_id, connection=conn)
                
                case 'flashcard.page', page_name:
                    return flashcard.parse_and_add(value, page_name=page_name, chat_id=chat_id, connection=conn)

                case 'event.today':
                    raise UserError("Not implemented")

                case _:
                    raise UserError("Unknown dynamic type")
                
    class addtolist(GeneralAction):
        @staticmethod
        def do_it(*, conn, chat_id, name, value):
            my_simple_sql = partial(simple_sql, connection=conn)

            listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))
            my_simple_sql((''' insert into ListElement(listid, value) values (?,?)''', (listid, value) ))

        async def print_usage(self):
            return await self.send("Usage:\n/addtolist value\n/addtolist listname value")
        
        async def run(self):
            with sqlite3.connect("db.sqlite") as conn:
                my_simple_sql = partial(simple_sql, connection=conn)
                conn.execute('begin transaction')

                match self.Args[0]:
                    case "":
                        name = "list"
                        value = ' '.join(self.Args[0:])
                    case _ as x:
                        if my_simple_sql((''' select 1 from List where chat_id=? and lower(name)=lower(?) ''', (self.get_chat_id(), x,) )):
                            name = x 
                            value = ' '.join(self.Args[1:])
                        else:
                            name = 'list'
                            value = ' '.join(self.Args[0:])

                if not value:
                    raise UsageError
                
                listsmodule.addtolist.do_it(name=name, chat_id=self.get_chat_id(), value=value, conn=conn)
                conn.execute('end transaction')

            return await self.send(f'''List named {name!r} edited''')        

    class removefromlist(GeneralAction):
        async def run(self):
            await self.send("To be implemented")

    class clearlist:
        @staticmethod
        def do_it(*, conn, chat_id, name):
            my_simple_sql = partial(simple_sql, connection=conn)
            listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))
            my_simple_sql(('''delete from ListElement where listid=?''', (listid, )))

    class editmultilist:
        @staticmethod
        def do_it(*, conn, chat_id, name, values):
            # 1: clear (in transaction)
            listsmodule.clearlist.do_it(conn=conn, chat_id=chat_id, name=name)
            # 2: set (in transaction)
            listsmodule.extendmultilist.do_it(conn=conn, chat_id=chat_id, name=name, values=values)

    class editmultitasklist:
        @staticmethod
        def do_it(*, conn, chat_id, name, values):
            # 1: clear (in transaction)
            listsmodule.clearlist.do_it(conn=conn, chat_id=chat_id, name=name)
            # 2: set (in transaction)
            listsmodule.extendmultitasklist.do_it(conn=conn, chat_id=chat_id, name=name, values=values)

    class extend_multi_dynamic:
        @staticmethod
        def do_it(*, conn, chat_id, name, values, dynamic_list):
            for value in values:
                listsmodule.dynamic_add.do_it(conn=conn, chat_id=chat_id, name=name, value=value, dynamic_list=dynamic_list)
    class extendmultilist:
        @staticmethod
        def do_it(*, conn, chat_id, name, values):
            my_simple_sql = partial(simple_sql, connection=conn)
            listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))
            for value in values:
                my_simple_sql(('''insert into ListElement(listid, value) values (?,?)''', (listid, value)))
    
    class extendmultitree:
        @staticmethod
        def do_it(*, conn, chat_id, name, values):
            listsmodule.extendmultilist.do_it(conn=conn, chat_id=chat_id, name=name, values=values)

    class extendmultitasklist:
        @staticmethod
        def do_it(*, conn, chat_id, name, values):
            new_values = list(map(listsmodule.make_task, values))
            listsmodule.extendmultilist.do_it(values=new_values, conn=conn, name=name, chat_id=chat_id)

    class extendmultitasktree:
        @staticmethod
        def do_it(*, conn, chat_id, name, values):
            new_values = list(map(listsmodule.make_task, values))
            listsmodule.extendmultitree.do_it(values=new_values, conn=conn, name=name, chat_id=chat_id)

    class shuffle:
        @staticmethod
        def do_it(*, conn, chat_id, name):
            # 1: load
            values = listsmodule.load(conn=conn, chat_id=chat_id, name=name)
            
            import random
            random.shuffle(values)

            listsmodule.dump(conn=conn, chat_id=chat_id, name=name, values=values)

    class printlist(GeneralAction):
        @staticmethod
        def it(*, conn, chat_id, name, parameters=None, space_between_lines:bool):
            my_simple_sql = partial(simple_sql, connection=conn)

            listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))
            if not parameters:
                result_list = [x[0] for x in my_simple_sql((''' select value from ListElement where listid=?''', (listid, ) ))]
            else:
                N = only_one(only_one(my_simple_sql(('''select count(*) from ListElement where listid=?''', (listid,)))))
                
                r = listsmodule.parse_interval_to_positive_range(parameters, N=N, based=1)
                limit, offset = listsmodule.limit_offset_based1(r)
                result_list = [x for x, in my_simple_sql((''' select value from ListElement where listid=? LIMIT ? OFFSET ?''', (listid, limit, offset)))]

            sep = lambda: '\n' if not space_between_lines else '\n\n'

            return sep().join(map('- {}'.format, result_list)) if result_list else '/'
            
        async def run(self):
            with sqlite3.connect("db.sqlite") as conn:
                my_simple_sql = partial(simple_sql, connection=conn)
                read_chat_settings = make_read_chat_settings(self.update, self.context)
                conn.execute('begin transaction')

                match self.Args[0]:
                    case "":
                        name = "list"
                    case _ as x:
                        if my_simple_sql((''' select 1 from List where chat_id=? and lower(name)=lower(?) ''', (self.get_chat_id(), x,) )):
                            name = x
                        else:
                            name = 'list'
                
                space_between_lines = do_if_setting_on(read_chat_settings('list.space_between_lines'))

                await self.send(listsmodule.printlist.it(conn=conn, chat_id=self.get_chat_id(), name=name, space_between_lines=space_between_lines))
                conn.execute('end transaction')
    
    class clear_dynamic:
        @staticmethod
        def do_it(*, conn, chat_id, name, dynamic_list):
            match listsmodule.dynamic_list_analyze(dynamic_list):
                case 'flashcard.current':
                    return flashcard.clear_flashcards(chat_id=chat_id, connection=conn, page_name=flashcard.Current)
                case 'flashcard.page', page_name:
                    return flashcard.clear_flashcards(chat_id=chat_id, page_name=page_name, connection=conn)
                case _:
                    raise UserError(f'Unknown dynamic list type {dynamic_list}')
    
    class print_dynamic(GeneralAction):
        @staticmethod
        def it(*, conn, chat_id, name, parameters, dynamic_list):
            if parameters:
                raise UserError("This dynamic list does not take parameters")
            
            match listsmodule.dynamic_list_analyze(dynamic_list):
                case 'flashcard.current':
                    return flashcard.print_current_flashcards(chat_id=chat_id, connection=conn)
                case 'flashcard.page', page_name:
                    return flashcard.print_page_flashcards(chat_id=chat_id, page_name=page_name, connection=conn)
                case _:
                    raise UserError(f'Unknown dynamic list type {dynamic_list}')
    
    class enumerate_dynamic(GeneralAction):
        @staticmethod
        def it(*, conn, chat_id, name, parameters, dynamic_list):
            if parameters:
                raise UserError("This dynamic list does not take parameters")
            match listsmodule.dynamic_list_analyze(dynamic_list):
                case 'flashcard.current':
                    return flashcard.enumerate_flashcards(page_name=flashcard.Current, chat_id=chat_id, connection=conn)
                case 'flashcard.page', page_name:
                    return flashcard.enumerate_flashcards(page_name=page_name, chat_id=chat_id, connection=conn)
                case _:
                    raise UserError(f'Unknown dynamic list type {dynamic_list}')
                
    class printtree(GeneralAction):
        @staticmethod
        def it(*, conn, chat_id, name, parameters, indent:int, space_between_lines:bool=False):
            my_simple_sql = partial(simple_sql, connection=conn)
            
            (listid,), = my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,)))
            
            ota = OnTreeAction(conn=conn, chat_id=chat_id, name=name)

            args = parameters.split()
            if len(args) == 0:
                itree = ''
            elif len(args) == 1:
                itree = ota.itree(args[0])
                node = ota.tree_getnode(itree)
            else:
                raise UserError("Too much parameters")

            bits = []

            def run_on(rowid, level):
                for x, xv in my_simple_sql((''' select rowid, value from ListElement where listid=? and tree_parent IS ? ''', (listid, rowid, ))):
                    bits.append((level, xv))
                    run_on(x, level+1)

            if itree:
                xv, = only_one(my_simple_sql((''' select value from ListElement where rowid = ? ''', (node, ))))
                bits.append((0, xv))
                run_on(node, 1)
            else:
                for x, xv in my_simple_sql((''' select rowid, value from ListElement where listid=? and tree_parent IS NULL ''', (listid, ))):
                    bits.append((0, xv))
                    run_on(x, 1)

            indented = lambda x: (x * '  ' if indent is None else x * indent * ' ')
            sep = lambda: '\n' if not space_between_lines else '\n\n'

            return sep().join('{}- {}'.format(indented(x), y) for x, y in bits) if bits else '/'

    class enumeratelist(GeneralAction):
        @staticmethod
        def it(*, conn, chat_id, name, parameters, space_between_lines:bool=False):
            my_simple_sql = partial(simple_sql, connection=conn)

            listid, = only_one(my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,))))
            if not parameters:
                start = 0
                result_list = [x[0] for x in simple_sql((''' select value from ListElement where listid=?''', (listid, ) ))]
            else:
                N = only_one(only_one(my_simple_sql(('''select count(*) from ListElement where listid=?''', (listid,)))))
                
                r = listsmodule.parse_interval_to_positive_range(parameters, N=N, based=1)
                
                limit, offset = listsmodule.limit_offset_based1(r)
                start = offset
                result_list = [x[0] for x in my_simple_sql((''' select value from ListElement where listid=? LIMIT ? OFFSET ?''', (listid, limit, offset)))]

            sep = lambda: '\n' if not space_between_lines else '\n\n'

            return sep().join(('{}. {}'.format(i + start, x) for i, x in enumerate(result_list, start=1))) if result_list else '/'

    class enumeratetree:
        @staticmethod
        def it(*, conn, chat_id, name, parameters='', indent:int, space_between_lines:bool):
            my_simple_sql = partial(simple_sql, connection=conn)
            args = parameters.split()

            (listid,), = my_simple_sql(('''select rowid from List where chat_id=? and lower(name)=lower(?)''', (chat_id, name,)))

            ota = OnTreeAction(conn=conn, chat_id=chat_id, name=name)

            itree = None
            the_range = None
            if len(args) == 0:
                pass
            elif len(args) == 1:
                if '.' in args[0]:
                    itree = ota.itree(args[0])
                    node = ota.tree_getnode(itree)
                else:
                    N = only_one(only_one(my_simple_sql((''' select count(*) from ListElement where listid=? and tree_parent IS NULL ''', (listid, )))))
                    the_range = listsmodule.parse_interval_to_positive_range(args[0], N=N, based=1)
            else:
                raise UserError("Too much parameters")

            bits = []

            def run_on(rowid, level):
                for i, (x, xv) in enumerate(my_simple_sql((''' select rowid, value from ListElement where listid=? and tree_parent IS ? ''', (listid, rowid, )))):
                    trail.append(str(i+1))
                    bits.append(('.'.join(trail), xv, level))
                    run_on(x, level+1)
                    trail.pop()
            
            if itree:
                trail = list(map(str, itree))
                xv, = only_one(my_simple_sql((''' select value from ListElement where rowid=?''', (node, ))))
                bits.append(('.'.join(trail), xv, 0))
                run_on(node, level=1)

            else:
                if the_range:
                    limit, offset = listsmodule.limit_offset_based1(the_range)
                    query = (''' select rowid, value from ListElement where listid=? and tree_parent IS NULL LIMIT ? OFFSET ?''', (listid, limit, offset))
                else:
                    offset = 0
                    query = (''' select rowid, value from ListElement where listid=? and tree_parent IS NULL ''', (listid, ))

                trail = []
                for i, (x, xv) in enumerate(my_simple_sql(query), start=1+offset):
                    trail.append(str(i))
                    bits.append(('.'.join(trail), xv, 0))
                    run_on(x, 1)
                    trail.pop()

            indented = lambda x: (x * '  ' if indent is None else x * indent * ' ')
            sep = lambda: '\n' if not space_between_lines else '\n\n'

            return sep().join('{}{}. {}'.format(indented(z), x, y) for x, y, z in bits) if bits else '/'
        
    class dirlist(GeneralAction):
        @staticmethod
        def it(*, conn, chat_id):
            my_simple_sql = partial(simple_sql, connection=conn)
            results = my_simple_sql(('''select name from List where chat_id=?''', (chat_id, )))
            return '\n'.join("- {}".format(name) for name, in results)
        
        async def run(self):
            with sqlite3.connect("db.sqlite") as conn:
                conn.execute('begin transaction')
                await self.send(self.it(conn=conn, chat_id=self.get_chat_id()) or '/')
                conn.execute('end transaction')

    class dellist(GeneralAction):
        @staticmethod
        def do_it(*, conn, chat_id, name):
            my_simple_sql = partial(simple_sql, connection=conn)
            listsmodule.clearlist.do_it(conn=conn, chat_id=chat_id, name=name)
            my_simple_sql(('''delete from list where chat_id=? and lower(name)=lower(?)''', (chat_id, name)))

        async def run(self):
            with sqlite3.connect("db.sqlite") as conn:
                conn.execute('begin transaction')
                
                my_simple_sql = partial(simple_sql, connection=conn)
                
                match self.Args[0]:
                    case "":
                        name = "list"
                    case _ as x:
                        if my_simple_sql((''' select 1 from List where chat_id=? and lower(name)=lower(?) ''', (self.get_chat_id(), x,) )):
                            name = x
                        else:
                            return await self.send(f"List '{x}' does not exist")
                
                self.do_it(conn=conn, chat_id=self.get_chat_id(), name=name)
                await self.send(f"List named {name!r} deleted")
                conn.execute('end transaction')


def list_del(li, i):
    copy = list(li)
    del copy[i]
    return copy

def natural_filter(x):
    return filter(None, x)


class EventDictAnalysed(TypedDict):
    """
    keys are lowercase
    """

import yaml
def addevent_analyse_yaml(update, context, text:str) -> EventDictAnalysed:
    text = '\n'.join(l for l in text.splitlines() if ':' in l)
    Y = yaml.safe_load(text)
    if not isinstance(Y, dict):
        raise EventAnalyseError('Each line should have a colon symbol, example:\n\nWhat: Party\nWhen: Tomorrow 16h')
    
    Y = {k.lower(): v for k,v in Y.items()}

    result = {}
    keys_lower = {k.lower(): k for k in Y.keys()}
    possibles = EventInfosAnalyse.possibles
    for field in possibles:
        if field in keys_lower:
            result[possibles[field].lower()] = Y.get(keys_lower[field], '')
    
    for field in Y.keys() - possibles.keys():
        result[field.lower()] = Y[field]
    
    if not result.get('when'):
        raise EventAnalyseError("When is mandatory")
    
    Interval = re.compile('(\d{2}:\d{2}) - (\d{2}:\d{2})')
    if match := Interval.search(result['when']):
        result['what'] = ' '.join(natural_filter([result.get('what'), '({})'.format(match.group(0))]))
        result['when'] = Interval.sub(match.group(1), result['when'])

    return result

def only_one(it, error=None, *, many=ValueError, none=ValueError):
    if error is not None:
        many = none = ValueError
    if len(L := list(it)) == 1:
        return L[0]
    else:
        if len(L) == 0:
            raise none
        else:
            raise many
    
def only_one_specific(it):
    return only_one(it, many=TooManyRecords, none=NoRecords)

def only_one_with_error(error):
    return partial(only_one, many=error, none=error)

def addevent_analyse_from_bot(update, context, text:str) -> EventDictAnalysed:
    my_timezone = induce_my_timezone_from_update(update)
    
    lines = GetOrEmpty(text.splitlines())
    if lines[0] in ("Event!", "Event added:", "Event edited:") or re.match('^Event from.*[:]', lines[0]):
        del lines[0]
        
    re_pattern = (
        '^'
        + '({})'.format('|'.join(map(re.escape, EventInfosAnalyse.emojis_meaning)))
        + '\\s*'
        + '(.*)'
    )
    from collections import defaultdict
    infos_raw = defaultdict(list)
    Re = re.compile(re_pattern, re.I)
    for line in lines:
        if match := Re.match(line):
            infos_raw[EventInfosAnalyse.emojis_meaning[match.group(1)].lower()].append(match.group(2))

    def deal_with_timezones(infos_raw):
        infos_raw = infos_raw.copy()

        def extract_timezone(data):
            Re2 = re.compile(
                '(.*)'
                + '\s*'
                + re.escape('(') + '(' + '.*?' + ')' + re.escape(')')
            )
            if match := Re2.search(data):
                time_str, tz = match.group(1), match.group(2)
                try:
                    tz = ZoneInfo(tz)
                except ZoneInfoNotFoundError:
                    tz = None
                if not tz:
                    return time_str.strip(), None
                return time_str, tz
            return data.strip(), None
        
        if infos_raw.get('time'):
            all_times = [
                (local_time_str, local_tz_str)
                for local_time_str, local_tz_str in map(extract_timezone, infos_raw['time'])
            ]

            if any(y is not None for x,y in all_times):
                local_only_one: Callable = partial(only_one, none=EventAnalyseError('Timezone not found'), many=EventAnalyseError(f'Multiple values for timezone {my_timezone!s}'))

                local_time_str, local_tz = local_only_one(
                    (local_time_str, local_tz)
                    for local_time_str, local_tz in all_times
                    if my_timezone == local_tz)

                infos_raw['time'] = [ local_time_str ]

        return infos_raw
    
    infos_raw = deal_with_timezones(infos_raw)
    
    def reduce_multi_values(infos_raw):
        return {k: ' & '.join(v) for k,v in infos_raw.items()} 

    infos_raw = reduce_multi_values(infos_raw)

    what = infos_raw.get('name', '')
    where = infos_raw.get('location', '')
    try:
        when = infos_raw['date'] + ((' ' + infos_raw['time']) if infos_raw.get('time') else '')
    except KeyError:
        raise EventAnalyseError("Missing Date in message")

    if match := re.search('\s*' + re.escape('(') + '.*' + re.escape(')'), when):
        when = when[:match.span(0)[0]] + when[match.span(0)[1]:]

    if match := re.match('|'.join(map(re.escape, DatetimeText.days_english)), when, re.I):
        when = when[:match.span(0)[0]] + when[match.span(0)[1]:]

    if match := re.search('(' + '\d+' + re.escape('/') + '\d+' + re.escape('/') + '\d+' + ')', when):
        when = when[:match.span(0)[0]] + '-'.join(reversed(match.group(1).split('/'))) + when[match.span(0)[1]:]

    when = when.strip()

    return {
        'what': what + ' @ ' + where if where else what,
        'when': when,
    }

def retrieve_event_from_db(update, context, what: str, when: str):
    from datetime import datetime
    event: ParsedEventFinal = parse_datetime_point(update, context, when_infos=when, what_infos=what)
    
    datetime_utc = event.datetime_utc
    name = event.name
    chat_id = update.effective_chat.id

    return only_one(simple_sql_dict(("SELECT rowid, date, name from Events where date=? and name=? and chat_id=?", (DatetimeDbSerializer.strftime(datetime_utc), name, chat_id))),
             many=UserError("Too many events matching this information (duplicate in db)"),
             none=UserError("No events matching this information (was probably deleted)"))


def enrich_event_with_where(event):
    if event.get('where'):
        return event
    event = dict(event)
    event['where'] = GetOrEmpty(re.compile('(?:[ ]|^)[@][ ]').split(event['what']))[1]
    if not event.get('where'):
        del event['where']
    return event

def split_event_name_into_what_where(event_name):
    X = GetOrEmpty(re.compile('(?:[ ]|^)[@][ ]').split(event_name, maxsplit=1))
    return (X[0], X[1]) # X[1] can be empty

def split_event_with_where_etc(event):
    event = dict(event)

    event['what'], event['where'] = split_event_name_into_what_where(event['what'])

    if not event.get('where') and 'where' in event:
        del event['where']
    if not event.get('link') and 'link' in event:
        del event['link']
    return event

def addevent_analyse(update, context) -> EventDictAnalysed:
    if not (reply := update_get_reply(update)):
        raise UserError("Cannot analyse if there is nothing to analyse")

    exceptions = []
    try:
        return addevent_analyse_yaml(update, context, reply.text) or {}
    except yaml.error.YAMLError as e:
        exceptions.append(EventAnalyseError("YAML Error:" + str(e)))
    except EventAnalyseError as e:
        exceptions.append(e)
    
    try:
        return addevent_analyse_from_bot(update, context, reply.text) or {}
    except EventAnalyseError as e:
        exceptions.append(e)
    
    if exceptions:
        raise exceptions[0] if len(exceptions) == 1 else EventAnalyseMultipleError(exceptions)
    else:
        raise EventAnalyseError("I cannot interpret this message as an event")

async def weekiso(update, context):
    send = make_send(update, context)
    await send("{0}-W{1}-{2}".format(*Datetime.today().isocalendar()))

async def weekisoroman(update, context):
    send = make_send(update, context)
    
    def int_to_roman(num):
        if not (0 < num < 4000):
            raise ValueError("Input must be between 1 and 3999")
        
        val = [
            1000, 900, 500, 400,
            100, 90, 50, 40,
            10, 9, 5, 4,
            1
        ]
        syms = [
            "M", "CM", "D", "CD",
            "C", "XC", "L", "XL",
            "X", "IX", "V", "IV",
            "I"
        ]
        roman_num = ''
        for i in range(len(val)):
            while num >= val[i]:
                roman_num += syms[i]
                num -= val[i]
        return roman_num

    await send("{}-W{}-{}".format(*map(int_to_roman, Datetime.today().isocalendar())))

async def whereisto(update, context, *, command: Literal['whereis', 'whereto']):
    send = make_send(update, context)

    key = None
    if reply := update_get_reply(update):
        try:
            infos_event = addevent_analyse(update, context)
            infos_event = enrich_event_with_where(infos_event)
            key = infos_event.get('where', None)  
        except UserError as e:
            reply_error = e

    if key is None:
        try:
            keys = context.args
            key = ' '.join(keys)
            if not key:
                raise ValueError
        except ValueError:
            return await send("Usage: /whereis place\n/whereis (on a event message)")
    
    key: str

    if command == 'whereis':
        results = simple_sql(('select value from EventLocation where chat_id=? and LOWER(key)=LOWER(?)', (chat_id := update.effective_chat.id, key,)))
        await send("I don't know ! :)" if not results else "→ " + only_one(results)[0])
    
    elif command == 'whereto':
        current_key = key
        results = []
        while True:
            current_result = simple_sql(('select value from EventLocation where chat_id=? and LOWER(key)=LOWER(?)', (chat_id := update.effective_chat.id, current_key, )))
            if not current_result:
                break
            current_key = current_result[0][0]
            if current_key in results:
                results.append(current_key)
                break  # we stop because there is a loop
            else:
                results.append(current_key)
                continue
        await send("I don't know ! :)" if not results else '\n'.join(map("→ {}".format, results)))

    else:
        raise AssertionError

async def whereis(update:Update, context:CallbackContext):
    return await whereisto(update:=update, context=context, command='whereis') 

async def whereto(update:Update, context:CallbackContext):
    return await whereisto(update:=update, context=context, command='whereto') 

async def thereis(update:Update, context:CallbackContext):
    send = make_send(update, context)

    arrows_symbols = ("->", "<-", "--", "→", "←")

    def strip_leading_arrow(text):
        """
        >>> strip_leading_arrow('Hello')
        'Hello'
        >>> strip_leading_arrow('→ Hello')
        'Hello'
        """
        Re = re.compile("^(→ |-> )")
        return Re.sub('', text)

    def split_by_equals(List):
        # example: List = ["A", "B", "=", "C", "D", "=", "F"]
        Is = [i for i in range(len(List)) if List[i] == "="]
        # Is = [2, 5]
        breaks = [' '.join(context.args[a+1:b]) for a, b in zip([-1] + Is, Is + [len(context.args)])]
        # breaks = ["A B", "C D", "F"]
        return breaks
    
    def split_by_arrows(List):
        # example: List = ["A", "B", "->", "C", "D", "--", "E"]
        Is = [i for i in range(len(List)) if List[i] in arrows_symbols]
        # Is = [2, 4]
        breaks_values = [' '.join(context.args[a+1:b]) for a, b in zip([-1] + Is, Is + [len(context.args)])]
        # break_values = ["A B", "C D", "E"]
        return breaks_values

    def parse_args(tries):
        match tries:
            case 1:
                # at least one equal: A = B = C means ((A -> C), (B -> C))
                assert_true("=" in context.args, ValueError('Must have at least one "=" for assignation expression'))
                assert_true(len(set(arrows_symbols) & set(context.args)) == 0, ValueError("Pure assignation in that block"))
                # = assignation
                breaks = split_by_equals(context.args)
                keys = breaks[:-1]
                values = [breaks[-1]]
                assert_true(values[0], UserError("Must be something after the ="))
            case 2:
                # at least one arrow symbol: A -> B -> C means ((A -> B), (B -> C))
                breaks = split_by_arrows(context.args)
                assert_true(len(breaks) > 1, ValueError)
                keys, values = [], []
                for i in range(len(breaks) - 1):
                    keys.append(breaks[i])
                    values.append(breaks[i+1])
            case 3:
                # length 2
                key, value = context.args
                values = [value]
                keys = [key]
            case 4:
                # no equal and no arrows
                key, *values = context.args
                value = ' '.join(values)
                values = [value]
                keys = [key]
        return keys, values

    def try_parse_args():
        for tries in (1, 2, 3, 4):
            try:
                return parse_args(tries)
            except UserError:
                raise
            except ValueError:
                continue
        else:
            raise UsageError

    if reply := update_get_reply(update):
        if reply.text.startswith('/whereis') or reply.text.startswith("/whereis@" + context.bot.username):
            keys = [GetOrEmpty(reply.text.split(maxsplit=1))[1]]
            values = [' '.join(context.args)]
        else:
            keys = split_by_equals(context.args)
            values = [strip_leading_arrow(reply.text)]
    else:
        try:
            keys, values = try_parse_args()
        except UsageError as e:
            return await send("Usage:\n/thereis place location\n/thereis place = location")
    
    for i, key in enumerate(keys):
        await save_thereis(key, values[0] if len(values) == 1 else values[i], update=update, context=context)

async def delthereis(update, context):
    send = make_send(update, context)
    
    loc = " ".join(context.args)

    do_delete_thereis_db(loc, chat_id=update.effective_chat.id)

    return await send("Location deleted")

def do_save_thereis_db(key, value, *, chat_id):
    assert_true(key and value, UserError("Key and Values must be non null"))
    
    with sqlite3.connect("db.sqlite") as conn:
        my_simple_sql = partial(simple_sql, connection=conn)
        conn.execute('begin transaction')
        
        my_simple_sql(('delete from EventLocation where chat_id=? and LOWER(key)=LOWER(?)', (chat_id, key)))
        my_simple_sql(('insert into EventLocation(key, value, chat_id) VALUES (?,?,?)', (key, value, chat_id)))
        conn.execute('end transaction')

def do_delete_thereis_db(key, *, chat_id):
    assert_true(key, UserError("Key and Values must be non null"))
    
    with sqlite3.connect("db.sqlite") as conn:
        my_simple_sql = partial(simple_sql, connection=conn)
        conn.execute('begin transaction')
        my_simple_sql(('delete from EventLocation where chat_id=? and LOWER(key)=LOWER(?)', (chat_id, key)))
        conn.execute('end transaction')

def do_update_thereis_db(key, value, *, chat_id):
    assert_true(key and value, UserError("Key and Values must be non null"))
    
    with sqlite3.connect("db.sqlite") as conn:
        my_simple_sql = partial(simple_sql, connection=conn)
        conn.execute('begin transaction')
        # if not my_simple_sql(('select 1 from EventLocation where chat_id=? and LOWER(key)=LOWER(?)', (chat_id, key))):
        my_simple_sql(('delete from EventLocation where chat_id=? and LOWER(key)=LOWER(?)', (chat_id, key)))
        my_simple_sql(('insert into EventLocation(key, value, chat_id) VALUES (?,?,?)', (key, value, chat_id)))
        conn.execute('end transaction')

async def save_thereis(key, value, *, update, context):
    send = make_send(update, context)

    do_save_thereis_db(key, value, chat_id=update.effective_chat.id)

    await send(f"Elephant remembers location:\n{key!r}\n→ {value!r}")

from datetime import datetime, timedelta
def sommeil(s, *, command) -> tuple[datetime, datetime]:
    if m := re.match("/%s (.*)" % command, s):
        s = m.group(1)
    args = s.split()
    a,tiret,b = args
    assert tiret == '-'
    a = map(int, a.split(":"))
    b = map(int, b.split(":"))
    ax,ay = a
    bx,by = b
    from datetime import time
    at = time(hour=ax, minute=ay)
    bt = time(hour=bx, minute=by)
    now = datetime.now().astimezone()
    adt = datetime.combine(now, at)
    bdt = datetime.combine(now, bt)
    if not adt < bdt:
        adt -= timedelta(days=1)
    assert adt < bdt
    return adt, bdt

async def sleep_(update, context):
    send = make_send(update, context)
    from_dt, to_dt = sommeil(update.effective_message.text, command='sleep')
    await send(str(to_dt - from_dt))

def parse_datetime_range(update, *, args, default="week", tz=None):
    from datetime import date as Date, time as Time, datetime as Datetime

    when_2 = None
    if not args:
        when = default
    elif len(args) == 1:
        when, = args  # beware of the ","
    elif len(args) == 2:
        when, when_2 = args
    else:
        raise UserError("<when> must be a day of the week, or a day of the month")
    
    time = Time(0, 0)
    tz = tz or induce_my_timezone_from_update(update)

    def make(when):
        beg_date, end_date = DatetimeText.to_date_range(when, tz=tz)
        beg_local, end_local = Datetime.combine(beg_date, time), Datetime.combine(end_date, time)
        beg, end = (x.replace(tzinfo=tz).astimezone(ZoneInfo('UTC')) for x in (beg_local, end_local))
        return dict(beg_utc=beg, end_utc=end, tz=tz, when=when, beg_local=beg_local, end_local=end_local)
    
    D1 = make(when)
    if not when_2:
        return D1
    else:
        D2 = make(when_2)
        return dict(beg_utc=D1['beg_utc'], end_utc=D2['end_utc'], tz=tz, when=when + '-' + when_2, beg_local=D1['beg_local'], end_local=D2['end_local'])
    
async def next_or_last_event(update: Update, context: CallbackContext, n:int, *, relative=False):
    from datetime import datetime as Datetime
    send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)

    do_event_admin_check('list', setting=read_chat_settings('event.admins'), user_id=update.effective_user.id)

    datetime_str = None
    skip_n = None
    if len(context.args) == 0:
        pass
    elif len(context.args) == 1:
        first_arg, = context.args
        try:
            skip_n = int(first_arg)
        except ValueError:
            pass
        if not skip_n > 0:
            raise UserError('n must be > 0')
        if skip_n is None:
            datetime_str = first_arg
            if len(datetime_str) <= len('2020-01-01'):
                datetime_str += ' ' + '00:00:00'
    elif len(context.args) == 2:
        date, hour = context.args
        if len(hour) <= len('08:00'):
            hour += ':00'
        datetime_str = date + ' ' + hour
    else:
        raise UserError("Usage: /nextevent\n/nextevent datetime\n/nextevent n")
    if skip_n is None:
        skip_n = 1
    
    chat_timezones = read_chat_settings("event.timezones")
    tz = induce_my_timezone_from_update(update)
    now = Datetime.now(UTC) if not datetime_str else DatetimeDbSerializer.strptime(datetime_str.replace('T', ' ')).replace(tzinfo=tz).astimezone(ZoneInfo('UTC'))

    events = simple_sql_dict(('''
        SELECT date as date_utc, name as name
        FROM Events
        WHERE %s
        AND chat_id=?
        ORDER BY date %s, rowid DESC
        LIMIT 1
        OFFSET ?
    ''' % ({1: "date >= ?", -1: "date <= ?"}[n], {1: 'ASC', -1:'DESC'}[n]), (now, update.effective_chat.id, skip_n - 1)))

    if len(events) == 0:
        return await send("No event !")

    strptime = DatetimeDbSerializer.strptime

    date_utc, name = events[0]
    datetime = strptime(date_utc).replace(tzinfo=ZoneInfo('UTC')).astimezone(tz)
    date, time = datetime.date(), datetime.time()
    infos = split_event_with_where_etc({'what': name})

    emojis = EventFormatting.emojis
    await send('\n'.join(natural_filter([
        f"Event!",
        f"{emojis.Name} {infos['what']}",
    ] + ([
        f"{emojis.Location} {infos['where']}",
    ] if infos.get('where') else []) + [
        f"{emojis.Date} {datetime:%A} {datetime.date():%d/%m/%Y}",
        (f"{emojis.Time} {time:%H:%M} ({tz})" if chat_timezones and set(chat_timezones) != {tz} else
         f"{emojis.Time} {time:%H:%M}") if time else None
    ] + ([
        f"{emojis.Time} {datetime_tz:%H:%M} ({timezone})" if datetime_tz.date() == datetime.date() else
        f"{emojis.Time} {datetime_tz:%H:%M} on {datetime_tz.date()} ({timezone})"
        for timezone in chat_timezones or []
        if timezone != tz
        for datetime_tz in [datetime.astimezone(timezone)]
    ] if time else [])) if not relative else natural_filter([
        f"Event!",
        f"{emojis.Name} {name}",
        f"{emojis.Date} {DatetimeText.format_td_T_minus(datetime - now)}",
    ])))

def format_event_emoji_style(*, name, datetime, date, time, tz, chat_timezones):
    infos = split_event_with_where_etc({'what': name})

    emojis = EventFormatting.emojis
    return '\n'.join(natural_filter([
        f"{emojis.Name} {infos['what']}",
    ] + ([
        f"{emojis.Location} {infos['where']}",
    ] if infos.get('where') else []) + [
        f"{emojis.Date} {datetime:%A} {datetime.date():%d/%m/%Y}",
        (f"{emojis.Time} {time:%H:%M} ({tz})" if chat_timezones and set(chat_timezones) != {tz} else
         f"{emojis.Time} {time:%H:%M}") if time else None
    ] + ([
        f"{emojis.Time} {datetime_tz:%H:%M} ({timezone})" if datetime_tz.date() == datetime.date() else
        f"{emojis.Time} {datetime_tz:%H:%M} on {datetime_tz.date()} ({timezone})"
        for timezone in chat_timezones or []
        if timezone != tz
        for datetime_tz in [datetime.astimezone(timezone)]
    ] if time else [])))

async def last_event(update, context, *, relative=False):
    return await next_or_last_event(update, context, -1, relative=relative)

async def next_event(update, context, *, relative=False):
    return await next_or_last_event(update, context, 1, relative=relative)

def setting_on_off(s, default):
    return (s if isinstance(s, bool) else
            True if isinstance(s, str) and s.lower() == 'on' else
            False if isinstance(s, str) and s.lower() == 'off' else 
            setting_on_off(default, default=False) if isinstance(default, str) and default.lower() in ('on', 'off') else
            default)

def do_if_setting_on(setting):
    return setting_on_off(setting, default=False)

def do_unless_setting_off(setting):
    return setting_on_off(setting, default=True)

def simplify_multi_line_location(location):
    if '\n' in location:
        without_link = [v for v in location.splitlines() if not v /fullmatches/ "https?://[^\s]+"]
        return ' / '.join(without_link)
    else:
        return location

def enrich_location_with_db(events, *, chat_id):
    new_events = []
    for event in events:
        new_event = event # by default

        event_obj = split_event_with_where_etc({'what': event['name']})
        if 'where' in event_obj and event_obj.get('where'):
            location_key = event_obj['where']

            if match := ImplicitLocations.Parens.fullmatch(location_key):
                base_loc, extension_loc = match.groups()
                if not extension_loc.strip():
                    ok = True # matches "Something ()" (empty parens)
                else:
                    ok = False # matches "Something (Extension)"
            else:
                base_loc = location_key
                ok = True
            
            if ok:
                mapped_loc = simple_sql(("""
                    SELECT value
                    FROM EventLocation
                    WHERE chat_id=?
                    AND LOWER(key) = LOWER(?)
                    """,
                    (chat_id, base_loc)))
                
                if mapped_loc:
                    value, = only_one(mapped_loc)
                    value = simplify_multi_line_location(value)
                    
                    event_obj['where'] += f' ({value})'
                    Canon = add_event_canon_infos(infos_event=event_obj)

                    new_event = {'date': event['date'], 'rowid': event['rowid'], 'name': Canon.what_infos}

        new_events.append(new_event)
    return new_events

def get_chat_language(update):
    return get_all_chat_languages(update)[0]

def get_all_chat_languages(update):
    read_chat_settings = make_read_chat_settings(update)
    x = read_chat_settings('main.language')
    L = read_chat_settings('main.languages')
    final_list = remove_dup_keep_order([x] if x else [] + (L if L else []))
    return final_list if final_list else ['EN']

def kwarg_prop_re_match(x: None|str, s):
    import regex
    X = regex.escape(x) if x is not None else r'\p{L}+'
    if m := regex.compile(rf'[:]({X})|({X})[:]([^\s]*)', re.I).fullmatch(s):
        return m.group(1) or m.group(2), m.group(3)
    else:
        return None

def clean_kwarg_args(args) -> tuple[list[str], dict[str, list[str]]]:
    D = dict()
    bits = [0]
    for i, x in enumerate(args):
        if kwarg_prop_re_match(None, x):
            bits.append(i)
    bits.append(len(args))

    parts = []
    for a,b in zip(bits[0:], bits[1:]):
        parts.append(args[a:b])
    
    D = {}
    for p in parts[1:]:
        mk, mv = kwarg_prop_re_match(None, p[0])
        D[mk] = p[1:] if not mv else [mv] + p[1:]
    return parts[0], D

async def list_days_or_today(
        update: Update,
        context: CallbackContext,
        mode: Literal['list', 'today', 'tomorrow', 'dayofweek'],
        mode_args={},
        relative=False,
        formatting:Literal['normal', 'linkdays', 'crazyjamdays', 'linkdayshtml', 'short', 'shorthtml']='normal'):
    """
    @param mode: 
        - list(): read the arguments from args
        - today(): for today, will put a "→" to mark past events
        - tomorrow(): for tomorrow
        - dayofweek(dayofweek=): example: "Monday" (dayofweek = 1)

    @param formatting: describe how each event is displayed
        - normal
        - linkdays: (?)
        - crazyjamdays: the format used by crazy jam
        - linkdayshtml: each event is possibily a link

    @param relative: display dates as "H-3" for example
    """
    from datetime import time as Time
    
    assert mode in ('list', 'today', 'tomorrow', 'dayofweek')
    if mode == 'dayofweek':
        assert (dayofweek := mode_args.get('dayofweek'))
        assert dayofweek in irange(1, 7)

    assert formatting in ('normal', 'linkdays', 'crazyjamdays', 'linkdayshtml', 'short', 'shorthtml')
    if with_link := formatting in ('linkdays', 'crazyjamdays', 'linkdayshtml', 'shorthtml'):
        assert not relative

    if with_html := formatting in ('linkdayshtml', 'shorthtml'):
        send = make_send(update, context, parse_mode='HTML', disable_web_page_preview=True)
    else:
        send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)

    do_event_admin_check('list', setting=read_chat_settings('event.admins'), user_id=update.effective_user.id)

    language = get_chat_language(update)

    chat_id = update.effective_chat.id

    args, kwargs = clean_kwarg_args(context.args)

    if kwargs.get('tz'):
        if kwargs['tz'] == ['chat']:
            tzs = read_chat_settings('event.timezones')
        elif kwargs['tz'] == ['revchat']:
            tzs = read_chat_settings('event.timezones')
            if tzs:
                tzs = list(reversed(tzs))
        else:
            tzs = list(map(partial(ZoneInfoOrAlias, chat_id=chat_id), kwargs['tz']))
    else:
        if do_if_setting_on(read_chat_settings('event.list.display_all_timezones')):
            tzs = read_chat_settings('event.timezones')
        else:
            tzs = None

    real_args = (args if mode == 'list' else
                 ('today',) if mode == 'today' else
                 ('tomorrow',) if mode == 'tomorrow' else 
                 (DatetimeText.days_english[dayofweek-1], ) if mode == 'dayofweek' else raise_error(AssertionError('mode must be a correct value')))

    datetime_range = parse_datetime_range(update, args=real_args, tz=tzs[0] if tzs else None)
    beg, end, tz, when = (datetime_range[x] for x in ('beg_utc', 'end_utc', 'tz', 'when'))
    tz = tzs[0] if tzs else tz

    strptime = DatetimeDbSerializer.strptime
    strftime = DatetimeDbSerializer.strftime

    events = simple_sql_dict(('''
        SELECT date, name, rowid
        FROM Events
        WHERE ? <= date AND date < ?
        AND chat_id=?
        ORDER BY date''',
        (strftime(beg), strftime(end), update.effective_chat.id,)))
    
    if with_link: 
        list_of_ids = [events['rowid'] for events in events]
        list_of_ids = ','.join(map(str, map(int, list_of_ids)))
        link_of_event = dict(simple_sql_dict((f'''
            SELECT event_id, link
            FROM EventLinkAttr
            WHERE event_id IN ({list_of_ids}) ''', ())))
    
    if do_unless_setting_off(read_chat_settings('event.location.autocomplete')):
        events = enrich_location_with_db(events, chat_id=update.effective_chat.id)

    from collections import defaultdict
    days = defaultdict(list)
    for event in events:
        date = strptime(event['date']).replace(tzinfo=ZoneInfo("UTC")).astimezone(tz)
        event_name = event['name']

        if with_link:
            event_link = link_of_event.get(event['rowid'])
        
        if with_html and event_link:
            if not event_link.startswith(('http://', 'https://')):
                event_link = 'https://' + event_link

        days[date.timetuple()[:3]].append((date, event_name) if formatting in ('normal', 'short') else (date, event_name, event_link))

    display_time_marker = (False if mode in ('list', 'tomorrow', 'dayofweek', ) else
                           do_unless_setting_off(read_chat_settings('event.listtoday.display_time_marker')) if mode == 'today' else 
                           raise_error(AssertionError))

    now_tz = datetime.now().astimezone(tz)
    def is_past(event_date):
        return event_date <= now_tz
    
    days_as_lines = []
    for day in sorted(days):
        date = days[day][0][0].date()
        if formatting == 'crazyjamdays':
          day_of_week = DatetimeText.days_short.in_lang(language)[date.weekday()]
          month_ru = DatetimeText.padezh_month(date.month, date.day)
          days_as_lines.append(
            f"\n    {date:%d} {month_ru} ({day_of_week})\n\n"
            + "\n\n".join(
                 f"{n}) {event_link}\n{event_name}" if event_link else 
                 f"{n}) {event_name}"
                for n, (event_date, event_name, event_link) in enumerate(days[day], start=1)
            ))
        elif formatting == 'linkdays':
          day_of_week = DatetimeText.days.in_lang(language)[date.weekday()]
          days_as_lines.append(
            f"\n    {date:%d/%m} ({day_of_week.capitalize()})\n\n"
            + "\n\n".join(
                 f"{n}) {event_link}\n{event_name}" if event_link and event_date.time() == Time(0, 0) else 
                 f"{n}) {event_link}\n{event_date:%H:%M}: {event_name}" if event_link and event_date.time() != Time(0, 0) else 
                 f"{n}) {event_name}" if event_date.time() == Time(0, 0) else 
                 f"{n}) {event_date:%H:%M}: {event_name}"
                for n, (event_date, event_name, event_link) in enumerate(days[day], start=1)
            ))
        elif formatting == 'linkdayshtml':
            day_of_week = DatetimeText.days.in_lang(language)[date.weekday()]
            days_as_lines.append(
            f"{day_of_week.capitalize()} {date:%d/%m}\n"
            + "\n".join(
                 f"""- <a href="{html.escape(event_link)}">{html.escape(event_name)}</a>""" if event_link and event_date.time() == Time(0, 0) else 
                 f"""- <a href="{html.escape(event_link)}">{event_date:%H:%M}: {html.escape(event_name)}</a>""" if event_link and event_date.time() != Time(0, 0) else 
                 f"""- {html.escape(event_name)}""" if event_date.time() == Time(0, 0) else 
                 f"""- {event_date:%H:%M}: {html.escape(event_name)}"""
                for n, (event_date, event_name, event_link) in enumerate(days[day], start=1)
            ))
        elif formatting in ('short', ):
            day_of_week = DatetimeText.days.in_lang(language)[date.weekday()]
            days_as_lines.append(
                f"{day_of_week.capitalize()} {date:%d/%m}\n"
                + "\n".join(
                    f"""- {event_pure_name}"""
                    for event_date, event_name in days[day]
                    for (event_pure_name, event_location) in [split_event_name_into_what_where(event_name)]
                )
            )
        elif formatting in ('shorthtml', ):
            day_of_week = DatetimeText.days.in_lang(language)[date.weekday()]
            days_as_lines.append(
                f"{day_of_week.capitalize()} {date:%d/%m}\n"
                + "\n".join(
                    f"""- <a href="{html.escape(event_link)}">{html.escape(event_pure_name)}</a>"""
                    for event_date, event_name, event_link in days[day]
                    for (event_pure_name, event_location) in [split_event_name_into_what_where(event_name)]
                )
            )
        else:
          day_of_week = DatetimeText.days.in_lang(language)[date.weekday()]
          days_as_lines.append(
            f"{day_of_week.capitalize()} {date:%d/%m}"
            + "\n"
            + "\n".join(
                f"-{marker} %s: {event_name}" % ' | '.join(f"{d:%H:%M}" for d in (event_date.astimezone(tz) for tz in tzs)) if tzs else
                f"-{marker} {event_date:%H:%M}: {event_name}" if not relative else
                f"-{marker} {DatetimeText.format_td_T_minus(event_date - now_tz)}: {event_name}"
                for event_date, event_name in days[day]
                for marker in ['>' if display_time_marker and is_past(event_date) else '']))
    
    msg = ('\n' if with_link and not with_html else '\n\n').join(days_as_lines)
    
    if formatting == 'crazyjamdays':
        msg = 'CRAZY JAM\n' + msg
    elif formatting == 'linkdays':
        msg = 'Events:\n' + msg

    chat_timezones = read_chat_settings("event.timezones")

    if msg and chat_timezones and set(chat_timezones) != {tz}:
        msg += '\n\n' + (f'Timezone: {tz}' if not tzs or len(tzs) == 1 else f"Timezones: {' '.join(map(str, tzs))}")

    await send(msg or (
        "No events for the next 7 days !" if when == 'week' else
        f"No events for {when} !" + (" 😱" if "today" in (mode, when) else "")
    ))

async def day_of_week_command(update, context, n):
    assert n in irange(1, 7)
    return await list_days_or_today(update, context, mode='dayofweek', mode_args={'dayofweek': n})

list_today = partial(list_days_or_today, mode='today')
list_days = partial(list_days_or_today, mode='list')

async def list_events(update: Update, context: CallbackContext, relative=False):
    send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)

    if relative:
        raise UserError("Not implemented yet, use /rlistdays for example")

    do_event_admin_check('list', setting=read_chat_settings('event.admins'), user_id=update.effective_user.id)

    datetime_range = parse_datetime_range(update, args=context.args)
    beg, end, tz, when = (datetime_range[x] for x in ('beg_utc', 'end_utc', 'tz', 'when'))

    chat_id = update.effective_chat.id
    with sqlite3.connect('db.sqlite') as conn:
        strptime = DatetimeDbSerializer.strptime
        strftime = DatetimeDbSerializer.strftime
        
        cursor = conn.cursor()
        query = ("""SELECT date, name
                    FROM Events
                    WHERE ? <= date AND date < ?
                    AND chat_id = ?
                    ORDER BY date""",
                (strftime(beg), strftime(end), chat_id))
        
        chat_timezones = read_chat_settings("event.timezones")
        msg = '\n'.join(f"- {DatetimeText.days_english[date.weekday()]} {date:%d/%m}: {event}" if not has_hour else 
                        f"- {DatetimeText.days_english[date.weekday()]} {date:%d/%m %H:%M}: {event}"
                        for date_utc, event in cursor.execute(*query)
                        for date in [strptime(date_utc).replace(tzinfo=ZoneInfo('UTC')).astimezone(tz)]
                        for has_hour in [True])
        if msg and chat_timezones and set(chat_timezones) != {tz}:
            msg += '\n\n' + f"Timezone: {tz}"
        await send(msg or (
            "No events for the next 7 days !" if when == 'week' else
            f"No events for {when} !" + (" 😱" if "today" == when else "")
        ))

async def delevent_from_answer(*, reply, update, context):
    event = addevent_analyse_from_bot(update, context, text=reply.text)
    event_db = retrieve_event_from_db(update=update, context=context, what=event['what'], when=event['when'])
    tz = induce_my_timezone(user_id=update.effective_user.id, chat_id=update.effective_chat.id)
    send = make_send(update, context)
    await db_delete_event(update, context, send=send, chat_id=update.effective_chat.id, event_id=event_db['rowid'], tz=tz)

async def delevent(update, context):
    send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)

    do_event_admin_check('del', setting=read_chat_settings('event.admins'), user_id=update.effective_user.id)

    if reply := update_get_reply(update):
        await delevent_from_answer(reply=reply, update=update, context=context)
        return ConversationHandler.END

    strptime = DatetimeDbSerializer.strptime

    strftime = DatetimeDbSerializer.strftime
    strftime_minutes = DatetimeDbSerializer.strftime_minutes

    datetime_range = parse_datetime_range(update, args=context.args, default="future")
    beg, end, tz = datetime_range['beg_utc'], datetime_range['end_utc'], datetime_range['tz']
    events = simple_sql_dict(('''
        SELECT rowid, date, name
        FROM Events
        WHERE ? <= date AND date < ?
        AND chat_id=?
        ORDER BY date''',
        (strftime(beg), strftime(end), update.effective_chat.id,)))
   
    saved_info_dict: dict = make_send_save_info(update, context)._asdict()

    keyboard = [
        [InlineKeyboardButton("{} - {}".format(
            strftime_minutes(strptime(event['date']).replace(tzinfo=ZoneInfo("UTC")).astimezone(tz)),
            event['name']
        ), callback_data=json.dumps(saved_info_dict | dict(rowid=str(event['rowid']))))]
        for event in events
    ]

    if not keyboard:
        await send("No events to delete !")
        return ConversationHandler.END
    
    cancel = [[InlineKeyboardButton("/cancel", callback_data=json.dumps(saved_info_dict | dict(rowid="null")))]]

    await send("Choose an event to delete:", reply_markup=InlineKeyboardMarkup(keyboard + cancel))

    return 0

async def do_delete_event(update, context):
    query = update.callback_query
    user = query.from_user
    await query.answer()

    data_dict: dict = json.loads(query.data)
    send = make_send(update, context, save_info=SendSaveInfo(chat_id=data_dict['chat_id'], thread_id=data_dict['thread_id']))
    
    rowid = data_dict["rowid"]
    if rowid == "null":
        await send("Cancelled: No event deleted")
    else:
        await db_delete_event(update, context, send, chat_id=update.effective_chat.id, event_id=rowid, tz=induce_my_timezone(user_id=user.id, chat_id=data_dict.get('chat_id')))

    # await query.edit_message_text
    # await query.edit_message_reply_markup()
    await query.delete_message()
        
    return ConversationHandler.END

async def selectevent(update, context):
    send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)

    do_event_admin_check('list', setting=read_chat_settings('event.admins'), user_id=update.effective_user.id)

    datetime_range = parse_datetime_range(update, args=context.args, default="future")
    beg, end, tz = datetime_range['beg_utc'], datetime_range['end_utc'], datetime_range['tz']
    events = simple_sql_dict(('''
        SELECT rowid, date, name
        FROM Events
        WHERE ? <= date AND date < ?
        AND chat_id=?
        ORDER BY date''',
        (DatetimeDbSerializer.strftime(beg), DatetimeDbSerializer.strftime(end), update.effective_chat.id,)))

    saved_info_dict: dict = make_send_save_info(update, context)._asdict()

    keyboard = [
        [InlineKeyboardButton("{} - {}".format(
            DatetimeDbSerializer.strftime_minutes(DatetimeDbSerializer.strptime(event['date']).replace(tzinfo=UTC).astimezone(tz)),
            event['name']
        ), callback_data=json.dumps(saved_info_dict | dict(rowid=str(event['rowid']))))]
        for event in events
    ]

    if not keyboard:
        await send("No events to select !")
        return ConversationHandler.END

    cancel = [[InlineKeyboardButton("/cancel", callback_data=json.dumps(saved_info_dict | dict(rowid="null")))]]

    await send("Choose an event:", reply_markup=InlineKeyboardMarkup(keyboard + cancel))

    return 0

async def do_selectevent(update, context):
    query = update.callback_query
    user = query.from_user
    await query.answer()

    data_dict: dict = json.loads(query.data)
    send = make_send(update, context, save_info=SendSaveInfo(chat_id=data_dict['chat_id'], thread_id=data_dict['thread_id']))

    rowid = data_dict["rowid"]
    if rowid == "null":
        await send("Cancelled: No event selected")
    else:
        await send(format_event_emoji_style_from_event_id(rowid, chat_id=query.message.chat.id, user_id=query.from_user.id))
    
    await query.delete_message()

    return ConversationHandler.END

async def db_delete_event(update, context, send, *, chat_id, event_id, tz):
    read_chat_settings = make_read_chat_settings(update, context)

    if read_chat_settings('event.delevent.display'):
        infos = dict(only_one(simple_sql_dict(('select date, name from Events where chat_id = ? and rowid = ?', (chat_id, event_id, )))))
    else:
        infos = None

    date_tz = None if not(infos and infos.get('date') and tz) else DatetimeDbSerializer.strftime(DatetimeDbSerializer.strptime(infos.get('date')).replace(tzinfo=UTC).astimezone(tz))

    do_event_admin_check('del', setting=read_chat_settings('event.admins'), user_id=update.effective_user.id)

    simple_sql(('delete from Events where chat_id = ? and rowid = ?', (chat_id, event_id)))
    
    await send(f"Event deleted" if infos is None else "Event deleted: {}".format(dict(date=date_tz, name=infos.get('name'))))

def n_to_1_dict(x:dict|Iterable):
    gen = x.items() if isinstance(x, dict) else x
    
    def is_cool_iterable(it):
        import collections.abc
        return (isinstance(it, collections.abc.Sequence) 
          and not isinstance(it, str))

    d = {}
    for key, value in gen:
        if is_cool_iterable(key):
            for v in key:
                d[v] = value
        else:
            d[key] = value
    return d

def fetch_event(key):
    return None

async def timedifference(update, context, command):
    from datetime import timedelta, datetime
    send = make_send(update, context)
    conversions = n_to_1_dict({
        ('minutes', 'min', 'minute'): timedelta(minutes=1), 
        ('days', 'day'): timedelta(days=1),
        ('h', 'hours', 'hour'): timedelta(hours=1),
        ('s', 'seconds', 'second'): timedelta(seconds=1),
        ('weeks', 'week'): timedelta(weeks=1),
    })
    is_timedelta = re.compile('|'.join(map(re.escape, conversions)))
    di = next((i for i, x in enumerate(context.args) if is_timedelta.fullmatch(x)), None)
    units = ('days' if di is None else
             context.args[di])
    dtunits = conversions[units]
    eventkey, = context.args
    event: datetime = fetch_event(eventkey) or DatetimeText.to_datetime_range(eventkey)[0]
    delta = event - datetime.now()
    if command == 'timesince':
        delta = -delta
    await send('{:.2f} {}'.format(delta / dtunits, units))

def ZoneInfoOrAlias(tz, *, chat_id):
    try:
        real, = only_one_specific(simple_sql((''' select real from TimezoneAlias where chat_id=? and lower(alias) = lower(?)''', (chat_id, tz))))
        return ZoneInfo(real)
    except NoRecords:
        return ZoneInfo(tz)

async def timein(update, context):
    from datetime import datetime
    send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)

    if not context.args:
        if not (tzs := read_chat_settings('event.timezones')):
            return await send("Usage: /timein [timezone]\nExample: /timein Europe/Brussels")
    else:
        tzs = None

    if len(context.args) == 1:
        tz_str, = context.args
    elif len(context.args) == 2:
        tz_str = '/'.join(context.args)
    elif len(context.args) > 2:
        raise ValueError("Too much arguments")
    
    if tzs is None:
        try:
            tz = ZoneInfoOrAlias(tz_str, chat_id=update.effective_chat.id)
        except (ZoneInfoNotFoundError, IsADirectoryError):
            raise UserError(f"{context.args[0]!r} is not a timezone")
        
        dt = datetime.now().astimezone(tz).replace(tzinfo=None)

        return await send(f"{dt:%H:%M} on {dt.date():%d/%m/%Y}")
    else:
        def get_dt(tz):
            return datetime.now().astimezone(tz).replace(tzinfo=None)

        def dt_tz_format_old(dt, tz):
            return "{} ({})".format(f"{dt:%H:%M} on {dt.date():%d/%m/%Y}", tz)

        def dt_tz_format_new(dt, tz):
            return "  {} ({})".format(f"{dt:%H:%M}", tz)

        def dt_tz_format_group(date):
            return f"{date:%d/%m/%Y} ({DatetimeText.days_english[date.weekday()].capitalize()})"

        dts = list(map(get_dt, tzs))
        data = sorted(zip(dts, tzs))

        lines = []
        for k, group in itertools.groupby(data, key=lambda t:t[0].date()):
            lines.append(dt_tz_format_group(k))
            lines.extend(dt_tz_format_new(dt, tz) for dt, tz in group)

        return await send('\n'.join(lines))

async def timeuntil(update, context):
    return await timedifference(update, context, command='timeuntil')

async def timesince(update, context):
    return await timedifference(update, context, command='timesince')

async def deletevent(update, context):
    send = make_send(update, context)
    key, = context.args
    await send("Not implemented yet!")

def get_my_timezone_from_timezone_table(user_id) -> ZoneInfo:
    query = ("""SELECT timezone FROM UserTimezone WHERE user_id=?""", (user_id,))
    with sqlite3.connect('db.sqlite') as conn:
        L = conn.execute(*query).fetchall()
        if len(L) == 0:
            return None
        elif len(L) == 1:
            return ZoneInfo(L[0][0])
        else:
            raise ValueError("Unique constraint failed: Multiple timezone for user {}".format(user_id))

def get_my_timezone(user_id) -> ZoneInfo:
    return (read_settings('event.timezone', id=user_id, settings_type='user')
            or get_my_timezone_from_timezone_table(user_id))

def set_my_timezone(user_id, tz:ZoneInfo):
    query_read = ("""SELECT timezone FROM UserTimezone WHERE user_id=?""", (user_id,))
    query_update = ("""UPDATE UserTimezone SET timezone=? WHERE user_id=?""", (str(tz), user_id))
    query_insert = ("""INSERT INTO UserTimezone(timezone, user_id) VALUES(?, ?)""", (str(tz), user_id))
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        L = conn.execute(*query_read).fetchall()
        if len(L) == 0:
            conn.execute(*query_insert)
        elif len(L) == 1:
            conn.execute(*query_update)
        else:
            raise ValueError("Unique constraint failed: Multiple timezone for user {}".format(user_id))
        conn.execute("end transaction")

def set_settings(*, id, key, value_raw:any, settings_type:Literal['chat'] | Literal['user'], list_type_and_extend:bool):
    conversion = CONVERSION_SETTINGS[settings_type][key]['to_db']

    value: any = conversion(value_raw)

    if list_type_and_extend:
        from_db = read_raw_settings(key=key, id=id, settings_type=settings_type)
        value = json.dumps(json.loads(from_db) + json.loads(value))

    table = SettingsInfo.TABLES[settings_type]
    field_id = SettingsInfo.FIELDS[settings_type]
    query_read = (f"""SELECT value FROM {table} WHERE {field_id}=? and key=?""", (id, key))
    query_update = (f"""UPDATE {table} SET value=? WHERE {field_id}=? and key=?""", (value, id, key))
    query_insert = (f"""INSERT INTO {table}({field_id}, key, value) VALUES(?, ?, ?)""", (id, key, value))
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        L = conn.execute(*query_read).fetchall()
        if len(L) == 0:
            conn.execute(*query_insert)
        elif len(L) == 1:
            conn.execute(*query_update)
        else:
            raise ValueError("Unique constraint failed: Multiple settings for {} {} and key {!r}".format(settings_type, id, key))
        conn.execute("end transaction")

def delete_settings(*, id, key, settings_type:Literal['chat'] | Literal['user']):
    table = SettingsInfo.TABLES[settings_type]
    field_id = SettingsInfo.FIELDS[settings_type]
    query_delete = (f"""DELETE FROM {table} WHERE {field_id}=? and key=?""", (id, key))
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute(*query_delete)

async def mytimezone(update: Update, context: CallbackContext):
    send = make_send(update, context)

    if not context.args:
        # get timezone
        tz = get_my_timezone_from_timezone_table(update.effective_user.id)
        base_text = ("You don't have any timezone set.\n"
                     "Use /mytimezone Continent/City to set it.\n"
                     "Example: /mytimezone Europe/Brussels\n"
                     "Example: /mytimezone America/Los_Angeles" if tz is None else
                     "Your timezone is: {}".format(tz))
        return await send(base_text)
    else:
        # set timezone
        default_error = UserError("This timezone is not known by the system.\nCorrect examples include:\n- America/Los_Angeles\n- Europe/Brussels")
        for tries in (1, 2):
            match tries:
                case 1:
                    tz_name, *_ = context.args
                case 2:
                    if len(context.args) < 2:
                        raise default_error
                    tz_continent, tz_city, *_ = context.args
                    tz_name = tz_continent + "/" + tz_city
            try:
                tz = ZoneInfoOrAlias(tz_name, chat_id=update.effective_chat.id)
                break
            except ZoneInfoNotFoundError:
                continue
            except Exception as e:
                if isinstance(e, IsADirectoryError):
                    continue
                else:
                    raise e
        else:
            raise default_error
        set_my_timezone(update.effective_user.id, tz)
        return await send("Your timezone is now: {}".format(tz))

async def timezonealias(update: Update, context: CallbackContext):
    send = make_send(update, context)
    chat_id = update.effective_chat.id

    if not context.args:
        # get timezone
        return await send("Usage: /timezonealias abc timezone")
    elif len(context.args) == 1:
        alias, = context.args
        tz = None
    elif len(context.args) == 2:
        alias, tz = context.args
    elif len(context.args) == 3:
        alias, tz_continent, tz_city = context.args
        tz = tz_continent + "/" + tz_city
    else:
        raise UserError("Too much arguments")
    
    if tz is None:
        # read
        try:
            return await send(get_timezonealias_from_table(update, context, alias=alias, chat_id=chat_id))
        except NoRecords:
            return await send(f"No timezone for {alias!r}")
    else:
        # write
        if tz.lower() == 'delete':
            pass # ok
        else:
            try:
                ZoneInfo(tz)
            except ZoneInfoNotFoundError:
                raise UserError(f"{tz!r} is not a timezone")
        set_timezonealias_to_table(update, context, chat_id=chat_id, alias=alias, real=tz)
        return await send(f"Alias {alias!r} saved")

class TooManyRecords(ValueError):
    pass

class NoRecords(ValueError):
    pass

def get_timezonealias_from_table(update, context, chat_id, alias):
    return only_one(
        only_one(
            simple_sql(('''select real from TimezoneAlias where chat_id = ? and alias = ?''', (chat_id, alias))),
            none=NoRecords,
            many=TooManyRecords,
        )
    )

def set_timezonealias_to_table(update, context, chat_id, alias, real):
    if real.lower() == 'delete':
        conn.execute('''delete from TimezoneAlias where chat_id=? and LOWER(?)=LOWER(alias)''', (chat_id, alias, ))
        return 
    
    # check correct TZ
    ZoneInfo(real)

    # db
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        if only_one(only_one(conn.execute('''select count(*) from TimezoneAlias where LOWER(?)=LOWER(alias) and chat_id=?''', (alias, chat_id)))) == 0:
            # insert
            conn.execute('''insert into TimezoneAlias(chat_id, alias, real) VALUES (?,?,?)''', (chat_id, alias, real))
        else:
            # update
            conn.execute('''update TimezoneAlias set real=? where chat_id=? and LOWER(?)=LOWER(alias)''', (real, chat_id, alias))
        conn.execute('end transaction')


def to_list(gen):
    def f(*a, **b):
        return list(gen(*a, **b))
    return f

@to_list
def remove_dup_keep_order(it):
    S = set()
    for x in it:
        if x not in S:
            S.add(x)
            yield x

async def menu(update, context):
    send = make_send(update, context)
    keyboard = [
        [
            InlineKeyboardButton("Help", callback_data="cmd:help")
        ],
        [
            InlineKeyboardButton("My timezone", callback_data="cmd:mytimezone"),
            InlineKeyboardButton("Chat timezone", callback_data="cmd:chattimezone"),
        ],
        [
            InlineKeyboardButton("User Settings", callback_data="cmd:mysettings"),
            InlineKeyboardButton("Delete user settings", callback_data="cmd:delsettings_command")
        ],
        [
            InlineKeyboardButton("Chat settings", callback_data="cmd:chatsettings"),
            InlineKeyboardButton("Delete chat settings", callback_data="cmd:delchatsettings")
        ],
        [
            InlineKeyboardButton("List events", callback_data="cmd:listdays"),
            InlineKeyboardButton("Next event", callback_data="cmd:nextevent"),
            InlineKeyboardButton("Last event", callback_data="cmd:lastevent")
        ],
        [
            InlineKeyboardButton("Today's events", callback_data="cmd:today"),
            InlineKeyboardButton("Tomorrow's events", callback_data="cmd:tomorrow")
        ],
        [
            InlineKeyboardButton("Add event", callback_data="cmd:addevent"),
            InlineKeyboardButton("Interactive Add event", callback_data="cmd:iaddevent"),
        ],
        [
            InlineKeyboardButton("Select event", callback_data="cmd:selectevent"),
            InlineKeyboardButton("Delete event", callback_data="cmd:delevent"),
        ],
        [
            InlineKeyboardButton("Event: Follow another chat", callback_data="cmd:eventfollow"),
            InlineKeyboardButton("Event: Accept a follower", callback_data="cmd:eventacceptfollow")
        ],
        [
            InlineKeyboardButton("Event: Delete a chat you follow", callback_data="cmd:deleventfollow"),
            InlineKeyboardButton("Event: Delete a follower", callback_data="cmd:deleventacceptfollow"),
        ],
        [
            InlineKeyboardButton("Close menu", callback_data="cmd:closemenu"),
        ],
    ]
    await send("The main menu\nChoose a command", reply_markup=InlineKeyboardMarkup(keyboard))

async def menu_button_handler(update, context):
    query = update.callback_query
    await query.answer()
    if not query.data.startswith("cmd:"):
        return
    cmd = query.data.split(":", 1)[1]

    @dataclass
    class HandlerInfo:
        function: Any
        params: list

    NoHandler = object()

    handlers = {
        'help': help,
        'mytimezone': mytimezone,
        'mysettings': mysettings,
        'delsettings_command': HandlerInfo(
            partial(settings_command, accepted_settings=ACCEPTED_SETTINGS_USER, settings_type='user', command_name='mysettings delete'),
            ['delete'],
        ),
        'delchatsettings': HandlerInfo(
            partial(settings_command, accepted_settings=ACCEPTED_SETTINGS_CHAT, settings_type='chat', command_name='chatsettings delete'),
            ['delete'],
        ),
        'addevent': add_event,
        'addschedule': addschedule,
        'delevent': NoHandler,  # delevent,
        'iaddevent': NoHandler, # InteractiveAddEvent.ask_when,
        'selectevent': NoHandler,
        'nextevent': next_event,
        'lastevent': last_event,
        'listdays': list_days,
        'listevents': list_events,
        'listtoday': list_today,
        'today': list_today,
        'tomorrow': partial(list_days_or_today, mode='tomorrow', relative=False),
        'timezonealias': timezonealias,
        'listallsettings': listallsettings,
        'chatsettings': chatsettings,
        # 'delchatsettings': del_chat_settings,
        'chattimezone': HandlerInfo(
            chatsettings,
            ['event.timezones'],
        ),
        'closemenu': None,
        'menu': menu
    }

    if cmd == 'closemenu':
        await query.delete_message()
        return

    handler_info = handlers.get(cmd)
    if not handler_info:
        return await make_send(update, context)(f"/{cmd} is not supported")
    
    if handler_info is NoHandler:
        send = make_send(update, context)
        return await send("Not Implemented yet")

    if isinstance(handler_info, HandlerInfo):
        fn, new_args = handler_info.function, handler_info.params or []
    else:
        fn, new_args = handler_info, []

    old_args = getattr(context, 'args', None)
    context.args = new_args 
    try:
        await fn(update,context)
    finally:
        context.args = old_args


@dataclass
class SettingsSpecs:
    type: Literal['single', 'list', 'bool'] = 'single'
    default: Optional[object] = None

ACCEPTED_SETTINGS_USER = {
    'event.timezone': SettingsSpecs('single'),
    'wikt.text': SettingsSpecs('single'),
    'wikt.description': SettingsSpecs('single'),
    'wikt.html': SettingsSpecs('bool'),
    'larousse.text': SettingsSpecs('single'),
    'larousse.description': SettingsSpecs('single'),
    'larousse.html': SettingsSpecs('bool'),
    'dict.text': SettingsSpecs('single'),
    'dict.description': SettingsSpecs('single'),
    'dict.engine': SettingsSpecs('single'),
    'dict.html': SettingsSpecs('bool'),
}
ACCEPTED_SETTINGS_CHAT = {
    'money.currencies': SettingsSpecs('list'),
    'money.known_currencies': SettingsSpecs('list'),
    'event.timezones': SettingsSpecs('list'),
    'event.admins': SettingsSpecs('list'),
    'event.addevent.help_file': SettingsSpecs('bool'),
    'event.addevent.display_link': SettingsSpecs('bool'),
    'event.addevent.display_file': SettingsSpecs('bool'),
    'event.addevent.display_forwarded_infos': SettingsSpecs('bool'),
    'event.addevent.required_time': SettingsSpecs('bool'),
    'event.list.display_all_timezones': SettingsSpecs('bool'),
    'event.listtoday.display_time_marker': SettingsSpecs('bool'),
    'event.delevent.display': SettingsSpecs('bool'),
    'event.location.autocomplete': SettingsSpecs('bool'),
    'event.commands.dayofweek': SettingsSpecs('bool'),
    'sharemoney.required_for': SettingsSpecs('bool'),
    'list.space_between_lines': SettingsSpecs('bool'),
    'list.indent': SettingsSpecs('single'),
    'main.language': SettingsSpecs('single'),
    'main.languages': SettingsSpecs('list'),
} | {
    setting + '.active': SettingsSpecs('bool', default) for _, setting, default in RESPONDERS
}

assert all(isinstance(x, SettingsSpecs) for x in (ACCEPTED_SETTINGS_USER | ACCEPTED_SETTINGS_CHAT).values()), "All settings must have a valid type Specification"
assert not(ACCEPTED_SETTINGS_USER.keys() & ACCEPTED_SETTINGS_CHAT.keys()), "Some settings are both user and chat settings, which is not allowed"

def assert_true(condition, error=AssertionError):
    if not condition:
        raise error
    return True

def is_timezone(x: str) -> bool:
    try:
        ZoneInfo(x)
        return True
    except ZoneInfoNotFoundError:
        return False

@dataclass
class EventAdmin:
    user_id: int | 0
    local_name: str
    permissions: list[Literal["add", "del", "edit", "list"]]
    # if add & del → edit
    # if add | del → list
    # if "*" → add, del, edit, list
    unknown = False

    def __init__(self, user_id:int | 0, local_name='', permissions=None):
        self.user_id = user_id
        self.local_name = local_name or ''
        self.permissions = permissions if permissions is not None else ['*']

        self.add_implicit_permissions()

        assert set(self.permissions) <= {'add', 'del', 'edit', 'list'}, str(self.permissions)

    def add_implicit_permissions(self):
        while True:
            S = set(self.permissions)

            if {'0'} <= S:
                S = set()

            if {'add', 'del'} <= S:
                S |= {"edit"}
            
            if {'add', 'del'} <= S:
                S |= {"list"}

            if {'*'} <= S:
                S |= {"add", "del", "edit", "list"}
                S -= {'*'}

            if set(self.permissions) == S:
                break

            self.permissions = sorted(S)
            
        return self

    def to_json(self):
        J = {
            'user_id': int(self.user_id),
            'local_name': str(self.local_name) if self.local_name else '',
            'permissions': list(map(str, self.permissions)),
            'unknown': self.unknown,
        }
        
        if not J.get('local_name'):
            del J['local_name']
        
        if J['unknown'] == False:
            del J['unknown']
        
        if J.get('permissions') == sorted(('*', 'list', 'add', 'del', 'edit')) or J.get('permissions') == sorted(('list', 'add', 'del', 'edit')):
            J['permissions'] = ['*']

        if J['permissions'] == []:
            J['permissions'] = ['0']
        
        if J['permissions'] == ['*'] and not J.get('local_name'):
            return int(J['user_id'])
        else:
            return J

    @staticmethod
    def from_json(J):
        if isinstance(J, int) and J == 0:
            return EventAdmin(user_id=0)
        
        if isinstance(J, int) or isinstance(J, str) and J.isdecimal():
            return EventAdmin(user_id=J)
        
        return EventAdmin(**{
            'user_id': int(J['user_id']),
            'local_name': str(J['local_name']) if J.get('local_name') else '',
            'permissions': list(map(str, J['permissions'])),
        })

def CONVERSION_SETTINGS_BUILDER():
    import json
    # serializers helper
    def list_of(obj):
        return {
            'from_db': lambda s: list(map(obj['from_db'], json.loads(s))),
            'to_db': lambda L: json.dumps(list(map(obj['to_db'], L))),
        }
    
    # serializers
    default_serializer = {
        'from_db': lambda x:x,
        'to_db': lambda x:x,
    }
    json_serializer = {
        'from_db': json.loads,
        'to_db': json.dumps,
    }
    timezone_serializer = {
        'from_db': ZoneInfo,
        'to_db': lambda x: assert_true(is_timezone(x), UserError(f"{x} is not a timezone"))
                 and x,
    }
    currency_serializer = {
        'from_db': str.upper,
        'to_db': str.upper,
    }
        
    int_serializer = {
        'from_db': int,
        'to_db': int,
    }
    list_of_event_admins = list_of({
        'from_db': EventAdmin.from_json,
        'to_db': lambda x: (
            EventAdmin(user_id=int(x), permissions=['0']) if x.isdecimal() and int(x) == 0 else
            EventAdmin(user_id=int(x), permissions=['*']) if x.isdecimal() else
            EventAdmin(user_id=int(x.split(':')[0]), permissions=x.split(':')[1].split(",")) if ':' in x else raise_error(ValueError)
        ).to_json(),
    })
    on_off_serializer = {
        'from_db': lambda x: x != 'off',
        'to_db': lambda x: assert_true(isinstance(x, str) and x.lower() in ('on', 'off', 'true', 'false', 'yes', 'no'), UserError(f"{x} must be on/off"))
                 and {'true': 'on', 'false': 'off', 'yes': 'on', 'no': 'off'}.get(x.lower(), x.lower()),
    }
    # mappings
    mapping_chat = {
        'money.currencies': list_of(currency_serializer),
        'money.known_currencies': list_of(currency_serializer),
        'event.timezones': list_of(timezone_serializer),
        'event.admins': list_of_event_admins,
        'list.indent': int_serializer,
        'main.languages': list_of(default_serializer),
    } | {
        name: on_off_serializer
        for name, specs in ACCEPTED_SETTINGS_CHAT.items()
        if specs.type == 'bool'
    }
    
    mapping_user = {
        'event.timezone': timezone_serializer,
        'dict.engine': {
            'from_db': lambda x: x,
            'to_db': lambda x: x if assert_true(x in DICT_ENGINES, UserError("{!r} is not a known engine.\nAvaiblable options: {}".format(x, ', '.join(DICT_ENGINES)))) else None
        }
    } | {
        name: on_off_serializer
        for name, specs in ACCEPTED_SETTINGS_USER.items()
        if specs.type == 'bool'
    }

    from collections import defaultdict
    return {
        'chat': defaultdict(lambda: default_serializer, mapping_chat),
        'user': defaultdict(lambda: default_serializer, mapping_user)
    }

CONVERSION_SETTINGS = CONVERSION_SETTINGS_BUILDER()
assert {'chat', 'user'} <= CONVERSION_SETTINGS.keys(), f'Missing keys in {CONVERSION_SETTINGS}'
assert all({'from_db', 'to_db'} <= x.keys() for y in CONVERSION_SETTINGS.values() for x in y.values()), f"Missing 'from_db' or 'to_db' in {CONVERSION_SETTINGS=}"

class SettingsInfo:
    TABLES = {'chat': 'ChatSettings', 'user': 'UserSettings'}
    FIELDS = {'chat': 'chat_id', 'user': 'user_id'}

SettingType = Literal['chat', 'user']

def read_settings(key, *, id, settings_type: SettingType):
    conversion = CONVERSION_SETTINGS[settings_type][key]['from_db']
    raw = read_raw_settings(key, id=id, settings_type=settings_type)
    return conversion(raw) if raw is not None else None

def read_raw_settings(key, *, id, settings_type: SettingType):
    with sqlite3.connect('db.sqlite') as conn:
        cursor = conn.cursor()

        table_name = SettingsInfo.TABLES[settings_type]
        field_id = SettingsInfo.FIELDS[settings_type]
        query = (
            f"""SELECT value from {table_name}
                WHERE {field_id}=?
                AND key=?""",
            (id, key)
        )

        results = cursor.execute(*query).fetchall()
        return results[0][0] if results else None

async def listallsettings(update: Update, context: CallbackContext, scope:Literal[None, 'chat', 'user']=None):
    send = make_send(update, context)

    Args = InfiniteEmptyList(context.args)
    scope = scope or Args[0]
    if scope and '.' in scope:
        scope_base, scope_bits = scope.split('.', maxsplit=1)
    else:
        scope_base, scope_bits = scope, ''
    assert (not scope_base) or scope_base in ('chat', 'user')

    await send('\n'.join("- {} ({})".format(
            setting,
            '|'.join(['user'] * (setting in ACCEPTED_SETTINGS_USER) + ['chat'] * (setting in ACCEPTED_SETTINGS_CHAT)))
        for setting in sorted(ACCEPTED_SETTINGS_USER | ACCEPTED_SETTINGS_CHAT)
        if scope_base == 'user' and setting in ACCEPTED_SETTINGS_USER or scope_base == 'chat' and setting in ACCEPTED_SETTINGS_CHAT or not scope_base
        if setting.startswith(scope_bits)
    ))

async def settings_command(update: Update, context: CallbackContext, *, command_name: str, settings_type:Literal['chat'] | Literal['user'], accepted_settings:list[str]):
    send = make_send(update, context)
    send_html = partial(send, parse_mode='HTML')
    import html

    async def print_usage():
        await send_html(
            f"Usage:" + "\n"
            f"/{command_name} command.key" + "\n"
            f"/{command_name} command.key value" + "\n"
            f"/{command_name} delete command.key" + "\n\n"
            f"- Type /listallsettings for a list of all settings" '\n'
            f"- Type <code>/listallsettings {settings_type}</code> for a list of {settings_type} settings")

    if len(context.args) == 0:
        return await print_usage()

    Args = InfiniteEmptyList(context.args)

    action: Literal["get", "set", "getset", "delete"]
    
    match Args[0].lower():
        case "del" | "delete":
            args = Args[1:]
            action = "delete"
        case "get":
            args = Args[1:]
            action = "get"
        case "set":
            args = Args[1:]
            action = "set"
        case _:
            args = Args[:]
            action = "getset"

    try:
        key, *rest = args
    except ValueError:
        return await send_html(
            'You must provide some settings:\n\n'
            f'- Type <code>/listallsettings {settings_type}</code> for a list of all settings\n')

    if key not in accepted_settings:
        return await send_html(
            f'Unknown settings: "{html.escape(key)}"' "\n\n"
            f"- Type /listallsettings for a list of all settings" "\n"
            f"- Type <code>/listallsettings {settings_type}.{html.escape(key)}</code> to see all settings starting with your command" "\n")

    if rest and rest[0] == '=' and action in ("set", "getset"):
        rest = rest[1:]
        if not rest:
            raise UserError('Must specify a value when setting a value with "a.b = c"')

    if action == "delete":
        if len(rest) > 0:
            raise UserError(f'Usage: /{command_name} delete command.key')
        return await delsettings_command(update, context, key=key, command_name=command_name + ' ' + 'delete', settings_type=settings_type, accepted_settings=accepted_settings)

    if action == "set" and len(rest) == 0:
        raise UserError(f"Usage: /{command_name} set command.key value")

    list_type_and_extend = False
    if list_type := accepted_settings[key].type == 'list':
        if InfiniteEmptyList(rest)[0] == '+=':
            rest = rest[1:]
            list_type_and_extend = True
        else:
            list_type_and_extend = False

        value = ([] if list(rest) in [['()'], ['[]']] else
                 None if not rest else
                 list(rest))
    else:
        if len(rest) not in (0, 1):
            return await print_usage()
        # default, single value no space string
        value = rest[0] if rest else None

    if settings_type == 'user':
        id = update.effective_user.id
    elif settings_type == 'chat':
        id = update.effective_chat.id
    else:
        raise ValueError(f'Invalid settings_type: {settings_type}')
    
    if value is None:
        # read
        value = read_raw_settings(id=id, key=key, settings_type=settings_type)

        await send(f'Settings: {key} = {value}' if value is not None else
                    f'No settings for {key!r}')
    
    else:
        # write value
        set_settings(value_raw=value, id=id, key=key, settings_type=settings_type, list_type_and_extend=list_type_and_extend)
        op = "+=" if list_type_and_extend else "="
        await send(f"Settings: {key} {op} {value}")

async def delsettings_command(update:Update, context: CallbackContext, *, key: str, accepted_settings:list[str], settings_type:Literal['chat'] | Literal['id'], command_name:str):
    send = make_send(update, context)
    
    if settings_type == 'user':
        id = update.effective_user.id
    elif settings_type == 'chat':
        id = update.effective_chat.id
    else:
        raise ValueError(f'Invalid settings_type: {settings_type}')
     
    delete_settings(id=id, key=key, settings_type=settings_type)
    await send(f"Deleted settings for {key!r}")

mysettings = partial(settings_command, accepted_settings=ACCEPTED_SETTINGS_USER, settings_type='user', command_name='mysettings')
chatsettings = partial(settings_command, accepted_settings=ACCEPTED_SETTINGS_CHAT, settings_type='chat', command_name='chatsettings')

def migration0():
    with sqlite3.connect('db.sqlite') as conn:
        cursor = conn.cursor()
        cursor.execute("CREATE TABLE Events(date datetime, name text)")

def migration1():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        data = conn.execute("select date, name from Events").fetchall()
        conn.execute("drop table Events")
        conn.execute("create table Events(date datetime, name text, chat_id, source_user_id)")
        conn.executemany("insert into Events(date, name, chat_id, source_user_id) values(?,?,NULL,NULL)", data)
        conn.execute("end transaction")

def migration2():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        conn.execute("create table UserTimezone(user_id, timezone text)")
        conn.execute("end transaction")

def migration3():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        conn.execute("create table UserSettings(user_id, key text, value text)")
        conn.execute("end transaction")

def migration4():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        conn.execute("create table EuroRates(datetime, rates)")
        conn.execute("end transaction")

def migration5():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        conn.execute("create table ChatSettings(chat_id, key text, value text)")
        conn.execute("end transaction")

def migration6():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        conn.execute("create table Flashcard(user_id, sentence, translation)")
        conn.execute("end transaction")

def migration7():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute("begin transaction")
        conn.execute("create table FlashcardPage(user_id, name, current)")
        conn.execute("alter table Flashcard add page_name default '1'")
        conn.execute("end transaction")

def migration8():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute("create table NamedChatDebt(chat_id, debitor_id, creditor_id, amount, currency)") # debitor owes creditor
        conn.execute('end transaction')

def migration9():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute('create table EventLocation(key, value, chat_id)')
        conn.execute('end transaction')

def migration10():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        # EventFollow(a,b) exists <=> a Follows b (Event wise)
        conn.execute('create table EventFollowPending(a_chat_id NOT NULL, b_chat_id NOT NULL)')
        conn.execute('create table EventFollow(a_chat_id NOT NULL, b_chat_id NOT NULL)')
        conn.execute('end transaction')

def migration11():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute('alter table EventFollowPending add column a_name')
        conn.execute('alter table EventFollowPending add column b_name')
        conn.execute('alter table EventFollow add column a_name')
        conn.execute('alter table EventFollow add column b_name')
        conn.execute('update EventFollowPending set a_name = a_chat_id, b_name = b_chat_id')
        conn.execute('update EventFollow set a_name = a_chat_id, b_name = b_chat_id')
        conn.execute('end transaction')

def migration12():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute('alter table EventFollowPending add column a_thread_id DEFAULT \'\'')
        conn.execute('alter table EventFollow add column a_thread_id DEFAULT \'\'')
        conn.execute('end transaction')

def migration13():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute('create table List(name, chat_id, source_user_id)')
        conn.execute('create table ListElement(listid, value)')
        conn.execute('end transaction')

def migration14():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute('alter table List add column type DEFAULT "list"')
        conn.execute('end transaction')

def migration15():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute('create table FwdRelation(original_message_id, fwd_message_id, original_chat_id, original_chat_username)')
        conn.execute('end transaction')

def migration16():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute('create table EventLinkAttr(event_id REFERENCES Events(rowid), link)')
        conn.execute('end transaction')

def migration17():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute('alter table ListElement add column tree_parent REFERENCES ListElement(rowid) DEFAULT NULL')
        conn.execute('end transaction')

def migration18():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute("alter table NamedChatDebt add column reason DEFAULT NULL")
        conn.execute('end transaction')

def migration19():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute("create table TimezoneAlias(chat_id, alias, real)")
        conn.execute('end transaction')

def migration20():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute("create table EnglishPracticeIrregularVerbs(chat_id, user_id, json)")
        conn.execute('end transaction')

def migration21():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute("alter table FlashcardPage rename column user_id to chat_id")
        conn.execute("alter table Flashcard add column page_id")
        for rowid, in conn.execute('select rowid from flashcard'):
            conn.execute('''
                update flashcard set page_id=(
                    select rowid from flashcardpage where flashcardpage.chat_id=flashcard.user_id and flashcardpage.name=flashcard.page_name
                ) where rowid=?
                ''', (rowid, )
            )
        conn.execute('''alter table Flashcard drop column page_name''')
        conn.execute('''alter table Flashcard drop column user_id''')
        conn.execute('end transaction')

def migration22():
    with sqlite3.connect('db.sqlite') as conn:
        conn.execute('begin transaction')
        conn.execute('create table LocationDistanceEdge(chat_id, source, dest, distance)')
        conn.execute('end transaction')

def migration23():
    with sqlite3.connect('db.sqlite') as conn:
        c = conn.cursor()
        c.execute('begin transaction')
        c.execute('create table LocationDistanceGraph(name, visibility, chat_id)')
        c.execute('create table LocationDistanceEdgeGraphRelation(name, visibility, chat_id)')
        c.execute('create table LocationDistanceCurrentGraphChat(chat_id, graph_id)')

        c.execute('alter table LocationDistanceEdge add column graph_id')
        for chat_id, in c.execute('select distinct chat_id from LocationDistanceEdge').fetchall():
            c.execute('insert into LocationDistanceGraph(name, visibility, chat_id) VALUES (?,?,?)', ('chat', 'chat', chat_id))
            c.execute('insert into LocationDistanceCurrentGraphChat(chat_id, graph_id) VALUES(?,?)', (chat_id, graph_id := c.lastrowid))
            c.execute('update LocationDistanceEdge set graph_id=? where chat_id=?', (graph_id, chat_id))
        c.execute('end transaction')

def migration24():
    with sqlite3.connect('db.sqlite') as conn:
        c = conn.cursor()
        c.execute('begin transaction')
        c.execute('''create table LocationDistanceImportedGraph(chat_id, graph_id)''')
        c.execute('end transaction')

def migration25():
    with get_connection() as conn:
        c = conn.cursor()
        c.execute('begin transaction')
        c.execute('''create table LocationDistanceGraphNamespace(chat_id, namespace, UNIQUE(namespace))''')
        c.execute('end transaction')

def get_latest_euro_rates_from_api() -> json:
    import requests
    from telegram_settings_local import FIXER_TOKEN
    response = requests.get(f'http://data.fixer.io/api/latest?access_key={FIXER_TOKEN}&base=EUR').json()
    assert response['success']
    assert response['base'] == 'EUR'
    # response['date']
    return response['rates']

class DatetimeDbSerializer:
    @staticmethod
    def strptime(x:str):
        from datetime import datetime
        return datetime.strptime(x, "%Y-%m-%d %H:%M:%S")
    
    @staticmethod
    def strftime(x:datetime):
        return x.strftime("%Y-%m-%d %H:%M:%S")

    def to_db(self, x: datetime):
        return self.strftime(x)

    def from_db(self, x: any):
        return self.strptime(x)
    
    @staticmethod
    def strftime_minutes(x:datetime):
        return x.strftime("%Y-%m-%d %H:%M")

class JsonDbSerializer:
    def to_db(self, x: json):
        import json
        return json.dumps(x)

    def from_db(self, x: any):
        import json
        return json.loads(x)

Rates = dict
def get_database_euro_rates() -> Rates:
    query_get_last_date = ('''select MAX(datetime), rates from EuroRates''', ())

    from datetime import datetime as Datetime, timedelta as Timedelta

    with sqlite3.connect('db.sqlite') as conn:
        latest_date_string, rates_string = conn.execute(*query_get_last_date).fetchone() or (None, None)
        latest_date: Datetime = latest_date_string and DatetimeDbSerializer().from_db(latest_date_string)
        rates: json = rates_string and JsonDbSerializer().from_db(rates_string)
    
    now = Datetime.now(UTC).replace(tzinfo=None)
    if latest_date is None or now - latest_date > Timedelta(days=1):
        rates = get_latest_euro_rates_from_api()
        with sqlite3.connect('db.sqlite') as conn:
            conn.execute('''INSERT INTO EuroRates(datetime, rates) VALUES(?, ?)''', (DatetimeDbSerializer().to_db(now), JsonDbSerializer().to_db(rates)))
        return rates
    else:
        return rates


from decimal import Decimal

def format_currency(*, currency_list:list[str], amount_list:list[Decimal]):
    return '\n'.join(
        "{}: {:.2f}".format(currency.upper(), amount)
        for currency, amount in zip(currency_list, amount_list))

def convert_money(amount: Decimal, currency_base:str, currency_converted:str, rates:Rates):
    currency_base, currency_converted = currency_base.upper(), currency_converted.upper()
    if currency_base == 'EUR':
        return amount * Decimal(rates[currency_converted])
    if currency_converted == 'EUR':
        return amount / Decimal(rates[currency_base.upper()])
    return convert_money(convert_money(amount, currency_base, 'EUR', rates=rates), 'EUR', currency_converted, rates=rates)

def make_money_command(name:str, currency:str):
    async def money(update: Update, context: CallbackContext):
        send = make_send(update, context)
        read_chat_settings = make_read_chat_settings(update, context)

        chat_currencies = read_chat_settings('money.currencies') or DEFAULT_CURRENCIES

        from decimal import Decimal
        value, *_ = context.args or ['1']
        amount_base = Decimal(value)
        rates = get_database_euro_rates()
        currencies_to_convert = [x for x in chat_currencies if x != currency]
        amounts_converted = [convert_money(amount_base, currency_base=currency, currency_converted=currency_to_convert, rates=rates) for currency_to_convert in currencies_to_convert]
        return await send(format_currency(currency_list=[currency] + currencies_to_convert, amount_list=[amount_base] + amounts_converted))
    return money

eur = make_money_command("eur", "eur")
brl = make_money_command("brl", "brl")
rub = make_money_command("rub", "rub")

async def convertmoney(update, context):
    send = make_send(update, context)
    read_chat_settings = make_read_chat_settings(update, context)

    try:
        if len(context.args) == 2:
            value, currency = context.args
            currency = currency.upper()
            mode = 'to_chat_currencies'
            direction, currency_converted = None, None
        elif len(context.args) == 4:
            value, currency, direction, currency_converted = context.args
            currency = currency.upper()
            currency_converted = currency_converted.upper()
            mode = 'to_one_currency'
            assert direction == 'to'
        else:
            raise Exception
    except:
        return await send("Usage: /convertmoney value currency [to currency]")

    currency = MONEY_CURRENCIES_ALIAS.get(currency.lower(), currency)
    if currency_converted is not None:
        currency_converted = MONEY_CURRENCIES_ALIAS.get(currency_converted.lower(), currency_converted)

    amount_base = Decimal(value)
    rates = get_database_euro_rates()

    if mode == 'to_chat_currencies':
        chat_currencies = read_chat_settings('money.currencies') or DEFAULT_CURRENCIES
        currencies_to_convert = [x.upper() for x in chat_currencies if x.upper() != currency.upper()]
    elif mode == 'to_one_currency':
        currencies_to_convert = [currency_converted.upper()]

    try:
        amounts_converted = [convert_money(amount_base, currency_base=currency, currency_converted=currency_to_convert, rates=rates) for currency_to_convert in currencies_to_convert]
    except KeyError as e:
        raise UserError(f"Unknown currency: {e}")
    await send(format_currency(currency_list=[currency] + currencies_to_convert, amount_list=[amount_base] + amounts_converted))

async def enable_disable_command(update, context, direction: Literal['enable', 'disable']):
    async def send_command(msg):
        import html
        return await make_send(update, context)("<code>" + html.escape(msg) + "</code>", parse_mode="HTML")

    if not context.args:
        return await make_send(update, context)(
            "Usage: /{'enable' if direction == 'enable' else 'disable'} responder\n\n"
            "Available commands: {}".format(', '.join(_[1] for _ in RESPONDERS)))

    responder, = context.args
    responder = responder.lower()
    if responder not in (_[1] for _ in RESPONDERS):
        raise UserError(f"Unknown responder: {responder}\n\nAvailable responders: {', '.join(_[1] for _ in RESPONDERS)}")
    
    return await send_command(f"/chatsettings {responder}.active {'on' if direction == 'enable' else 'off'}")

async def implicit_setting_command(update, context, type: Literal['disable', 'only', 'known']):
    reply = update_get_reply(update)
    if type == 'disable' and not reply:
        return await enable_disable_command(update, context, direction='disable')
    if not reply:
        raise UserError("Use this on a message")
    if not reply.from_user.id == context.bot.id:
        raise UserError("Use this on a bot message")

    send = make_send(update, context)
    async def send_command(msg):
        import html
        return await send("<code>" + html.escape(msg) + "</code>", parse_mode="HTML")
    
    def is_currency_list():
        return all(re.compile('^[A-Z]{3}[:].*$').match(line) for line in reply.text.splitlines())

    def is_locationdistance_list():
        return all(x /fullmatchesI/ '• (\d+) [|] (.*)' for x in reply.text.splitlines())

    if type == 'disable':
        if reply.text:
            if is_currency_list():
                return await send_command('/chatsettings money.active off')
            if reply.text == 'Click the file below to add the event to your calendar:':
                return await send_command('/chatsettings event.addevent.help_file off')
            if reply.text.startswith('Error: Time must be specified (policy of the group)'):
                return await send_command('/chatsettings event.addevent.required_time off')
            if reply.text /fullmatches/ 'Forwarded to \d+ chats':
                return await send_command('/chatsettings event.addevent.display_forwarded_infos off')
            if reply.text.startswith('Error: You must specify a reason'):
                return await send_command('/chatsettings sharemoney.required_for off')
            if is_locationdistance_list():
                return await send_command('/chatsettings locationdistance.active off')
        elif reply.document and reply.document.file_name == 'event.ics':
            return await send_command('/chatsettings event.addevent.display_file off')
            
    elif type in ('only', 'known') and reply.text and is_currency_list():
        if type == 'only':
            read_chat_settings = make_read_chat_settings(update, context)
            currencies = read_chat_settings('money.currencies')
            new_currencies = ordered_set_remove(currencies, list(map(str.upper, context.args)))
        elif type == 'known':
            new_currencies = list(map(str.upper, context.args))
        return await send_command('/chatsettings money.known_currencies {}'.format(' '.join(new_currencies)))
    
    return await send("Unknown type of message")

async def sharemoney(update, context):
    send = make_send(update, context)
    Args = GetOrEmpty(context.args)
    on_off = Args[0]
    try:
        if on_off == 'on':
            return await send("Run: /chatsettings sharemoney.active on")
        elif on_off == 'off':
            return await send("Run: /delchatsettings sharemoney.active")
        else:
            raise UsageError
    except UsageError:
        return await send('Usage: /sharemoney on|off')

async def listdebts(update, context):
    send = make_send(update, context)
    chat_id = update.effective_chat.id
    lines = simple_sql_dict(('select chat_id, debitor_id, creditor_id, amount, currency, reason from NamedChatDebt where chat_id=?', (chat_id,)))
    debts_sum = {}
    
    for debt in (NamedChatDebt(**x) for x in lines):
        if (debt.debitor_id, debt.creditor_id) in debts_sum or (debt.creditor_id, debt.debitor_id) in debts_sum:
            key = ((debt.debitor_id, debt.creditor_id) if (debt.debitor_id, debt.creditor_id) in debts_sum else
                   (debt.creditor_id, debt.debitor_id))
            
            sign = (+1 if (debt.debitor_id, debt.creditor_id) in debts_sum else
                    -1)
            
            debts_sum[key] += sign * Decimal(debt.amount)
        else:
            debts_sum[debt.debitor_id, debt.creditor_id] = Decimal(debt.amount)
    
    name_re = regex.compile(r"(\p{L}\w*)([.]([A-Za-z]+))?")

    def my_sort(keys):
        def key(x):
            full_a, full_b = x
            a_name, b_name = name_re.fullmatch(full_a).group(1), name_re.fullmatch(full_b).group(1)
            a_cur,  b_cur  = name_re.fullmatch(full_a).group(3), name_re.fullmatch(full_b).group(3)
            return (a_name, b_name, a_cur or '', b_cur or '')
        return sorted(keys, key=key)
    
    debts_sum_sorted = [(k, debts_sum[k]) for k in my_sort(debts_sum)]

    return await send('\n'.join(
        "{} owes {} {}".format(debitor, creditor, amount) if amount > 0 else
        "{} owes {} {}".format(creditor, debitor, -amount) if amount < 0 else
        "{} and {} are even".format(debitor, creditor)
        for (debitor, creditor), amount in debts_sum_sorted) or 'No debts in this chat !')

async def detaildebts(update, context):
    Args = GetOrEmpty(context.args)

    if Args[0].isdecimal():
        last_n: int = int(Args[0])
        Args = GetOrEmpty(Args[1:])
    else:
        last_n: int = 30
        Args = GetOrEmpty(Args[0:])
    
    account_filter: None | tuple['simple', str] | tuple['multi', list[str]]

    if not Args[0]:
        account_filter = None

    elif not Args[1]:
        # one person filter: filter all where that person happen
        account_filter = ('simple', Args[0])

        Args = GetOrEmpty(Args[1:])

    elif Args[0] and Args[1]:
        # multiway filter
        if Args[2]:
            raise UserError("Currently not avaialbe to filter with more than 2 persons")
        account_filter = ('multi', [Args[0], Args[1]])

        Args = GetOrEmpty(Args[2:])

    send = make_send(update, context)
    chat_id = update.effective_chat.id

    sql_filter = ('1=1' if account_filter is None else
                  'debitor_id=? OR creditor_id=?' if account_filter[0] == 'simple' else
                  'debitor_id=? AND creditor_id=? OR debitor_id=? AND creditor_id=?')
    
    filter_params = (() if account_filter is None else
                     (account_filter[1], account_filter[1]) if account_filter[0] == 'simple' else 
                     (account_filter[1][0], account_filter[1][1], account_filter[1][1], account_filter[1][0]))

    sql = 'select chat_id, debitor_id, creditor_id, amount, currency, reason from NamedChatDebt where chat_id=? AND (%s) ORDER BY rowid DESC LIMIT ?' % sql_filter

    count = simple_sql_dict(('select count(*) as c from NamedChatDebt where chat_id=? AND (%s)' % sql_filter, (chat_id, ) + filter_params))[0]['c']

    lines = simple_sql_dict((sql, (chat_id, ) + filter_params + (last_n, )))

    to_print = []
    if count > len(lines):
        to_print.append('...')
    
    if account_filter and account_filter[0] == 'multi':
        to_print.append('// "{}" owes "{}"'.format(*account_filter[1]))
        total_displayed = 0

    debt: NamedChatDebt
    for debt in reversed([NamedChatDebt(**x) for x in lines]):
        name_re = regex.compile(r"(\p{L}\w*)([.]([A-Za-z]+))?")
        debitor_name = name_re.fullmatch(debt.debitor_id).group(1)
        creditor_name = name_re.fullmatch(debt.creditor_id).group(1)
        currency_1 = name_re.fullmatch(debt.creditor_id).group(3)
        currency_2 = name_re.fullmatch(debt.creditor_id).group(3)
        if currency_1 or currency_2:
            assert currency_1.upper() == currency_2.upper()
        currency = currency_1 or currency_2
        currency = currency and currency.upper()
    
        if account_filter and account_filter[0] == 'multi':
            
            if (debt.debitor_id, debt.creditor_id, ) == tuple(account_filter[1]):
                directed_amount = debt.amount
            elif (debt.creditor_id, debt.debitor_id, ) == tuple(account_filter[1]):
                directed_amount = - debt.amount
            else:
                raise ValueError('Logic problem in identify directed_amount')
            
            to_print.append(' '.join(filter(None, (
                '+' if directed_amount >= 0 else '-',
                str(abs(directed_amount)),
                str(currency) if currency else '',
                f'# {debt.reason}' if debt.reason else '',
            ))))

            total_displayed += directed_amount

        else:
            to_print.append(' '.join(filter(None, (
                'Debt',
                f'"{debitor_name}"',
                'owes',
                f'"{creditor_name}"',
                f'{debt.amount}',
                f'{currency}' if currency else '',
                f'# {debt.reason}' if debt.reason else ''
            ))))
    
    if account_filter and account_filter[0] == 'multi':
        to_print.append("// Total: {}".format(total_displayed))

    return await send('\n'.join(to_print) or 'No debts in that chat !')

class EnglishPracticeData:
    class IrregularVerbs:
        TRIPLETS = ["put", "cut", "let", "shut", "split", "spread", "hurt", "cost", "burst","fit", "bet", "hit", "set", "broadcast", "read"]
        TWINS = ["bend", "teach", "think", "catch", "bring", "build", "buy", "creep", "deal", "dig", "feed", "feel", "fight", "find", "flee","get","hang","have", "hear", "hold", "keep", "kneel", "lay", "lead", "leave", "lend", "light", "lose", "make", "mean", "meet", "pay", "say", "seek", "sell", "send", "sew", "shine", "shoot", "sit", "sleep", "slide","spend", "spit", "stand", "stick", "sting", "strike", "sweep", "swing", "tell", "think", "understand", "weep", "win"]

        @staticmethod
        def phrase_for_verb(word):
            if word == "put":
                return("Обычно, Петя [кладет] ручку на стол. Вчера он [положил] ручку на пол, но чаще всего ручка в его доме [кладется] на стол")
            elif word == "cut":
                return("Она никогда не [режет] хлеб. Неделю назад она [порезала] свой палец. Вообще, колбасу [режут] другим ножом .")
            elif word == "let":
                return("Она редко [позволяет] другим помочь. 5 минут назад она [позволила] мне помочь ей открыть дверь. Обычно, нагетивные эмоции [выпускаются] все вместе.")
            elif word == "shut":
                return("Она всегда [захлопывает] двень и книгу очень громко. На прошлой неделе я сказал ей замолчать и она [замолчала] (....up). Эта дверь [не закрывается] силой.")
            elif word == "split":
                return("Наш учитель никогда [не делит] класс на группы. Вчера она [разделила] маффин и дала мне половину. Обычно, на свидании в Европе счет [делится пополвм].")
            elif word == "spread":
                return("Она часто [распространяет] слухи про других. Час назад огонь [распространился] на второй этаж больницы. Холодное масло[размазано] крупными кусками на хлебе.")
            elif word == "hurt":
                return("Она часто [делает больно] другим. 5 минут назад она [повредила] свою коленку. В автомобильной аварии никто [не пострадал].")
            elif word == "cost":
                return("Это [стоит] слишком дорого. Неделю назад эта сумка [стоила] меньше. Квартиры в Москве никогда [не стоили] так дешево, как сейчас.\n\nПОДСКАЗКА: третье предложение это Present Perfect Simple с третьей формой глагола")
            elif word == "burst":
                return("Каждый понедельник после уроков она [раздается смехом] ***ПОДСКАЗКА*** :инфинитив тут TO BURST INTO LAUGHTER. 5 минут назад она [разревелась] ***ПОДСКАЗКА*** :инфинитив тут TO BURST INTO TEARS. Трубы [были прорваны] всю неделю") 
            elif word == "fit":
                return("Это платье мне[не подходит по размеру]. В детстве она всегда [вливалась] в любой коллектив ***ПОДСКАЗКА*** :инфинитив тут TO FIT IN - ВЛИВАТЬСЯ. Одежда была [подогнана] под конкретные мерки клиента.")
            elif word == "bet":
                return("Она никогда [не делает ставки] на свою лошадь. Неделю назад она [поставила ставки] на свою любимую команду. Вообще, на детей [ставки не ставят]  .")
            elif word == "hit":
                return("Она никогда не [ударяет] своих детей . Неделю назад она [ударила] свой палец. Вчера 3 дрона [было сбито] СВО .")
            elif word == "set":
                return("Она никогда не [накрывает] на стол. Неделю назад она [отложила] немного денег ***ПОДСКАЗКА*** :инфинитив тут TO SET ASIDE MONEY/TIME -ОТЛОЖИТЬ . Время для урока [определено]. ***ПОДСКАЗКА*** :инфинитив тут TO SET A TIME/DATE.")
            elif word == "broadcast":
                return("Она никогда не [транслирует] свои страхи. Неделю назад она [транслировала] свое интервью в грам. Матч будет [транслирован] со стадиона.")
            elif word == "read":
                return("Она никогда не [читает] книги. Неделю назад она [прочитала] 3 поста в соц сетях. Поэма [будет прочтена] со стадиона.")
            raise ValueError

def grouped_dict(it):
    from collections import defaultdict
    d = defaultdict(list)
    for k, v in it:
        d[k].append(v)
    return dict(d)

class LanguagePractice:
  async def practice_command(update, context):
    # only accessible when the practiceenglish.active is
    send = make_send(update, context)
    await send("Translate the verbs in [ ] to English.\n\nSeparate them with spaces.\nFor example: put put put.\n\nKEEP IN MIND that:\n1)every 1st verb is in Present Simple\n2)Every 2nd verb is in Past Simple\n3)Every third verb is in PASSIVE VOICE.")
    # table EnglishPracticeIrregularVerbs
    VERBS = EnglishPracticeData.IrregularVerbs.TRIPLETS
    
    import random
    session = {'verbs': (session_verbs := random.sample(VERBS, 5)), 'i': 1}

    user_id = update.effective_user.id
    chat_id = update.effective_chat.id

    import json
    simple_sql(('insert into EnglishPracticeIrregularVerbs(user_id, chat_id, json) VALUES (?,?,?)', (user_id, chat_id, json.dumps(session))))

    await send(EnglishPracticeData.IrregularVerbs.phrase_for_verb(session_verbs[0]))

    return 'next-verb'

  async def practice_command_next_verb_state(update, context):
    send = make_send(update, context)

    user_id = update.effective_user.id
    chat_id = update.effective_chat.id

    with sqlite3.connect("db.sqlite") as conn:
        conn.execute('begin transaction')
        my_simple_sql = partial(simple_sql, connection=conn)
        session_rowid, session = only_one(my_simple_sql(('select rowid, json from EnglishPracticeIrregularVerbs where user_id=? and chat_id=? ORDER BY rowid DESC LIMIT 1', (user_id, chat_id))))
        
        import json
        session = json.loads(session)
        
        finished = (session['i'] == len(session['verbs']))

        if finished:
            await send('Congratulations!')

        else:
            await send(EnglishPracticeData.IrregularVerbs.phrase_for_verb(session['verbs'][session['i']]))
            session['i'] += 1

            my_simple_sql(('update EnglishPracticeIrregularVerbs set json = ? where rowid=?', (json.dumps(session), session_rowid)))

        conn.execute('end transaction')
    
    return 'next-verb' if not finished else ConversationHandler.END

async def help(update, context):
    send = make_send(update, context)

    Args = InfiniteEmptyList(context.args)
    
    module_filter = Args[0].lower() if Args[0].lower() in COMMAND_LIST_ALL_MODULES else ''
    display_modules = Args[0].lower() in ('module', 'modules')

    one_command_display = Args[0].lower()
    if one_command_display.startswith('/'):
        one_command_display = one_command_display[1:]

    bot_father = '--botfather' in context.args

    if display_modules:
        return await send('\n'.join(map("- {}".format, COMMAND_LIST_ALL_MODULES)))

    fmt = ('{} - {}' if bot_father else
           '/{} {}')
    
    li = ([x.name for x in COMMAND_LIST_HELP if x.botfather] if bot_father else
          [x.name for x in COMMAND_LIST_HELP])

    if bot_father:
        return await send('\n'.join(fmt.format(command, COMMAND_DESC.get(command, command)) for command in li))

    def top_most_module(module):
        if all_modules_parent.get(module):
            return top_most_module(all_modules_parent[module])
        else:
            return module
        
    by_modules = grouped_dict((top_most_module(COMMAND_LIST_HELP_DICT[c].module), c) for c in li)

    if one_command_display in commandspecs.MAN:
        import textwrap
        return await send(textwrap.dedent(commandspecs.MAN[Args[0]]).strip())
    else:
        lines = []
        first = True
        for mod, L in by_modules.items():
            if module_filter and module_filter != mod:
                continue
            if first:
                first = False
            else:
                lines.append('')

            lines.append(f'[Module "{mod}"]')
            for command in L:
                lines.append('  ' + fmt.format(command, COMMAND_DESC.get(command, command)))

        return await send('\n'.join(lines) or '?')
    
class UserError(ValueError):
    pass

class EventAdminError(UserError):
    def __init__(self, msg="You are not allowed to do this"):
        super().__init__(msg)

class UnknownDateError(UserError):
    pass

class DictJsLike(dict):
    def __getattribute__(self, x):
        if x in self:
            return self[x]
        return super().__getattribute__(x)

class EventFormatting:
    emojis = DictJsLike(
         Name="📃",
         Time="⌚",
         Date="🗓️",
         Location="📍",
         Link="🔗",
    )

class Unicode:
    BULLET = "•"

class EventInfosAnalyse:
    possibles = {
        'what': 'what', 'when': 'when', 'where': 'where',
        'quand': 'when', 'quoi': 'what', 'où': 'where',
        'name': 'what', 'location': 'where', 
    }

    emojis_meaning = {y:x for x,y in EventFormatting.emojis.items()}

class EventAnalyseError(UserError):
    pass

class EventAnalyseMultipleError(EventAnalyseError):
    def __init__(self, exceptions):
        self.exceptions = exceptions
        
    def __str__(self):
        return '\n---\n'.join(map(str, self.exceptions))

async def log_error(error, send):
    if isinstance(error, UserError):
        return await send("Error: {}".format(error))
    elif isinstance(error, CrazyJamFwdError):
        logging.error("Error", exc_info=error)
        return
    else:
        logging.error("Error", exc_info=error)
        return await send("An unknown error occured in your command, ask @robertvend to fix it !")

async def general_error_callback(update:Update, context:CallbackContext):
    async def send_on_error(m):
        if update and update.effective_chat:
            send = make_send(update, context)
            await send(m)
    
    return await log_error(context.error, send_on_error)

import unittest
import unittest.mock
from unittest import IsolatedAsyncioTestCase, TestCase

class SyncTests(TestCase):
    def test_detect_currencies(self):
        self.assertIn(('5', 'eur'), detect_currencies("This is 5€"))
        self.assertIn(('5', 'eur'), detect_currencies("This is 5 eur"))
        self.assertIn(('5', 'eur'), detect_currencies("This is 5 euros"))
        self.assertIn(('5', 'eur'), detect_currencies("This is 5 EUR"))

async def test_simple_output(function, input:list[str]):
    # setup
    context = unittest.mock.AsyncMock()
    context.args = input
    update = unittest.mock.Mock()
    update.effective_chat = Chat(type=Chat.PRIVATE, id='123')
    
    # call function
    await function(update, context)
    
    # asserts
    context.bot.send_message.assert_called_once()
    return context.bot.send_message.mock_calls[0].kwargs['text']

async def test_simple_responder(function, msg:str):
    # setup 
    send = unittest.mock.AsyncMock()
    
    # call
    await function(msg, send)
    
    # assert
    send.assert_called_once()
    return send.mock_calls[0].args[0]

async def test_multiple_responder(function, msg:str):
    # setup 
    send = unittest.mock.AsyncMock()
    
    # call
    await function(msg, send)
    
    # assert
    return [send.mock_calls[i].args[0] for i in range(len(send.mock_calls))]

class AsyncTests(IsolatedAsyncioTestCase):
    async def test_ru(self):
        self.assertEqual(await test_simple_output(ru, ['azerty']), 'азерты', "One letter mapping")
        self.assertNotEqual(await test_simple_output(ru, ['azerty']), 'лалала', "Wrong output")
        self.assertEqual(await test_simple_output(ru, ['zhina']), 'жина', "Two letters mapping")
        self.assertEqual(await test_simple_output(ru, ["hello'"]), 'хеллоь', "Soft sign")
        self.assertEqual(await test_simple_output(ru, ["hello''"]), 'хеллоъ', "Hard sign")
        self.assertEqual(await test_simple_output(ru, ["xw"]), 'хв', "x and w")
        self.assertEqual(await test_simple_output(ru, ['hello', 'shchashasha']), 'хелло щашаша', "Multiple words")
        self.assertEqual(await test_simple_output(ru, ['Chto']), 'Что', 'Mix of capital and small letters')
    
    @unittest.skip
    async def test_hello_responder(self):
        self.assertIn("hello", (await test_simple_responder(hello_responder, "Hello")).lower())
        self.assertIn("hello", (await test_simple_responder(hello_responder, "Hello World !")).lower())
        self.assertEqual(0, len(await test_multiple_responder(hello_responder, "Tada")))
    
    @unittest.skip
    async def test_money_responder(self):
        results = await test_multiple_responder(money_responder, "This is 5€")
        self.assertEqual(1, len(results))
        self.assertIn("EUR: 5", results[0])
        self.assertIn("BRL: 26", results[0])


COMMAND_DESC = {
    "help": "Help !",
    "caps": "Returns the list of parameters in capital letters",
    "addevent": "Add event",
    "addschedule": "Add multiple events",
    'eventfollow': "Follow another chat to receive their events",
    'eventacceptfollow': "Accept an event follow request",
    'deleventfollow': "Stop to event follow some chat",
    'renameeventfollow': "Rename a follow relation",
    'renameeventacceptfollow': "Rename an accepted follow relation",
    'deleventacceptfollow': "Stop some chat from event following you",
    "nextevent": "Display the next event in emoji row format",
    "lastevent": "Display the last event in emoji row format",
    "listevents": "List events",
    "listdays": "List events grouped by days",
    "listtoday": "Shortcut for /listdays today, can add time marker",
    "today": "Shortcut for /listtoday",
    "whereis": "Remember a place/directions for events",
    "thereis": "Set a place a place/directions for events",
    "whereto": "Remember a place/directions for events in cascade mode",
    "delevent": "Delete event",
    "ru": "Latin alphabet to Cyrillic using Russian convention",
    "dict": "Shows definition of each word using dictionary and settings engine",
    "wikt": "Shows definition of each word using wiktionary",
    "larousse": "Show definition of each word using french dictionary Larousse.fr",
    'eur': "Convert euros to other currencies",
    'brl': "Convert brazilian reals to other currencies",
    'rub': "Convert russian rubles to other currencies",
    'convertmoney': 'Convert money to chat currencies or to specific currency',
    "mytimezone": "Set your timezone to use events commands",
    "mysettings": "Change user settings that are usable for commands",
    "delsettings": "Delete user settings that are usable for commands",
    "chatsettings": "Change chat settings that are usable for commands",
    "delchatsettings": "Delete chat settings that are usable for commands",
    "flashcard": "Add a new flashcard to help memorize words more easily",
    "exportflashcards": "Export your flashcards in excel format",
    "praticeflashcards": "Practice your flashcards to train your memory",
    "switchpageflashcard": "Switch to a page to group your flashcards",
    "uniline": "Describe in Unicode each character or symbol or emoji",
    "nuniline": "Describe in Unicode each non ascii character or symbol or emoji",
    "timeuntil": "Tell the time until an event",
    "timesince": "Tell the elapsed time since an event",
    "sleep": "Record personal sleep cycle and make graphs", 
    "sharemoney": "Manage money between users (shared bank account, add a debt)",
    "listdebts": "List debts between users (sharemoney)",
    "detaildebts": "List debts between users in details (sharemoney)",
    "tzalias": "Set short codes for timezones",

    'createlist': 'Create a list (of strings, by default)',
    'addtolist': 'Add to the end of a list',
    'appendtolist': 'Alias for addtolist',
    'removefromlist': 'Remove the first element from a list or print error',
    'delfromlist': 'Alias for delfromlist',
    'deletefromlist': 'Alias for removefromlist',
    'printlist': 'Print a list using dashes',
    'dirlist': 'List all lists',
    'dellist': 'Delete a list from all lists',
    'menu': 'Menu with buttons'
}

import itertools

def ordered_set_remove(A, B):
    return (''.join if isinstance(A, str) else type(A))(x for x in A if x not in B)

def ordered_set_union(A, B):
    return (''.join if isinstance(A, str) else type(A))(x for x in itertools.chain(A, B))


CommandModules = Literal['event', 'eventlocation', 'money', 'sharemoney', 'lang', 'dict', 'flashcard', 'list']
@dataclass
class CommandInfoSpecs:
    name: str
    module: CommandModules
    botfather: bool = True

class commandspecs:
    MAN = {
        'flashcard': '''
            /flashcard adds a flashcard to a page.
            A flashcard is a card witch two faces, the face-up and face-down

            There are multiple ways to do that:
                /flashcard a b
                Just two words, face-up will be a, face-down will be b

                /flashcard a = b
                Here a or b can be multiple words, we use the "=" symbol

                /flashcard a / b
                Same thing but with the "/" separator

                Replying to a message
                The replied message will be face-up, and the arguments will be face-down

            EXAMPLES

                /flashcard Bonjour Hello
                It adds the card face-up = Bonjour, and face-down = Hello to the current page

                /flashcard Comment ça va = How are you ?
                We use the "=" separator between face-up and face-down.

                Message 1: J'ai mangé une pomme
                Reply to Message 1: /flashcard I ate an apple
                This is useful when the card must have a sentence said by someone else,
                here we only write the face-down and reuse the text written by someone else
                It Adds the flashcard "J'ai mangé une pomme" / "I ate an apple"
        ''',
        'practiceflashcards': '''
            /practiceflashcards allows to practice flashcards by shuffling a flashcard page them
            The command shuffle all cards and display the face-up.
            Then waits for a message (the user should answer the cards).
            And finally the cards are revealed and the conversation is over

                n: the number of random flashcards drawn
                reverse: ask for the face down and gives the face up as an answer
                page: use another page
                user: use deck from user (beware, you might reveal private info in the group)

            EXAMPLES
            
                /practiceflashcards
                Practice the current flashcard page

                /practiceflashcards n:10 page:french
                Practice the "french" page with 10 random flashcards

            SEE ALSO

            - help `switchpageflashcard` to change current flashcard 
            - help `flashcard` to add a flashcard to current page
            - help `list.dynamic.flashcard.current` to control the current page
        '''
    }

all_modules = ('event', 'eventlocation', 'money', 'sharemoney', 'lang', 'dict', 'flashcard', 'list')
all_modules_parent = {
    'event': None,
    'eventlocation': 'event',
    'money': None,
    'sharemoney': 'money',
    'lang': None,
    'dict': 'lang',
    'flashcard': 'lang',
    'list': None,
}

COMMAND_LIST_HELP = (
    CommandInfoSpecs('addevent', 'event'),
    CommandInfoSpecs('addschedule', 'event'),
    CommandInfoSpecs('delevent', 'event'),
    CommandInfoSpecs('iaddevent', 'event'),
    CommandInfoSpecs('nextevent', 'event'),
    CommandInfoSpecs('lastevent', 'event'),
    CommandInfoSpecs('listevents', 'event'),
    CommandInfoSpecs('listdays', 'event'),
    CommandInfoSpecs('listtoday', 'event'),
    CommandInfoSpecs('today', 'event'),
    CommandInfoSpecs('tomorrow', 'event'),
    CommandInfoSpecs('mytimezone', 'event'),
    
    CommandInfoSpecs('eventfollow', 'event'),
    CommandInfoSpecs('eventacceptfollow', 'event'),
    CommandInfoSpecs('deleventfollow', 'event'),
    CommandInfoSpecs('deleventacceptfollow', 'event'),

    CommandInfoSpecs('whereis', 'eventlocation'),
    CommandInfoSpecs('whereto', 'eventlocation'),
    CommandInfoSpecs('delwhereis', 'eventlocation'),
    CommandInfoSpecs('delthereis', 'eventlocation'),

    CommandInfoSpecs('convertmoney', 'money'),
    CommandInfoSpecs('sharemoney', 'sharemoney'),

    CommandInfoSpecs('ru', 'lang'),

    CommandInfoSpecs('dict', 'dict'),
    CommandInfoSpecs('wikt', 'dict'),
    CommandInfoSpecs('larousse', 'dict'),
    
    CommandInfoSpecs('flashcard', 'flashcard'),
    CommandInfoSpecs('myflashcard', 'flashcard'),
    CommandInfoSpecs('exportflashcards', 'flashcard'),
    CommandInfoSpecs('practiceflashcards', 'flashcard'),
    CommandInfoSpecs('switchpageflashcard', 'flashcard'),
    CommandInfoSpecs('listflashcards', 'flashcard'),
    CommandInfoSpecs('listpageflashcard', 'flashcard'),

    CommandInfoSpecs('mysettings', 'settings'),
    CommandInfoSpecs('delmysettings', 'settings'),
    CommandInfoSpecs('chatsettings', 'settings'),
    CommandInfoSpecs('delchatsettings', 'settings'),
    CommandInfoSpecs('listallsettings', 'settings'),

    CommandInfoSpecs('createlist', 'list'),
    CommandInfoSpecs('addtolist', 'list'),
    CommandInfoSpecs('removefromlist', 'list'),
    CommandInfoSpecs('delfromlist', 'list'),
    CommandInfoSpecs('deletefromlist', 'list'),
    CommandInfoSpecs('printlist', 'list'),
    CommandInfoSpecs('dirlist', 'list'),
    CommandInfoSpecs('dellist', 'list'),
    CommandInfoSpecs('menu', 'menu'),
)

COMMAND_LIST_HELP_DICT = {x.name: x for x in COMMAND_LIST_HELP}

COMMAND_LIST_ALL_MODULES = list(remove_dup_keep_order(x.module for x in COMMAND_LIST_HELP))


if __name__ == '__main__':
    application = ApplicationBuilder().token(TOKEN).build()
    
    start_handler = CommandHandler('start', start)
    application.add_handler(start_handler)

    application.add_handler(MessageHandler(CrazyJamFilter(), on_crazy_jam_message))

    application.add_handler(ConversationHandler(
        entry_points=[CommandHandler('practice', LanguagePractice.practice_command, filters=EnglishPracticeFilter())],
        states={
            'next-verb': [
                MessageHandler(filters.TEXT & ~filters.COMMAND, LanguagePractice.practice_command_next_verb_state)
            ],
        },
        fallbacks=[]
    ))

    message_handler = MessageHandler(filters.TEXT & (~filters.COMMAND), on_message)
    application.add_handler(message_handler)
    
    application.add_handler(ConversationHandler(
        entry_points=[CommandHandler('befluent', befluent)],
        states = {
            
        },
        fallbacks=[

        ]
    ))
    application.add_handler(CommandHandler('caps', caps))
    application.add_handler(CommandHandler('ids', ids))
    application.add_handler(CommandHandler('addevent', add_event))
    application.add_handler(CommandHandler('iameventadmin', iameventadmin))
    application.add_handler(ConversationHandler(
        entry_points=[CommandHandler('iaddevent', InteractiveAddEvent.ask_when)],
        states={
            'ask-what-or-time': [
                MessageHandler(filters.TEXT & ~filters.COMMAND, InteractiveAddEvent.ask_what_or_time),
            ],
            'ask-time': [
                MessageHandler(filters.TEXT & ~filters.COMMAND, InteractiveAddEvent.ask_time),
            ],
            'ask-what': [
                MessageHandler(filters.TEXT & ~filters.COMMAND, InteractiveAddEvent.ask_what),
                CommandHandler('empty', InteractiveAddEvent.ask_what_empty),
                CommandHandler('midnight', InteractiveAddEvent.ask_what_empty),
            ],
            'ask-where': [
                MessageHandler(filters.TEXT & ~filters.COMMAND, InteractiveAddEvent.ask_where),
                CommandHandler('empty', InteractiveAddEvent.ask_where_empty),
                CommandHandler('skip', InteractiveAddEvent.ask_where_empty),
            ],
            'ask-confirm': [
                MessageHandler(filters.TEXT & ~filters.COMMAND, InteractiveAddEvent.ask_confirm),
                CommandHandler('empty', InteractiveAddEvent.ask_confirm_empty),
                CommandHandler('skip', InteractiveAddEvent.ask_confirm_empty),
            ],
            'do-add-event': [MessageHandler(filters.TEXT & ~filters.COMMAND, InteractiveAddEvent.do_add_event)]
        },
        fallbacks=[]
    ), group=3)
    application.add_handler(CommandHandler('events', events()))
    application.add_handler(CommandHandler('addschedule', addschedule))
    application.add_handler(CommandHandler('eventfollow', macro_event_follow))
    application.add_handler(CommandHandler('nextevent', next_event))
    application.add_handler(CommandHandler('rnextevent', partial(next_event, relative=True)))
    application.add_handler(CommandHandler('listevents', list_events))
    application.add_handler(CommandHandler('rlistevents', partial(list_events, relative=True)))
    application.add_handler(CommandHandler('listdays', list_days))
    application.add_handler(CommandHandler('crazyjamdays',  partial(list_days,formatting='crazyjamdays')))
    application.add_handler(CommandHandler('listdayslink',  partial(list_days,formatting='linkdays')))
    application.add_handler(CommandHandler('listdayshtml',  partial(list_days,formatting='linkdayshtml')))
    application.add_handler(CommandHandler('listdaysshort',  partial(list_days,formatting='short')))
    application.add_handler(CommandHandler('listdaysshorthtml',  partial(list_days,formatting='shorthtml')))
    application.add_handler(CommandHandler('rlistdays', partial(list_days, relative=True)))
    application.add_handler(CommandHandler('listoday', list_today)) # hidden command, for typo
    application.add_handler(CommandHandler('rlistoday', partial(list_today, relative=True))) # hidden command, for typo
    application.add_handler(CommandHandler('listtoday', list_today))
    application.add_handler(CommandHandler('rlisttoday', partial(list_today, relative=True)))
    application.add_handler(CommandHandler('today', list_today))
    application.add_handler(CommandHandler('tomorrow', partial(list_days_or_today, mode='tomorrow', relative=False)))
    application.add_handler(CommandHandler('rtoday', partial(list_today, relative=True)))
    application.add_handler(CommandHandler('lastevent', last_event))
    application.add_handler(CommandHandler('rlastevent', partial(last_event, relative=True)))
    application.add_handler(CommandHandler('weekiso', weekiso))
    application.add_handler(CommandHandler('weekisoroman', weekisoroman))
    application.add_handler(CommandHandler('whereis', whereis))
    application.add_handler(CommandHandler('whereto', whereto))
    application.add_handler(CommandHandler('thereis', thereis))
    application.add_handler(CommandHandler('delthereis', delthereis))
    application.add_handler(CommandHandler('delwhereis', delthereis))
    application.add_handler(CommandHandler('graph', locationdistance.locationinfo))
    application.add_handler(CommandHandler('distfrom', distfrom))
    application.add_handler(CommandHandler('pathfrom', pathfrom))
    application.add_handler(ConversationHandler(
        entry_points=[CommandHandler("delevent", delevent)],
        states={
            0: [CallbackQueryHandler(do_delete_event)],
        },
        fallbacks=[],
    ), group=2)
    application.add_handler(ConversationHandler(
        entry_points=[CommandHandler("selectevent", selectevent)],
        states={
            0: [CallbackQueryHandler(do_selectevent)],
        },
        fallbacks=[],
    ), group=4)
    
    application.add_handler(CommandHandler('ru', ru))
    application.add_handler(CommandHandler('ipa', ipa))
    application.add_handler(CommandHandler('iparu', iparu))
    application.add_handler(CommandHandler('pron', pron))
    application.add_handler(CommandHandler('dict', dict_))
    application.add_handler(CommandHandler('wikt', wikt))
    application.add_handler(CommandHandler('larousse', larousse))
    application.add_handler(CommandHandler('eur', eur))
    application.add_handler(CommandHandler('brl', brl))
    application.add_handler(CommandHandler('rub', rub))
    application.add_handler(CommandHandler('convertmoney', convertmoney))
    application.add_handler(CommandHandler('mytimezone', mytimezone))
    application.add_handler(CommandHandler('timezonealias', timezonealias))
    application.add_handler(CommandHandler('listallsettings', listallsettings))
    application.add_handler(CommandHandler('mysettings', mysettings))
    application.add_handler(CommandHandler('chatsettings', chatsettings))
    application.add_handler(CommandHandler('flashcard', partial(add_flashcard, scope='general')))
    application.add_handler(CommandHandler('myflashcard', partial(add_flashcard, scope='personal')))
    application.add_handler(CommandHandler('switchpageflashcard', switchpageflashcard))
    application.add_handler(CommandHandler('listflashcards', listflashcards))
    application.add_handler(CommandHandler('listpageflashcards', listpageflashcards))
    application.add_handler(CommandHandler('exportflashcards', exportflashcards))
    application.add_handler(CommandHandler('sharemoney', sharemoney))
    application.add_handler(ConversationHandler(
        entry_points=[CommandHandler('practiceflashcards', practiceflashcards)],
        states={
            0: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, guessing_word),
                CommandHandler('cancel', guessing_word),
                CommandHandler('stop', guessing_word),
                CommandHandler('finish', guessing_word),
            ]
        },
        fallbacks=[]
    ), group=1)
    application.add_handler(CommandHandler('help', help))
    application.add_handler(CommandHandler('uniline', uniline))
    application.add_handler(CommandHandler('nuniline', nuniline))
    application.add_handler(CommandHandler('timein', timein))
    application.add_handler(CommandHandler('timeuntil', timeuntil))
    application.add_handler(CommandHandler('timesince', timesince))
    #application.add_handler(CommandHandler('sleep', sleep_))
    application.add_handler(CommandHandler('listdebts', listdebts))
    application.add_handler(CommandHandler('detaildebts', detaildebts))

    application.add_handler(CommandHandler('createlist', listsmodule.createlist()))
    application.add_handler(CommandHandler('addtolist', listsmodule.addtolist()))
    application.add_handler(CommandHandler('appendtolist', listsmodule.addtolist()))
    application.add_handler(CommandHandler('removefromlist', listsmodule.removefromlist()))
    application.add_handler(CommandHandler('delfromlist', listsmodule.removefromlist()))
    application.add_handler(CommandHandler('deletefromlist', listsmodule.removefromlist()))
    application.add_handler(CommandHandler('printlist', listsmodule.printlist()))
    application.add_handler(CommandHandler('dirlist', listsmodule.dirlist()))
    application.add_handler(CommandHandler('dellist', listsmodule.dellist()))
    application.add_handler(CommandHandler('delist', listsmodule.dellist()))
    application.add_handler(CommandHandler('enable', partial(enable_disable_command, direction='enable')))
    application.add_handler(CommandHandler('disable', partial(implicit_setting_command, type='disable')))
    application.add_handler(CommandHandler('only', partial(implicit_setting_command, type='only')))
    application.add_handler(CommandHandler('known', partial(implicit_setting_command, type='known')))

    application.add_handler(CommandHandler('menu', menu))
    application.add_handler(CallbackQueryHandler(menu_button_handler, pattern="^cmd:"))

    for i in irange(1, 7):
        application.add_handler(CommandHandler(DatetimeText.days_english[i-1], partial(day_of_week_command, n=i), filters=DayOfWeekFilter()))

    application.add_error_handler(general_error_callback)
    
    application.run_polling()

